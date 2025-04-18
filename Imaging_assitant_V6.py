import logging
import traceback
import sys
import svgwrite
import tempfile
from tempfile import NamedTemporaryFile
import base64
from PIL import ImageGrab, Image, ImageQt  # Import Pillow's ImageGrab for clipboard access
from io import BytesIO
import io
from PyQt5.QtWidgets import (
    QDesktopWidget, QSpacerItem, QDialogButtonBox,QTableWidget, QTableWidgetItem, QScrollArea, QInputDialog, QShortcut, QFrame, QApplication, QSizePolicy, QMainWindow, QApplication, QTabWidget, QLabel, QPushButton, QVBoxLayout, QTextEdit, QHBoxLayout, QCheckBox, QGroupBox, QGridLayout, QWidget, QFileDialog, QSlider, QComboBox, QColorDialog, QMessageBox, QLineEdit, QFontComboBox, QSpinBox
)
from PyQt5.QtGui import QPixmap, QKeySequence, QImage, QPolygonF,QPainter, QColor, QFont, QKeySequence, QClipboard, QPen, QTransform,QFontMetrics,QDesktopServices
from PyQt5.QtCore import Qt, QBuffer, QPoint,QPointF, QRectF,QUrl
import json
import os
import numpy as np
import matplotlib.pyplot as plt
import platform
import openpyxl
from openpyxl.styles import Font
# import ctypes

from PyQt5.QtWidgets import QDialog, QVBoxLayout, QLabel, QPushButton, QSlider,QMenuBar, QMenu, QAction
from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg as FigureCanvas
from matplotlib.gridspec import GridSpec
from skimage.restoration import rolling_ball 
from scipy.signal import find_peaks
from scipy.ndimage import gaussian_filter1d
from scipy.interpolate import interp1d # Needed for interpolation
import cv2

# --- Style Sheet Definition ---
STYLE_SHEET = """
QMainWindow {
    background-color: #f0f0f0; /* Light gray background */
}

QTabWidget::pane { /* The tab widget frame */
    border-top: 1px solid #C2C7CB;
    padding: 10px;
    background-color: #ffffff; /* White background for tab content */
}

QTabBar::tab {
    background: qlineargradient(x1: 0, y1: 0, x2: 0, y2: 1,
                                stop: 0 #E1E1E1, stop: 0.4 #DDDDDD,
                                stop: 0.5 #D8D8D8, stop: 1.0 #D3D3D3);
    border: 1px solid #C4C4C3;
    border-bottom-color: #C2C7CB; /* same as the pane border */
    border-top-left-radius: 4px;
    border-top-right-radius: 4px;
    min-width: 8ex;
    padding: 5px 10px;
    margin-right: 2px; /* space between tabs */
}

QTabBar::tab:selected, QTabBar::tab:hover {
    background: qlineargradient(x1: 0, y1: 0, x2: 0, y2: 1,
                                stop: 0 #fafafa, stop: 0.4 #f4f4f4,
                                stop: 0.5 #e7e7e7, stop: 1.0 #fafafa);
}

QTabBar::tab:selected {
    border-color: #9B9B9B;
    border-bottom-color: #ffffff; /* same as pane background */
    margin-left: -2px; /* make selected tab look connected */
    margin-right: -2px;
}

QTabBar::tab:!selected {
    margin-top: 2px; /* make non-selected tabs look smaller */
}

QPushButton {
    background-color: #e0e0e0;
    border: 1px solid #c0c0c0;
    padding: 5px 10px;
    border-radius: 4px;
    min-height: 20px; /* Ensure minimum height */
}

QPushButton:hover {
    background-color: #d0d0d0;
    border: 1px solid #b0b0b0;
}

QPushButton:pressed {
    background-color: #c0c0c0;
}

QPushButton:disabled {
    background-color: #f5f5f5;
    color: #a0a0a0;
    border: 1px solid #d5d5d5;
}

QGroupBox {
    background-color: #fafafa; /* Slightly off-white */
    border: 1px solid #d0d0d0;
    border-radius: 5px;
    margin-top: 1ex; /* spacing above the title */
    padding: 10px; /* internal padding */
}

QGroupBox::title {
    subcontrol-origin: margin;
    subcontrol-position: top left; /* position at the top left */
    padding: 0 3px;
    left: 10px; /* position title slightly indented */
    color: #333;
    font-weight: bold;
}

QLabel {
    color: #333; /* Darker text for labels */
    padding-bottom: 2px; /* Small spacing below labels */
}

QLineEdit, QTextEdit, QSpinBox, QComboBox, QFontComboBox {
    border: 1px solid #c0c0c0;
    border-radius: 3px;
    padding: 3px 5px;
    background-color: white;
    min-height: 20px;
}

QLineEdit:focus, QTextEdit:focus, QSpinBox:focus, QComboBox:focus, QFontComboBox:focus {
    border: 1px solid #88aaff; /* Highlight focus */
}

QSlider::groove:horizontal {
    border: 1px solid #bbb;
    background: white;
    height: 8px; /* Slider groove height */
    border-radius: 4px;
}

QSlider::handle:horizontal {
    background: qlineargradient(x1:0, y1:0, x2:1, y2:1, stop:0 #eee, stop:1 #ccc);
    border: 1px solid #777;
    width: 13px; /* Handle width */
    margin: -2px 0; /* handle is placed vertically centered */
    border-radius: 4px;
}

QSlider::add-page:horizontal {
    background: #d0d0d0; /* Color for the part after the handle */
    border: 1px solid #bbb;
    border-radius: 4px;
}

QSlider::sub-page:horizontal {
    background: #88aaff; /* Color for the part before the handle */
    border: 1px solid #bbb;
    border-radius: 4px;
}

QTableWidget {
    border: 1px solid #c0c0c0;
    gridline-color: #d0d0d0; /* Lighter grid lines */
    background-color: white;
}

QHeaderView::section {
    background-color: #e8e8e8; /* Header background */
    padding: 4px;
    border: 1px solid #c0c0c0;
    font-weight: bold;
}

QTableWidgetItem {
    padding: 3px;
}

QScrollArea {
    border: none; /* Remove border from scroll area itself */
}
/* Make the LiveViewLabel border slightly softer */
#LiveViewLabel {
    border: 1px solid #c0c0c0;
}
"""

# --- End Style Sheet Definition ---
# Configure logging to write errors to a log file
logging.basicConfig(
    filename="error_log.txt",
    level=logging.ERROR,
    format="%(asctime)s - %(levelname)s - %(message)s"
)

def log_exception(exc_type, exc_value, exc_traceback):
    """Log uncaught exceptions to the error log."""
    if issubclass(exc_type, KeyboardInterrupt):
        sys.__excepthook__(exc_type, exc_value, exc_traceback)
        return
    # Log the exception
    logging.error("Uncaught exception", exc_info=(exc_type, exc_value, exc_traceback))

    # Display a QMessageBox with the error details
    error_message = f"An unexpected error occurred:\n\n{exc_type.__name__}: {exc_value}"
    QMessageBox.critical(
        None,  # No parent window
        "Unexpected Error",  # Title of the message box
        error_message,  # Error message to display
        QMessageBox.Ok  # Button to close the dialog
    )
    

# # Set the custom exception handler
sys.excepthook = log_exception

class TableWindow(QDialog):
    # Add 'calculated_quantities' parameter
    def __init__(self, peak_areas, standard_dictionary, standard, calculated_quantities, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Table Export")
        self.setGeometry(100, 100, 600, 400)
        self.calculated_quantities = calculated_quantities # <-- Store passed quantities

        self.scroll_area = QScrollArea(self)
        self.scroll_area.setWidgetResizable(True)

        self.table = QTableWidget()
        self.table.setColumnCount(4)
        self.table.setHorizontalHeaderLabels(["Band", "Peak Area", "Percentage (%)", "Quantity (Unit)"])
        self.table.setSelectionBehavior(QTableWidget.SelectRows)
        self.table.setSelectionMode(QTableWidget.ContiguousSelection)

        # Populate the table using the passed data
        self.populate_table(peak_areas, standard_dictionary, standard)

        self.scroll_area.setWidget(self.table)
        self.export_button = QPushButton("Export to Excel")
        self.export_button.clicked.connect(self.export_to_excel)

        layout = QVBoxLayout()
        layout.addWidget(self.scroll_area)
        layout.addWidget(self.export_button)
        self.setLayout(layout)

    def populate_table(self, peak_areas, standard_dictionary, standard):
        total_area = sum(peak_areas)
        self.table.setRowCount(len(peak_areas))

        for row, area in enumerate(peak_areas):
            band_label = f"Band {row + 1}"
            self.table.setItem(row, 0, QTableWidgetItem(band_label))

            peak_area_rounded = round(area, 3)
            self.table.setItem(row, 1, QTableWidgetItem(str(peak_area_rounded)))

            if total_area != 0:
                percentage = (area / total_area) * 100
                percentage_rounded = round(percentage, 2)
            else:
                percentage_rounded = 0.0
            self.table.setItem(row, 2, QTableWidgetItem(f"{percentage_rounded:.2f}%"))

            # --- Use pre-calculated quantities if available ---
            quantity_str = "" # Default empty string
            if standard and row < len(self.calculated_quantities): # Check if standard mode and index is valid
                quantity_str = f"{self.calculated_quantities[row]:.2f}"
            # Always set the item, even if empty
            self.table.setItem(row, 3, QTableWidgetItem(quantity_str))
            # --- End quantity handling ---


        self.table.setEditTriggers(QTableWidget.NoEditTriggers)
        self.table.setContextMenuPolicy(Qt.CustomContextMenu)
        self.table.customContextMenuRequested.connect(self.copy_table_data)

    # (copy_table_data and export_to_excel remain the same)
    def copy_table_data(self):
        """Copy selected table data to the clipboard."""
        selected_items = self.table.selectedItems()
        if not selected_items:
            return

        # Get the selected rows and columns
        rows = set(item.row() for item in selected_items)
        cols = set(item.column() for item in selected_items)

        # Prepare the data for copying
        data = []
        for row in sorted(rows):
            row_data = []
            for col in sorted(cols):
                item = self.table.item(row, col)
                row_data.append(item.text() if item else "")
            data.append("\t".join(row_data))

        # Copy the data to the clipboard
        clipboard = QApplication.clipboard()
        clipboard.setText("\n".join(data))

    def export_to_excel(self):
        """Export the table data to an Excel file."""
        # Prompt the user to select a save location
        options = QFileDialog.Options()
        file_path, _ = QFileDialog.getSaveFileName(
            self, "Save Excel File", "", "Excel Files (*.xlsx)", options=options
        )
        if not file_path:
            return

        # Create a new Excel workbook and worksheet
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = "Peak Areas"

        # Write the table headers to the Excel sheet
        headers = [self.table.horizontalHeaderItem(col).text() for col in range(self.table.columnCount())]
        for col, header in enumerate(headers, start=1):
            worksheet.cell(row=1, column=col, value=header)
            worksheet.cell(row=1, column=col).font = Font(bold=True)  # Make headers bold

        # Write the table data to the Excel sheet
        for row in range(self.table.rowCount()):
            for col in range(self.table.columnCount()):
                item = self.table.item(row, col)
                worksheet.cell(row=row + 2, column=col + 1, value=item.text() if item else "")

        # Save the Excel file
        try:
            workbook.save(file_path)
            QMessageBox.information(self, "Success", f"Table data exported to {file_path}")
        except Exception as e:
             QMessageBox.critical(self, "Export Error", f"Could not save Excel file:\n{e}")
                

class PeakAreaDialog(QDialog):
    """
    Interactive dialog to adjust peak regions and calculate peak areas.
    Handles 8-bit ('L') and 16-bit ('I;16', 'I') grayscale PIL Image input.
    Area calculations use original data range (inverted).
    Peak detection uses a 0-255 scaled profile (inverted).
    """

    def __init__(self, cropped_data, current_settings, persist_checked, parent=None):
        """
        Initializes the dialog. Args updated.
        """
        super().__init__(parent)
        self.setWindowTitle("Adjust Peak Regions and Calculate Areas")
        self.setGeometry(100, 100, 1100, 850) # Keep original size for now

        # --- Validate and Store Input Image ---
        # (Keep image validation and processing logic as is)
        if not isinstance(cropped_data, Image.Image):
             raise TypeError("Input 'cropped_data' must be a PIL Image object")
        self.cropped_image_for_display = cropped_data # Keep original PIL for display

        self.original_max_value = 255.0 # Default assumption
        pil_mode = cropped_data.mode
        print(f"PeakAreaDialog: Initializing with PIL mode = {pil_mode}")

        if pil_mode.startswith('I;16') or pil_mode == 'I' or pil_mode == 'I;16B' or pil_mode == 'I;16L':
            self.intensity_array_original_range = np.array(cropped_data, dtype=np.float64)
            self.original_max_value = 65535.0
        elif pil_mode == 'L':
            self.intensity_array_original_range = np.array(cropped_data, dtype=np.float64)
            self.original_max_value = 255.0
        elif pil_mode == 'F':
             self.intensity_array_original_range = np.array(cropped_data, dtype=np.float64)
             max_in_float = np.max(self.intensity_array_original_range) if np.any(self.intensity_array_original_range) else 1.0
             self.original_max_value = 1.0 if max_in_float <= 1.0 and max_in_float > 0 else max_in_float
             print(f"PeakAreaDialog: Float max value detected: {max_in_float}, using original_max_value={self.original_max_value}")
             try:
                 scaled_for_display = np.clip(self.intensity_array_original_range * 255.0 / self.original_max_value, 0, 255).astype(np.uint8)
                 self.cropped_image_for_display = Image.fromarray(scaled_for_display, mode='L')
             except: pass
        else:
            print(f"PeakAreaDialog: Warning - Mode '{pil_mode}' converting to 8-bit 'L'.")
            try:
                gray_img = cropped_data.convert("L")
                self.intensity_array_original_range = np.array(gray_img, dtype=np.float64)
                self.original_max_value = 255.0
                self.cropped_image_for_display = gray_img
            except Exception as e: raise TypeError(f"Could not convert '{pil_mode}' to 'L': {e}")

        if self.intensity_array_original_range.ndim != 2:
             raise ValueError(f"Intensity array must be 2D, shape {self.intensity_array_original_range.shape}")

        self.profile_original_inverted = None
        self.profile = None # Scaled, inverted, SMOOTHED profile for detection
        self.background = None

        # --- Settings and State ---
        self.rolling_ball_radius = current_settings.get('rolling_ball_radius', 50)
        # ** NEW: Add smoothing sigma setting **
        self.smoothing_sigma = current_settings.get('smoothing_sigma', 2.0)
        # ... (rest of settings loading remains the same) ...
        self.peak_height_factor = current_settings.get('peak_height_factor', 0.1)
        self.peak_distance = current_settings.get('peak_distance', 30)
        self.peak_prominence_factor = current_settings.get('peak_prominence_factor', 0.02)
        self.peak_spread_pixels = current_settings.get('peak_spread_pixels', 10)
        self.band_estimation_method = current_settings.get('band_estimation_method', "Mean")
        self.area_subtraction_method = current_settings.get('area_subtraction_method', "Valley-to-Valley")
        self.peaks = np.array([])
        self.initial_peak_regions = []
        self.peak_regions = []
        self.peak_areas_rolling_ball = []
        self.peak_areas_straight_line = []
        self.peak_areas_valley = []
        self.peak_sliders = []
        self._final_settings = {}
        self._persist_enabled_on_exit = persist_checked

        # Build UI & Initial Setup
        self._setup_ui(persist_checked)
        self.regenerate_profile_and_detect() # Initial calculation

    def _setup_ui(self, persist_checked_initial):
        """Creates and arranges the UI elements."""
        # --- Main Layout ---
        main_layout = QVBoxLayout(self)
        main_layout.setSpacing(15)

        # --- Matplotlib Plot Canvas ---
        self.fig = plt.figure(figsize=(10, 5))
        self.canvas = FigureCanvas(self.fig)
        self.canvas.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        main_layout.addWidget(self.canvas, stretch=3)

        # --- Controls Layout (Horizontal Box for side-by-side groups) ---
        controls_hbox = QHBoxLayout()
        controls_hbox.setSpacing(15)

        # --- Left Controls Column (Global & Detection) ---
        left_controls_vbox = QVBoxLayout()

        # Group 1: Global Settings
        global_settings_group = QGroupBox("Global Settings")
        global_settings_layout = QGridLayout(global_settings_group)
        global_settings_layout.setSpacing(8)

        # Band Estimation Method Dropdown (Row 0)
        self.band_estimation_combobox = QComboBox()
        self.band_estimation_combobox.addItems(["Mean", "Percentile:5%", "Percentile:10%", "Percentile:15%", "Percentile:30%"])
        self.band_estimation_combobox.setCurrentText(self.band_estimation_method)
        self.band_estimation_combobox.currentIndexChanged.connect(self.regenerate_profile_and_detect)
        global_settings_layout.addWidget(QLabel("Band Profile:"), 0, 0)
        global_settings_layout.addWidget(self.band_estimation_combobox, 0, 1, 1, 2) # Span 2

        # Area Subtraction Method Dropdown (Row 1)
        self.method_combobox = QComboBox()
        self.method_combobox.addItems(["Valley-to-Valley", "Rolling Ball", "Straight Line"])
        self.method_combobox.setCurrentText(self.area_subtraction_method)
        self.method_combobox.currentIndexChanged.connect(self.update_plot) # Only needs plot update
        global_settings_layout.addWidget(QLabel("Area Method:"), 1, 0)
        global_settings_layout.addWidget(self.method_combobox, 1, 1, 1, 2) # Span 2

        # Rolling Ball Radius Slider (Row 2) - ** MODIFIED LAYOUT **
        global_settings_layout.addWidget(QLabel("Rolling Ball Radius:"), 2, 0) # Static Label
        self.rolling_ball_slider = QSlider(Qt.Horizontal)
        self.rolling_ball_slider.setRange(1, 500)
        self.rolling_ball_slider.setValue(int(self.rolling_ball_radius))
        self.rolling_ball_slider.valueChanged.connect(self.update_plot) # Update plot when radius changes
        # Separate Label for the value, prevents layout jumps
        self.rolling_ball_value_label = QLabel(f"({int(self.rolling_ball_radius)})")
        # Ensure value label has enough space initially
        fm = QFontMetrics(self.rolling_ball_value_label.font())
        self.rolling_ball_value_label.setMinimumWidth(fm.horizontalAdvance("(500) ")) # Width for max value
        self.rolling_ball_value_label.setAlignment(Qt.AlignLeft | Qt.AlignVCenter) # Align left
        self.rolling_ball_slider.valueChanged.connect(lambda val, lbl=self.rolling_ball_value_label: lbl.setText(f"({val})")) # Update value label
        global_settings_layout.addWidget(self.rolling_ball_slider, 2, 1) # Slider in middle column
        global_settings_layout.addWidget(self.rolling_ball_value_label, 2, 2) # Value label in last column


        left_controls_vbox.addWidget(global_settings_group)

        # Group 2: Peak Detection Parameters
        peak_detect_group = QGroupBox("Peak Detection Parameters")
        peak_detect_layout = QGridLayout(peak_detect_group)
        peak_detect_layout.setSpacing(8)

        # Manual Peak Number Input & Update Button (Row 0)
        self.peak_number_label = QLabel("Detected Peaks:")
        self.peak_number_input = QLineEdit()
        self.peak_number_input.setPlaceholderText("#")
        self.peak_number_input.setMaximumWidth(60)
        self.update_peak_number_button = QPushButton("Set")
        self.update_peak_number_button.setToolTip("Manually override the number of peaks detected.")
        self.update_peak_number_button.clicked.connect(self.manual_peak_number_update)
        peak_detect_layout.addWidget(self.peak_number_label, 0, 0)
        peak_detect_layout.addWidget(self.peak_number_input, 0, 1)
        peak_detect_layout.addWidget(self.update_peak_number_button, 0, 2)

        # ** NEW: Smoothing Sigma Slider (Row 1) **
        self.smoothing_label = QLabel(f"Smoothing Sigma ({self.smoothing_sigma:.1f})")
        self.smoothing_slider = QSlider(Qt.Horizontal)
        # Range 0-100 represents sigma 0.0 to 10.0
        self.smoothing_slider.setRange(0, 100)
        self.smoothing_slider.setValue(int(self.smoothing_sigma * 10)) # Initial value
        self.smoothing_slider.valueChanged.connect(lambda val, lbl=self.smoothing_label: lbl.setText(f"Smoothing Sigma ({val/10.0:.1f})")) # Update label text
        # Changing sigma requires re-smoothing and re-detecting peaks
        self.smoothing_slider.valueChanged.connect(self.regenerate_profile_and_detect) # Connect to regenerate
        peak_detect_layout.addWidget(self.smoothing_label, 1, 0)
        peak_detect_layout.addWidget(self.smoothing_slider, 1, 1, 1, 2) # Span slider


        # Peak Prominence Slider (Row 2) - Shifted down one row
        self.peak_prominence_slider_label = QLabel(f"Min Prominence ({self.peak_prominence_factor:.2f}) (rel 0-255)")
        self.peak_prominence_slider = QSlider(Qt.Horizontal)
        self.peak_prominence_slider.setRange(0, 100) # 0.0 to 1.0 factor
        self.peak_prominence_slider.setValue(int(self.peak_prominence_factor * 100))
        # Changing prominence only requires re-detection, not full regeneration
        self.peak_prominence_slider.valueChanged.connect(self.detect_peaks)
        self.peak_prominence_slider.valueChanged.connect(lambda val, lbl=self.peak_prominence_slider_label: lbl.setText(f"Min Prominence ({val/100.0:.2f}) (rel 0-255)"))
        peak_detect_layout.addWidget(self.peak_prominence_slider_label, 2, 0)
        peak_detect_layout.addWidget(self.peak_prominence_slider, 2, 1, 1, 2)

        # Peak Height Slider (Row 3) - Shifted down one row
        self.peak_height_slider_label = QLabel(f"Min Height ({self.peak_height_factor:.2f}) % (of 0-255 range)")
        self.peak_height_slider = QSlider(Qt.Horizontal)
        self.peak_height_slider.setRange(0, 100)
        self.peak_height_slider.setValue(int(self.peak_height_factor * 100))
        self.peak_height_slider.valueChanged.connect(self.detect_peaks)
        self.peak_height_slider.valueChanged.connect(lambda val, lbl=self.peak_height_slider_label: lbl.setText(f"Min Height ({val/100.0:.2f}) % (of 0-255 range)"))
        peak_detect_layout.addWidget(self.peak_height_slider_label, 3, 0)
        peak_detect_layout.addWidget(self.peak_height_slider, 3, 1, 1, 2)

        # Peak Distance Slider (Row 4) - Shifted down one row
        self.peak_distance_slider_label = QLabel(f"Min Distance ({self.peak_distance}) px")
        self.peak_distance_slider = QSlider(Qt.Horizontal)
        self.peak_distance_slider.setRange(1, 200)
        self.peak_distance_slider.setValue(self.peak_distance)
        self.peak_distance_slider.valueChanged.connect(self.detect_peaks)
        self.peak_distance_slider.valueChanged.connect(lambda val, lbl=self.peak_distance_slider_label: lbl.setText(f"Min Distance ({val}) px"))
        peak_detect_layout.addWidget(self.peak_distance_slider_label, 4, 0)
        peak_detect_layout.addWidget(self.peak_distance_slider, 4, 1, 1, 2)


        left_controls_vbox.addWidget(peak_detect_group)
        left_controls_vbox.addStretch(1)

        controls_hbox.addLayout(left_controls_vbox, stretch=1)

        # --- Right Controls Column (Peak Region Adjustments - No changes needed here) ---
        right_controls_vbox = QVBoxLayout()
        peak_spread_group = QGroupBox("Peak Region Adjustments")
        peak_spread_layout = QGridLayout(peak_spread_group)
        peak_spread_layout.setSpacing(8)
        self.peak_spread_label = QLabel(f"Peak Spread (+/- {self.peak_spread_pixels} px)")
        self.peak_spread_slider = QSlider(Qt.Horizontal)
        self.peak_spread_slider.setRange(0, 100)
        self.peak_spread_slider.setValue(self.peak_spread_pixels)
        self.peak_spread_slider.setToolTip(
            "Adjusts the width of all detected peak regions simultaneously.\n"
            "Regions expand/contract around the initial detected peak center."
        )
        self.peak_spread_slider.valueChanged.connect(self.apply_peak_spread)
        self.peak_spread_slider.valueChanged.connect(
            lambda value, lbl=self.peak_spread_label: lbl.setText(f"Peak Spread (+/- {value} px)")
        )
        peak_spread_layout.addWidget(self.peak_spread_label, 0, 0)
        peak_spread_layout.addWidget(self.peak_spread_slider, 0, 1)
        right_controls_vbox.addWidget(peak_spread_group)

        scroll_area = QScrollArea()
        scroll_area.setWidgetResizable(True)
        scroll_area.setMinimumHeight(250)
        scroll_area.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        self.container = QWidget()
        self.peak_sliders_layout = QVBoxLayout(self.container)
        self.peak_sliders_layout.setSpacing(10)
        scroll_area.setWidget(self.container)
        right_controls_vbox.addWidget(scroll_area, stretch=1)
        controls_hbox.addLayout(right_controls_vbox, stretch=2)
        main_layout.addLayout(controls_hbox)

        # --- Bottom Button Layout (No changes needed here) ---
        bottom_button_layout = QHBoxLayout()
        self.persist_settings_checkbox = QCheckBox("Persist Settings")
        self.persist_settings_checkbox.setChecked(persist_checked_initial)
        self.persist_settings_checkbox.setToolTip("Save current detection parameters for the next time this dialog opens during this session.")
        bottom_button_layout.addWidget(self.persist_settings_checkbox)
        bottom_button_layout.addStretch(1)
        self.ok_button = QPushButton("OK")
        self.ok_button.setMinimumWidth(100)
        self.ok_button.setDefault(True)
        self.ok_button.clicked.connect(self.accept_and_close)
        bottom_button_layout.addWidget(self.ok_button)
        main_layout.addLayout(bottom_button_layout)


    # --- Methods for Parent Interaction (No changes needed here) ---
    def accept_and_close(self):
        self._final_settings = {
            'rolling_ball_radius': self.rolling_ball_slider.value(),
            'peak_height_factor': self.peak_height_slider.value() / 100.0,
            'peak_distance': self.peak_distance_slider.value(),
            'peak_prominence_factor': self.peak_prominence_slider.value() / 100.0,
            'peak_spread_pixels': self.peak_spread_slider.value(),
            'band_estimation_method': self.band_estimation_combobox.currentText(),
            'area_subtraction_method': self.method_combobox.currentText(),
            # ** NEW: Save smoothing sigma **
            'smoothing_sigma': self.smoothing_slider.value() / 10.0,
        }
        self._persist_enabled_on_exit = self.persist_settings_checkbox.isChecked()
        self.accept()

    def get_current_settings(self): return self._final_settings
    def should_persist_settings(self): return self._persist_enabled_on_exit


    # --- Core Logic Methods ---

    def regenerate_profile_and_detect(self):
        """
        Calculates the raw inverted profile, applies smoothing based on the slider,
        stores this as the main profile for plotting and calculation,
        then scales a copy for peak detection.
        """
        self.band_estimation_method = self.band_estimation_combobox.currentText()
        self.area_subtraction_method = self.method_combobox.currentText()
        if hasattr(self, 'smoothing_slider'):
             self.smoothing_sigma = self.smoothing_slider.value() / 10.0
        else:
            print("Warning: Smoothing slider not ready.")

        # --- Calculate profile from ORIGINAL intensity data ---
        # (Calculation logic remains the same)
        profile_temp = None
        # ... (Mean/Percentile calculation as before) ...
        if self.band_estimation_method == "Mean":
            profile_temp = np.mean(self.intensity_array_original_range, axis=1)
        elif self.band_estimation_method.startswith("Percentile"):
            try:
                percent = int(self.band_estimation_method.split(":")[1].replace('%', ''))
                profile_temp = np.percentile(self.intensity_array_original_range, max(0, min(100, percent)), axis=1)
            except: profile_temp = np.percentile(self.intensity_array_original_range, 5, axis=1); print("Warning: Defaulting to 5th percentile.")
        else: profile_temp = np.mean(self.intensity_array_original_range, axis=1)

        if not np.all(np.isfinite(profile_temp)):
            print("Warning: Original profile NaN/Inf. Setting to zero.")
            profile_temp = np.zeros(self.intensity_array_original_range.shape[0])


        # --- Create INVERTED Original Profile ---
        profile_original_inv_raw = self.original_max_value - profile_temp.astype(np.float64)
        min_inverted_raw = np.min(profile_original_inv_raw)
        profile_original_inv_raw -= min_inverted_raw # Shift baseline to zero

        # --- Apply Smoothing to the INVERTED ORIGINAL profile ---
        # This smoothed version becomes the primary profile for plotting and calculations
        self.profile_original_inverted = profile_original_inv_raw # Start with raw
        try:
            current_sigma = self.smoothing_sigma
            if current_sigma > 0.1 and len(self.profile_original_inverted) > 6:
                # Smooth the full-range inverted profile
                self.profile_original_inverted = gaussian_filter1d(
                    self.profile_original_inverted, sigma=current_sigma
                )
                print(f"  Applied Gaussian smoothing (sigma={current_sigma:.1f}) to main profile.")
            elif current_sigma <= 0.1:
                 print("  Skipping smoothing on main profile (sigma <= 0.1)")
            else: print("Warning: Main profile too short for smoothing.")
        except Exception as smooth_err:
            print(f"Error smoothing main profile: {smooth_err}")
            # self.profile_original_inverted remains the raw version on error


        # --- Create the SCALED (0-255) version FOR PEAK DETECTION ONLY ---
        # Scale the *already smoothed* self.profile_original_inverted
        prof_min_inv, prof_max_inv = np.min(self.profile_original_inverted), np.max(self.profile_original_inverted)
        if prof_max_inv > prof_min_inv:
            self.profile = (self.profile_original_inverted - prof_min_inv) / (prof_max_inv - prof_min_inv) * 255.0
        else:
            self.profile = np.zeros_like(self.profile_original_inverted) # Handle flat profile

        # No need to smooth self.profile again, it was derived from the smoothed original

        # Detect peaks using the scaled profile
        self.detect_peaks()


    def detect_peaks(self):
        """Detect peaks using the 0-255 SCALED, INVERTED, and **potentially smoothed** self.profile."""
        # self.profile should already be smoothed (or not) by regenerate_profile_and_detect

        if self.profile is None or len(self.profile) == 0:
            print("Profile (for detection) not generated yet.")
            # (Rest of handling for no profile remains the same)
            self.peaks, self.initial_peak_regions, self.peak_regions = np.array([]), [], []
            self.peak_number_input.setText("0"); self.update_sliders(); self.update_plot()
            return

        # --- Update parameters from sliders (excluding smoothing, already done) ---
        # (Parameter update logic remains the same)
        self.peak_height_factor = self.peak_height_slider.value() / 100.0
        self.peak_distance = self.peak_distance_slider.value()
        self.peak_prominence_factor = self.peak_prominence_slider.value() / 100.0
        self.peak_spread_pixels = self.peak_spread_slider.value()
        self.rolling_ball_radius = self.rolling_ball_slider.value()

        # --- Update UI Labels ---
        # (Label update logic remains the same)
        # ** NEW: Update smoothing label **
        if hasattr(self, 'smoothing_label'): # Check if UI element exists
            self.smoothing_label.setText(f"Smoothing Sigma ({self.smoothing_sigma:.1f})")
        self.peak_height_slider_label.setText(f"Min Height ({self.peak_height_factor:.2f}) % (of 0-255 range)")
        self.peak_distance_slider_label.setText(f"Min Distance ({self.peak_distance}) px")
        self.peak_prominence_slider_label.setText(f"Min Prominence ({self.peak_prominence_factor:.2f}) (rel 0-255)")
        self.peak_spread_label.setText(f"Peak Spread (+/- {self.peak_spread_pixels} px)")
        # Rolling ball label is now updated automatically via its connection

        # --- Calculate thresholds ON THE SCALED 0-255 PROFILE ---
        # (Threshold calculation remains the same)
        profile_range = np.ptp(self.profile)
        if profile_range < 1e-6 : profile_range = 1.0
        min_val_profile = np.min(self.profile)
        min_height_abs = min_val_profile + profile_range * self.peak_height_factor
        min_prominence_abs = profile_range * self.peak_prominence_factor
        min_prominence_abs = max(1.0, min_prominence_abs) # Ensure minimum

        # --- Detect peaks using the potentially smoothed self.profile ---
        try:
            # (find_peaks call remains the same)
            peaks_indices, properties = find_peaks(
                self.profile, # Use the potentially smoothed profile
                height=min_height_abs,
                prominence=min_prominence_abs,
                distance=self.peak_distance,
                width=1, rel_height=0.5
            )
            print(f"  Found {len(peaks_indices)} peaks at indices: {peaks_indices}") # Debug

            # (Rest of setting initial_peak_regions remains the same)
            left_ips = properties.get('left_ips', [])
            right_ips = properties.get('right_ips', [])
            self.peaks = peaks_indices
            self.initial_peak_regions = []
            profile_len = len(self.profile_original_inverted) # Use original length for bounds

            if len(left_ips) == len(self.peaks) and len(right_ips) == len(self.peaks):
                 for i, peak_idx in enumerate(self.peaks):
                     start = int(np.floor(left_ips[i])); end = int(np.ceil(right_ips[i]))
                     start = max(0, start); end = min(profile_len - 1, end)
                     if start >= end: # Fallback
                         fb_width = max(1, self.peak_distance // 4); start = max(0, peak_idx - fb_width); end = min(profile_len - 1, peak_idx + fb_width)
                         if start >= end: end = min(profile_len - 1, start + 1)
                     self.initial_peak_regions.append((start, end))
            else: # Fallback
                 print("Warning: Width properties inconsistent, using distance fallback.")
                 for i, peak_idx in enumerate(self.peaks):
                     wd_est = self.peak_distance // 2; start = max(0, peak_idx - wd_est); end = min(profile_len - 1, peak_idx + wd_est)
                     if start >= end: start = max(0, peak_idx - 2); end = min(profile_len - 1, peak_idx + 2)
                     if start >= end: end = min(profile_len - 1, start + 1)
                     self.initial_peak_regions.append((start, end))

        except Exception as e:
            print(f"Error during peak detection: {e}")
            QMessageBox.warning(self, "Peak Detection Error", f"Peak detection error:\n{e}")
            self.peaks = np.array([]); self.initial_peak_regions = []

        # (Rest of the function remains the same)
        if not self.peak_number_input.hasFocus() or self.peak_number_input.text() == "":
             self.peak_number_input.setText(str(len(self.peaks)))
        self.apply_peak_spread(self.peak_spread_slider.value()) # Calls update_sliders & update_plot

    # --- apply_peak_spread (No changes needed from previous version) ---
    def apply_peak_spread(self, spread_value):
        """Applies the spread value to the initial peak regions."""
        self.peak_spread_pixels = spread_value
        self.peak_regions = []
        if self.profile_original_inverted is None or len(self.profile_original_inverted) == 0:
            self.update_sliders(); self.update_plot(); return
        profile_len = len(self.profile_original_inverted)
        num_initial = min(len(self.peaks), len(self.initial_peak_regions))
        if len(self.peaks) != len(self.initial_peak_regions): print(f"Warning: Peak/initial region mismatch.")
        for i in range(num_initial):
            peak_idx = self.peaks[i]; center = peak_idx
            new_start = max(0, int(center - self.peak_spread_pixels))
            new_end = min(profile_len - 1, int(center + self.peak_spread_pixels))
            if new_start > new_end: new_start = new_end
            self.peak_regions.append((new_start, new_end))
        if len(self.peak_regions) != len(self.peaks): self.peak_regions = self.peak_regions[:len(self.peaks)]
        self.update_sliders()
        self.update_plot()

    # --- manual_peak_number_update (No changes needed from previous version) ---
    def manual_peak_number_update(self):
        """Handles manual changes to the number of peaks."""
        if self.profile_original_inverted is None or len(self.profile_original_inverted) == 0:
            QMessageBox.warning(self, "Error", "Profile must be generated first."); return
        profile_len = len(self.profile_original_inverted)
        try:
            num_peaks_manual = int(self.peak_number_input.text())
            if num_peaks_manual < 0: raise ValueError("Negative number")
            current_num_peaks = len(self.peaks)
            if num_peaks_manual == current_num_peaks: return
            if num_peaks_manual == 0: self.peaks, self.initial_peak_regions, self.peak_regions = np.array([]), [], []
            elif num_peaks_manual < current_num_peaks: self.peaks, self.initial_peak_regions = self.peaks[:num_peaks_manual], self.initial_peak_regions[:num_peaks_manual]
            else: # Add dummies
                num_to_add = num_peaks_manual - current_num_peaks; profile_center = profile_len // 2
                peaks_list = self.peaks.tolist(); initial_regions_list = list(self.initial_peak_regions)
                for i in range(num_to_add):
                    new_peak_pos = max(0, min(profile_len - 1, profile_center + np.random.randint(-50, 50)))
                    peaks_list.append(new_peak_pos)
                    pl_width = 5; initial_start = max(0, new_peak_pos - pl_width); initial_end = min(profile_len - 1, new_peak_pos + pl_width)
                    if initial_start >= initial_end: initial_end = min(profile_len-1, initial_start + 1)
                    initial_regions_list.append((initial_start, initial_end))
                if peaks_list:
                    min_len = min(len(peaks_list), len(initial_regions_list))
                    if len(peaks_list) != len(initial_regions_list): print("Warning: Peak/initial region mismatch manual add."); peaks_list, initial_regions_list = peaks_list[:min_len], initial_regions_list[:min_len]
                    combined = sorted(zip(peaks_list, initial_regions_list), key=lambda pair: pair[0])
                    if combined: sorted_peaks, sorted_initial_regions = zip(*combined); self.peaks, self.initial_peak_regions = np.array(sorted_peaks), list(sorted_initial_regions)
                    else: self.peaks, self.initial_peak_regions = np.array([]), []
                else: self.peaks, self.initial_peak_regions = np.array([]), []
            self.apply_peak_spread(self.peak_spread_slider.value())
        except ValueError: self.peak_number_input.setText(str(len(self.peaks))); QMessageBox.warning(self, "Input Error", "Please enter a valid integer.")
        except Exception as e: print(f"Error manual peak update: {e}"); QMessageBox.critical(self, "Error", f"Manual peak update error:\n{e}"); self.peak_number_input.setText(str(len(self.peaks)))

    # --- update_sliders (No changes needed from previous version) ---
    def update_sliders(self):
        """Update sliders based on self.peak_regions."""
        while self.peak_sliders_layout.count():
            item = self.peak_sliders_layout.takeAt(0); widget = item.widget()
            if widget: widget.deleteLater()
            elif not isinstance(item, QSpacerItem): del item
        self.peak_sliders.clear()
        if self.profile_original_inverted is None or len(self.profile_original_inverted) == 0: return
        profile_len = len(self.profile_original_inverted)
        num_items = len(self.peak_regions)
        if len(self.peaks) != num_items: print(f"Warning: Peak/region count mismatch sliders.")
        for i in range(num_items):
            try:
                start_val, end_val = int(self.peak_regions[i][0]), int(self.peak_regions[i][1])
                peak_index = int(self.peaks[i]) if i < len(self.peaks) else -1
            except Exception as e: print(f"Warning: Invalid data slider index {i}: {e}"); continue
            peak_group = QGroupBox(f"Peak {i + 1} (Idx: {peak_index if peak_index != -1 else 'N/A'})")
            peak_layout = QGridLayout(peak_group)
            start_slider = QSlider(Qt.Horizontal); start_slider.setRange(0, profile_len - 1); start_val_clamped = max(0, min(profile_len - 1, start_val)); start_slider.setValue(start_val_clamped)
            start_label = QLabel(f"Start: {start_val_clamped}"); start_slider.valueChanged.connect(lambda val, lbl=start_label, idx=i: self._update_region_from_slider(idx, 'start', val, lbl)); start_slider.valueChanged.connect(self.update_plot)
            peak_layout.addWidget(start_label, 0, 0); peak_layout.addWidget(start_slider, 0, 1)
            end_slider = QSlider(Qt.Horizontal); end_slider.setRange(0, profile_len - 1); end_val_clamped = max(start_val_clamped, min(profile_len - 1, end_val)); end_slider.setValue(end_val_clamped)
            end_label = QLabel(f"End: {end_val_clamped}"); end_slider.valueChanged.connect(lambda val, lbl=end_label, idx=i: self._update_region_from_slider(idx, 'end', val, lbl)); end_slider.valueChanged.connect(self.update_plot)
            peak_layout.addWidget(end_label, 1, 0); peak_layout.addWidget(end_slider, 1, 1)
            self.peak_sliders_layout.addWidget(peak_group); self.peak_sliders.append((start_slider, end_slider))
        if num_items > 0:
            last_item = self.peak_sliders_layout.itemAt(self.peak_sliders_layout.count() - 1)
            if not isinstance(last_item, QSpacerItem): self.peak_sliders_layout.addStretch(1)
        if hasattr(self, 'container') and self.container: self.container.update()

    # --- _update_region_from_slider (No changes needed from previous version) ---
    def _update_region_from_slider(self, index, boundary_type, value, label_widget):
        """Helper to update self.peak_regions."""
        if 0 <= index < len(self.peak_regions):
            current_start, current_end = self.peak_regions[index]
            start_slider_widget, end_slider_widget = self.peak_sliders[index]
            if boundary_type == 'start':
                new_start = min(value, current_end); self.peak_regions[index] = (new_start, current_end)
                label_widget.setText(f"Start: {new_start}")
                if start_slider_widget.value() != new_start: start_slider_widget.blockSignals(True); start_slider_widget.setValue(new_start); start_slider_widget.blockSignals(False)
            elif boundary_type == 'end':
                new_end = max(value, current_start); self.peak_regions[index] = (current_start, new_end)
                label_widget.setText(f"End: {new_end}")
                if end_slider_widget.value() != new_end: end_slider_widget.blockSignals(True); end_slider_widget.setValue(new_end); end_slider_widget.blockSignals(False)


    # --- update_plot (Modified for Correct Profiles) ---
    


    def update_plot(self):
        """
        Update plot using the **SMOOTHED** original inverted profile range for
        line display and area calculations.
        """
        if self.canvas is None: return

        # --- Profile for Plotting and Calculation is now the SMOOTHED one ---
        profile_to_plot_and_calc = self.profile_original_inverted

        if profile_to_plot_and_calc is None or len(profile_to_plot_and_calc) == 0 :
            try: self.fig.clf(); self.ax = self.fig.add_subplot(111); self.ax.text(0.5, 0.5, "No Profile Data", ha='center', va='center', transform=self.ax.transAxes); self.canvas.draw_idle()
            except Exception as e: print(f"Error clearing plot: {e}")
            return

        self.method = self.method_combobox.currentText()
        self.rolling_ball_radius = self.rolling_ball_slider.value()

        # --- Calculate Rolling Ball Background (on the SMOOTHED profile) ---
        try:
            profile_float = profile_to_plot_and_calc # Already float64
            safe_radius = max(1, min(self.rolling_ball_radius, len(profile_float) // 2 - 1))
            # Rolling ball background calculation needs to happen *after* smoothing
            if safe_radius != self.rolling_ball_radius: print(f"Adjusted rolling ball radius to {safe_radius}")
            if len(profile_float) > 1 :
                background_smoothed = rolling_ball(profile_float, radius=safe_radius)
                self.background = np.maximum(background_smoothed, 0) # Ensure background >= 0
            else: self.background = profile_float.copy()
        except ImportError: self.background = np.zeros_like(profile_to_plot_and_calc); print("Scikit-image needed")
        except Exception as e: print(f"Error rolling ball: {e}."); self.background = np.zeros_like(profile_to_plot_and_calc)

        # --- Setup Plot ---
        self.fig.clf(); gs = GridSpec(2, 1, height_ratios=[3, 1], figure=self.fig); self.ax = self.fig.add_subplot(gs[0]); ax_image = self.fig.add_subplot(gs[1], sharex=self.ax)

        # --- Plot Smoothed Profile ---
        self.ax.plot(profile_to_plot_and_calc, label=f"Profile (Smoothed σ={self.smoothing_sigma:.1f})", color="black", lw=1.2) # Label indicates smoothing

        # --- Plot Detected Peak Markers (on the smoothed profile line) ---
        if len(self.peaks) > 0:
             valid_peaks = self.peaks[(self.peaks >= 0) & (self.peaks < len(profile_to_plot_and_calc))]
             if len(valid_peaks) > 0:
                 # Get Y values from the SMOOTHED profile at the detected peak indices
                 peak_y_on_smoothed = profile_to_plot_and_calc[valid_peaks]
                 self.ax.scatter(valid_peaks, peak_y_on_smoothed, color="red", marker='x', s=50, label="Detected Peaks", zorder=5)

        # --- Process Peak Regions (Calculations now use the SMOOTHED profile) ---
        self.peak_areas_rolling_ball.clear(); self.peak_areas_straight_line.clear(); self.peak_areas_valley.clear()
        num_items_to_plot = len(self.peak_regions)
        profile_range_plot = np.ptp(profile_to_plot_and_calc) if np.ptp(profile_to_plot_and_calc) > 0 else 1.0
        max_text_y_position = np.min(profile_to_plot_and_calc) if len(profile_to_plot_and_calc) > 0 else 0

        for i in range(num_items_to_plot):
            start, end = int(self.peak_regions[i][0]), int(self.peak_regions[i][1])
            if start >= end:
                 self.peak_areas_rolling_ball.append(0.0); self.peak_areas_straight_line.append(0.0); self.peak_areas_valley.append(0.0); continue

            x_region = np.arange(start, end + 1)
            # Use SMOOTHED profile data for region calculations
            profile_region_smoothed = profile_to_plot_and_calc[start : end + 1]
            # Use background calculated from SMOOTHED profile
            bg_start = max(0, min(start, len(self.background)-1))
            bg_end = max(0, min(end + 1, len(self.background)))
            if bg_start < bg_end: background_region = self.background[bg_start:bg_end]
            else: background_region = np.zeros_like(profile_region_smoothed)
            # Interpolate background if needed
            if len(background_region) != len(profile_region_smoothed):
                 try:
                    interp_func = interp1d(np.arange(len(self.background)), self.background, kind='linear', fill_value="extrapolate")
                    background_region = interp_func(x_region)
                 except Exception as interp_err:
                    print(f"Warning: Background interpolation failed for peak {i+1}: {interp_err}")
                    background_region = np.zeros_like(profile_region_smoothed)

            if profile_region_smoothed.shape != background_region.shape:
                print(f"Shape mismatch peak {i+1}. Skip area calculation.");
                self.peak_areas_rolling_ball.append(0.0); self.peak_areas_straight_line.append(0.0); self.peak_areas_valley.append(0.0); continue

            # --- Area Calculations on SMOOTHED Profile ---
            # 1. Rolling Ball Area
            area_rb = max(0, np.trapz(profile_region_smoothed - background_region, x=x_region))
            self.peak_areas_rolling_ball.append(area_rb)

            # 2. Straight Line Area
            area_sl = 0.0
            if start < len(profile_to_plot_and_calc) and end < len(profile_to_plot_and_calc):
                y_baseline_pts_sl = np.array([profile_to_plot_and_calc[start], profile_to_plot_and_calc[end]])
                y_baseline_interp_sl = np.interp(x_region, [start, end], y_baseline_pts_sl)
                area_sl = max(0, np.trapz(profile_region_smoothed - y_baseline_interp_sl, x=x_region))
            self.peak_areas_straight_line.append(area_sl)

            # 3. Valley-to-Valley Area
            area_vv = 0.0
            try:
                search_range = max(15, int((end - start) * 1.5))
                # Use SMOOTHED profile to find valleys
                valley_start_idx, valley_end_idx = self._find_valleys_inverted(profile_to_plot_and_calc, start, end, search_range)
                if 0 <= valley_start_idx < len(profile_to_plot_and_calc) and 0 <= valley_end_idx < len(profile_to_plot_and_calc):
                    y_baseline_pts_vv = np.array([profile_to_plot_and_calc[valley_start_idx], profile_to_plot_and_calc[valley_end_idx]])
                    y_baseline_interp_vv = np.interp(x_region, [valley_start_idx, valley_end_idx], y_baseline_pts_vv)
                    area_vv = max(0, np.trapz(profile_region_smoothed - y_baseline_interp_vv, x=x_region))
                else: print(f"Warning: Invalid indices VV peak {i+1}.")
            except Exception as e_vv: print(f"Error VV calc peak {i+1}: {e_vv}.")
            self.peak_areas_valley.append(area_vv)


            # --- Plot Baselines and Fills (based on smoothed profile) ---
            current_area = 0.0
            if self.method == "Rolling Ball":
                if i == 0: self.ax.plot(np.arange(len(self.background)), self.background, color="green", ls="--", lw=1, label="Rolling Ball BG")
                self.ax.fill_between(x_region, background_region, profile_region_smoothed, where=profile_region_smoothed >= background_region, color="yellow", alpha=0.4, interpolate=True)
                current_area = area_rb
            elif self.method == "Straight Line" and start < len(profile_to_plot_and_calc) and end < len(profile_to_plot_and_calc):
                 self.ax.plot([start, end], y_baseline_pts_sl, color="purple", ls="--", lw=1, label="SL BG" if i == 0 else "")
                 self.ax.fill_between(x_region, y_baseline_interp_sl, profile_region_smoothed, where=profile_region_smoothed >= y_baseline_interp_sl, color="cyan", alpha=0.4, interpolate=True)
                 current_area = area_sl
            elif self.method == "Valley-to-Valley" and 0 <= valley_start_idx < len(profile_to_plot_and_calc) and 0 <= valley_end_idx < len(profile_to_plot_and_calc):
                 self.ax.plot([valley_start_idx, valley_end_idx], y_baseline_pts_vv, color="orange", ls="--", lw=1, label="VV BG" if i == 0 else "")
                 self.ax.fill_between(x_region, y_baseline_interp_vv, profile_region_smoothed, where=profile_region_smoothed >= y_baseline_interp_vv, color="lightblue", alpha=0.4, interpolate=True)
                 current_area = area_vv

            # --- Plot Area Text and Markers ---
            area_text_format = "{:.0f}"
            combined_text = f"Peak {i + 1}\n{area_text_format.format(current_area)}"
            # Position text relative to the smoothed profile's peak
            text_y_base = np.max(profile_region_smoothed)
            text_y_pos = text_y_base + profile_range_plot * 0.06 # Use profile range for spacing
            self.ax.text((start + end) / 2, text_y_pos, combined_text, ha="center", va="bottom", fontsize=7, color='black', zorder=6)
            max_text_y_position = max(max_text_y_position, text_y_pos + profile_range_plot*0.03)

            self.ax.axvline(start, color="gray", ls=":", lw=1.0, alpha=0.8)
            self.ax.axvline(end, color="gray", ls=":", lw=1.0, alpha=0.8)

        # --- Final Plot Configuration ---
        self.ax.set_ylabel("Intensity (Smoothed, Inverted)") # Label reflects smoothing
        self.ax.legend(fontsize='small', loc='upper right')
        self.ax.set_title(f"Smoothed Intensity Profile (σ={self.smoothing_sigma:.1f}) and Peak Regions")
        self.ax.tick_params(axis='x', which='both', bottom=False, top=False, labelbottom=False)

        # Set plot limits based on SMOOTHED profile range
        if len(profile_to_plot_and_calc) > 1: self.ax.set_xlim(0, len(profile_to_plot_and_calc) - 1)
        prof_min_smooth, prof_max_smooth = (np.min(profile_to_plot_and_calc), np.max(profile_to_plot_and_calc)) if len(profile_to_plot_and_calc) > 0 else (0, 1)
        y_max_limit = max(prof_max_smooth, max_text_y_position) + profile_range_plot * 0.05
        y_min_limit = prof_min_smooth - profile_range_plot * 0.05
        if y_max_limit <= y_min_limit: y_max_limit = y_min_limit + 1
        self.ax.set_ylim(y_min_limit, y_max_limit)
        if prof_max_smooth > 10000: self.ax.ticklabel_format(style='sci', axis='y', scilimits=(0,0))

        # --- Display Cropped Image ---
        # (Image display logic remains the same)
        ax_image.clear()
        if hasattr(self, 'cropped_image_for_display') and isinstance(self.cropped_image_for_display, Image.Image):
             try:
                 rotated_pil_image = self.cropped_image_for_display.rotate(90, expand=True)
                 im_array_disp = np.array(rotated_pil_image)
                 ax_image.imshow(im_array_disp, cmap='gray', aspect='auto',
                                 extent=[0, len(profile_to_plot_and_calc)-1, 0, rotated_pil_image.height],
                                 vmin=0, vmax=self.original_max_value)
                 ax_image.set_xlabel("Pixel Index Along Profile Axis")
                 ax_image.set_yticks([]); ax_image.set_ylabel("Lane Width", fontsize='small')
             except Exception as img_e:
                 print(f"Error displaying cropped image preview: {img_e}")
                 ax_image.text(0.5, 0.5, 'Error loading preview', ha='center', va='center', transform=ax_image.transAxes); ax_image.set_xticks([]); ax_image.set_yticks([])
        else:
             ax_image.text(0.5, 0.5, 'No Image Preview', ha='center', va='center', transform=ax_image.transAxes); ax_image.set_xticks([]); ax_image.set_yticks([])


        # --- Adjust and Draw ---
        try: self.fig.subplots_adjust(left=0.08, right=0.98, top=0.92, bottom=0.1, hspace=0.05)
        except Exception as layout_e: print(f"Error adjusting layout: {layout_e}")
        try: self.canvas.draw_idle()
        except Exception as draw_e: print(f"Error drawing canvas: {draw_e}")
        
        plt.close(self.fig)


    # --- NEW HELPER for finding minima on inverted profile ---
    def _find_valleys_inverted(self, profile_data, start, end, search_range):
        """Helper to find valley MINIMA near start and end points on INVERTED profile."""
        profile_len = len(profile_data)

        # Left valley (find minimum)
        valley_start_idx = start
        search_start = max(0, start - search_range)
        if start > search_start : # Check if search range is valid
            # Find index of minimum value within the search window BEFORE the peak start
            min_idx_in_search = np.argmin(profile_data[search_start:start]) + search_start
            valley_start_idx = min_idx_in_search
            # Optional: Check if point immediately before start is even lower
            # if start > 0 and profile_data[start-1] < profile_data[valley_start_idx]:
            #     valley_start_idx = start-1
        # else: valley stays at start

        # Right valley (find minimum)
        valley_end_idx = end
        search_end = min(profile_len, end + search_range + 1)
        slice_start = end + 1
        if slice_start < search_end: # Check if search range is valid
            # Find index of minimum value within the search window AFTER the peak end
            min_idx_in_search = np.argmin(profile_data[slice_start : search_end]) + slice_start
            valley_end_idx = min_idx_in_search
            # Optional: Check if point immediately after end is even lower
            # if end < profile_len - 1 and profile_data[end+1] < profile_data[valley_end_idx]:
            #     valley_end_idx = end+1
        # else: valley stays at end


        # Validate and clamp indices (same as before)
        valley_start_idx = max(0, valley_start_idx)
        valley_end_idx = min(profile_len - 1, valley_end_idx)
        if valley_start_idx > start: valley_start_idx = start # Valley cannot be inside peak region
        if valley_end_idx < end: valley_end_idx = end # Valley cannot be inside peak region
        if valley_end_idx <= valley_start_idx:
            print(f"Warning: Inverted valley detection invalid. Using region boundaries ({start},{end}).")
            valley_start_idx, valley_end_idx = start, end

        return valley_start_idx, valley_end_idx

    # --- get_final_peak_area (Identical to previous versions) ---
    def get_final_peak_area(self):
        """Return the list of calculated peak areas based on the selected method."""
        num_valid_peaks = len(self.peak_regions)
        current_area_list = []
        if self.method == "Rolling Ball": current_area_list = self.peak_areas_rolling_ball
        elif self.method == "Straight Line": current_area_list = self.peak_areas_straight_line
        elif self.method == "Valley-to-Valley": current_area_list = self.peak_areas_valley
        else: return []
        if len(current_area_list) != num_valid_peaks:
            print(f"Warning: Area list length mismatch for method '{self.method}'.")
            return current_area_list[:num_valid_peaks]
        else:
            return current_area_list
    
        
    


class LiveViewLabel(QLabel):
    def __init__(self, font_type, font_size, marker_color, parent=None):
        super().__init__(parent)
        self.setMouseTracking(True)  # Enable mouse tracking
        self.preview_marker_enabled = False
        self.preview_marker_text = ""
        self.preview_marker_position = None
        self.marker_font_type = font_type
        self.marker_font_size = font_size
        self.marker_color = marker_color
        self.setFocusPolicy(Qt.StrongFocus)
        self.bounding_box_preview = []
        self.measure_quantity_mode = False
        self.counter = 0
        self.zoom_level = 1.0  # Initial zoom level
        self.pan_start = None  # Start position for panning
        self.pan_offset = QPointF(0, 0)  # Offset for panning
        self.quad_points = []  # Stores 4 points of the quadrilateral
        self.selected_point = -1  # Index of selected corner (-1 = none)
        self.drag_threshold = 10  # Pixel radius for selecting corners
        self.bounding_box_complete = False
        self.mode=None
        # Add rectangle-related attributes
        self.rectangle_start = None  # Start point of the rectangle
        self.rectangle_end = None    # End point of the rectangle
        self.rectangle_points = []   # Stores the rectangle points
        self.drag_start_pos = None  # For tracking drag operations
        self.draw_edges=True

    def mouseMoveEvent(self, event):
        if self.preview_marker_enabled:
            self.preview_marker_position = event.pos()
            self.update()  # Trigger repaint to show the preview
        if self.selected_point != -1 and self.measure_quantity_mode:# and self.mode=="quad":
            # Update dragged corner position
            self.quad_points[self.selected_point] = self.transform_point(event.pos())
            self.update()  # Show the bounding box preview
        if self.selected_point != -1 and self.mode=="move":
            self.drag_start_pos=event.pos()
        
        super().mouseMoveEvent(event)

    def mousePressEvent(self, event):
        if self.preview_marker_enabled:
            # Place the marker text permanently at the clicked position
            parent = self.parent()
            parent.place_custom_marker(event, self.preview_marker_text)
            self.update()  # Clear the preview
        if self.measure_quantity_mode and self.mode=="quad":
            # Check if clicking near existing corner
            for i, p in enumerate(self.quad_points):
                if (self.transform_point(event.pos()) - p).manhattanLength() < self.drag_threshold:
                    self.selected_point = i
                    return
    
            # Add new point if < 4 corners
            if len(self.quad_points) < 4:
                self.quad_points.append(self.transform_point(event.pos()))
                self.selected_point = len(self.quad_points) - 1
    
            if len(self.quad_points) == 4 and self.zoom_level != 1.0 and not self.bounding_box_complete:
                # self.quad_points = [self.transform_point(p) for p in self.quad_points]
                self.bounding_box_complete = True
            self.update()  # Trigger repaint
        super().mousePressEvent(event)

    def mouseReleaseEvent(self, event):
        if self.mode=="quad":
            self.selected_point = -1
            self.update()
        super().mouseReleaseEvent(event)

    def transform_point(self, point):
        """Transform a point from widget coordinates to image coordinates."""
        if self.zoom_level != 1.0:
            return QPointF(
                (point.x() - self.pan_offset.x()) / self.zoom_level,
                (point.y() - self.pan_offset.y()) / self.zoom_level
            )
        return QPointF(point)

    def zoom_in(self):
        self.zoom_level *= 1.1
        self.update()

    def zoom_out(self):
        self.zoom_level /= 1.1
        if self.zoom_level < 1.0:
            self.zoom_level = 1.0
            self.pan_offset = QPointF(0, 0)  # Reset pan offset when zoom is reset
        self.update()

    def paintEvent(self, event):
        super().paintEvent(event)
        painter = QPainter(self)
        
        painter.setRenderHint(QPainter.Antialiasing, True)
        painter.setRenderHint(QPainter.TextAntialiasing, True)
        painter.setRenderHint(QPainter.SmoothPixmapTransform, True)


        if self.zoom_level != 1.0:
            painter.translate(self.pan_offset)
            painter.scale(self.zoom_level, self.zoom_level)

        # Draw the preview marker if enabled
        if self.preview_marker_enabled and self.preview_marker_position:
            painter.setOpacity(0.5)  # Semi-transparent preview
            font = QFont(self.marker_font_type)
            font.setPointSize(self.marker_font_size)
            painter.setFont(font)
            painter.setPen(self.marker_color)
            text_width = painter.fontMetrics().horizontalAdvance(self.preview_marker_text)
            text_height = painter.fontMetrics().height()
            # Draw the text at the cursor's position
            x, y = self.preview_marker_position.x(), self.preview_marker_position.y()
            if self.zoom_level != 1.0:
                x = (x - self.pan_offset.x()) / self.zoom_level
                y = (y - self.pan_offset.y()) / self.zoom_level
            painter.drawText(int(x - text_width / 2), int(y + text_height / 4), self.preview_marker_text)
            
        if len(self.quad_points) > 0 and len(self.quad_points) <4 :
            for p in self.quad_points:
                painter.setPen(QPen(Qt.red, 2))
                painter.drawEllipse(p, 1, 1)
    
        if len(self.quad_points) == 4 and self.draw_edges==True:
            painter.setPen(QPen(Qt.red, 2))
            painter.drawPolygon(QPolygonF(self.quad_points))
            # Draw draggable corners
            for p in self.quad_points:
                painter.drawEllipse(p, self.drag_threshold, self.drag_threshold)
    
        # Draw the bounding box preview if it exists
        if self.bounding_box_preview:
            painter.setPen(QPen(Qt.red, 2))  # Use green color for the bounding box
            start_x, start_y, end_x, end_y = self.bounding_box_preview
            rect = QRectF(QPointF(start_x, start_y), QPointF(end_x, end_y))
            painter.drawRect(rect)
    
        painter.end()
        self.update()

    def keyPressEvent(self, event):
        if event.key() == Qt.Key_Escape:
            if self.preview_marker_enabled:
                self.preview_marker_enabled = False  # Turn off the preview
                self.update()  # Clear the overlay
            self.measure_quantity_mode = False
            self.counter = 0
            self.bounding_box_complete = False
            self.quad_points = []
            self.mode=None
            self.update()
        super().keyPressEvent(event)

class CombinedSDSApp(QMainWindow):
    def __init__(self):
        super().__init__()
        self.screen = QDesktopWidget().screenGeometry()
        self.screen_width, self.screen_height = self.screen.width(), self.screen.height()
        window_width = int(self.screen_width * 0.5)  # 60% of screen width
        window_height = int(self.screen_height * 0.75)  # 95% of screen height
        self.window_title="IMAGING ASSISTANT V6.0"
        self.setWindowTitle(self.window_title)
        self.setWindowFlags(Qt.Window | Qt.CustomizeWindowHint | Qt.WindowMinimizeButtonHint | Qt.WindowCloseButtonHint)
        self.setSizePolicy(QSizePolicy.Minimum, QSizePolicy.Minimum)
        self.undo_stack = []
        self.redo_stack = []
        self.quantities_peak_area_dict = {}
        self.latest_peak_areas = []
        self.latest_calculated_quantities = []
        self.image_path = None
        self.image = None
        self.image_master= None
        self.image_before_padding = None
        self.image_contrasted=None
        self.image_before_contrast=None
        self.contrast_applied=False
        self.image_padded=False
        self.predict_size=False
        self.warped_image=None
        self.left_markers = []
        self.right_markers = []
        self.top_markers = []
        self.custom_markers=[]
        self.current_left_marker_index = 0
        self.current_right_marker_index = 0
        self.current_top_label_index = 0
        self.font_rotation=-45
        self.image_width=0
        self.image_height=0
        self.new_image_width=0
        self.new_image_height=0
        self.base_name="Image"
        self.image_path=""
        self.x_offset_s=0
        self.y_offset_s=0
        self.peak_area=[]
        self.is_modified=False
        
        self.peak_dialog_settings = {
            'rolling_ball_radius': 50,
            'peak_height_factor': 0.1,
            'peak_distance': 30,
            'peak_prominence_factor': 0.02,
            'peak_spread_pixels': 10,
            'band_estimation_method': "Mean",
            'area_subtraction_method': "Valley-to-Valley"
        }
        self.persist_peak_settings_enabled = True # State of the checkbox
        
        # Variables to store bounding boxes and quantities
        self.bounding_boxes = []
        self.up_bounding_boxes = []
        self.standard_protein_areas=[]
        self.quantities = []
        self.protein_quantities = []
        self.measure_quantity_mode = False
        # Initialize self.marker_values to None initially
        self.marker_values_dict = {
            "Precision Plus All Blue/Unstained": [250, 150, 100, 75, 50, 37, 25, 20, 15, 10],
            "1 kB Plus": [15000, 10000, 8000, 7000, 6000, 5000, 4000, 3000, 2000, 1500, 1000, 850, 650, 500, 400, 300, 200, 100],
        }
        self.top_label=["MWM" , "S1", "S2", "S3" , "S4", "S5" , "S6", "S7", "S8", "S9", "MWM"]
        self.top_label_dict={"Precision Plus All Blue/Unstained": ["MWM", "S1", "S2", "S3", "S4", "S5", "S6", "S7", "S8", "S9", "MWM"]}
        
        
        self.custom_marker_name = "Custom"
        # Load the config file if exists
        
        self.marker_mode = None
        self.left_marker_shift = 0   # Additional shift for marker text
        self.right_marker_shift = 0   # Additional shift for marker tex
        self.top_marker_shift=0 
        self.left_marker_shift_added=0
        self.right_marker_shift_added=0
        self.top_marker_shift_added= 0
        self.left_slider_range=[-1000,1000]
        self.right_slider_range=[-1000,1000]
        self.top_slider_range=[-1000,1000]
               
        
        self.top_padding = 0
        self.font_color = QColor(0, 0, 0)  # Default to black
        self.custom_marker_color = QColor(0, 0, 0)  # Default to black
        self.font_family = "Arial"  # Default font family
        self.font_size = 12  # Default font size
        self.image_array_backup= None
        self.run_predict_MW=False
        
        self.create_menu_bar()
        
        # Main container widget
        main_widget = QWidget()
        self.setCentralWidget(main_widget)
        layout = QVBoxLayout(main_widget)

        # Upper section (Preview and buttons)
        upper_layout = QHBoxLayout()

        self.label_width=int(self.screen_width * 0.28)
    
        self.live_view_label = LiveViewLabel(
            font_type=QFont("Arial"),
            font_size=int(24),
            marker_color=QColor(0,0,0),
            parent=self,
        )
        # Image display
        self.live_view_label.setStyleSheet("border: 1px solid black;")
        # self.live_view_label.setCursor(Qt.CrossCursor)
        self.live_view_label.setFixedSize(self.label_width, self.label_width)
        # self.live_view_label.mousePressEvent = self.add_band()
        # self.live_view_label.mousePressEvent = self.add_band
        
        
       

        # Buttons for image loading and saving
        # Load, save, and crop buttons
        buttons_layout = QVBoxLayout()
        load_button = QPushButton("Load Image")
        load_button.setToolTip("Load an image or a previously saved file. Shortcut: Ctrl+O or CMD+O")
        load_button.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)  # Expand width
        load_button.clicked.connect(self.load_image)
        buttons_layout.addWidget(load_button)
        
        paste_button = QPushButton('Paste Image')
        paste_button.setToolTip("Paste an image from clipboard or folder. Shortcut: Ctrl+V or CMD+V")
        paste_button.clicked.connect(self.paste_image)  # Connect the button to the paste_image method
        paste_button.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)  # Expand width
        buttons_layout.addWidget(paste_button)
    
        
        reset_button = QPushButton("Reset Image")  # Add Reset Image button
        reset_button.setToolTip("Reset all image manipulations and marker placements. Shortcut: Ctrl+R or CMD+R")
        reset_button.clicked.connect(self.reset_image)  # Connect the reset functionality
        reset_button.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)  # Expand width
        buttons_layout.addWidget(reset_button)
        
        
        
        copy_button = QPushButton('Copy Image to Clipboard')
        copy_button.setToolTip("Copy the modified image to clipboard. Shortcut: Ctrl+C or CMD+C")
        copy_button.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)  # Expand width
        copy_button.clicked.connect(self.copy_to_clipboard)
        buttons_layout.addWidget(copy_button)
        
        save_button = QPushButton("Save Image with Configuration")
        save_button.setToolTip("Save the modified image with the configuration files. Shortcut: Ctrl+S or CMD+S")
        save_button.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)  # Expand width
        save_button.clicked.connect(self.save_image)
        buttons_layout.addWidget(save_button)
        
        #if platform.system() == "Windows": # "Darwin" for MacOS # "Windows" for Windows
        #    copy_svg_button = QPushButton('Copy SVG Image to Clipboard')
        #    copy_svg_button.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)  # Expand width
        #    copy_svg_button.clicked.connect(self.copy_to_clipboard_SVG)
        #    buttons_layout.addWidget(copy_svg_button)
            
        
        save_svg_button = QPushButton("Save SVG Image (MS Word Import)")
        save_svg_button.setToolTip("Save the modified image as an SVG file so that it can be modified in MS Word or similar. Shortcut: Ctrl+M or CMD+M")
        save_svg_button.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)  # Expand width
        save_svg_button.clicked.connect(self.save_image_svg)
        buttons_layout.addWidget(save_svg_button)
        
        undo_redo_layout=QHBoxLayout()
        
        undo_button = QPushButton("Undo")
        undo_button.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)  # Expand width
        undo_button.clicked.connect(self.undo_action)
        undo_button.setToolTip("Undo settings related to image. Cannot Undo Marker Placement. Use remove last option. Shortcut: Ctrl+Z or CMD+Z")
        undo_redo_layout.addWidget(undo_button)
        
        redo_button = QPushButton("Redo")
        redo_button.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)  # Expand width
        redo_button.clicked.connect(self.redo_action)
        redo_button.setToolTip("Redo settings related to image. Cannot Undo Marker Placement. Use remove last option.Shortcut: Ctrl+Y or CMD+Y")        
        undo_redo_layout.addWidget(redo_button)
        
        buttons_layout.addLayout(undo_redo_layout)
        
        # New Zoom buttons layout
        zoom_layout = QHBoxLayout()
        
        # Zoom In button
        zoom_in_button = QPushButton("Zoom In")
        zoom_in_button.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        zoom_in_button.clicked.connect(self.zoom_in)
        zoom_in_button.setToolTip("Increase zoom level. Click the display window and use arrow keys for moving")
        
        # Zoom Out button
        zoom_out_button = QPushButton("Zoom Out")
        zoom_out_button.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        zoom_out_button.clicked.connect(self.zoom_out)
        zoom_out_button.setToolTip("Decrease zoom level. Click the display window and use arrow keys for moving")
        
        # Add Zoom buttons to the zoom layout
        zoom_layout.addWidget(zoom_in_button)
        zoom_layout.addWidget(zoom_out_button)
        
        # Add Zoom layout below Undo/Redo
        buttons_layout.addLayout(zoom_layout)
        
        # buttons_layout.addStretch()
        
        
        upper_layout.addWidget(self.live_view_label, stretch=1)
        upper_layout.addLayout(buttons_layout, stretch=1)
        layout.addLayout(upper_layout)

        # Lower section (Tabbed interface)
        self.tab_widget = QTabWidget()
        # self.tab_widget.setToolTip("Change the tabs quickly with shortcut: Ctrl+1,2,3 or 4 and CMD+1,2,3 or 4")
        self.tab_widget.addTab(self.font_and_image_tab(), "Image and Font")
        self.tab_widget.addTab(self.create_cropping_tab(), "Transform")
        self.tab_widget.addTab(self.create_white_space_tab(), "Padding")
        self.tab_widget.addTab(self.create_markers_tab(), "Markers")
        self.tab_widget.addTab(self.combine_image_tab(), "Overlap Images")
        self.tab_widget.addTab(self.analysis_tab(), "Analysis")
        
        layout.addWidget(self.tab_widget)
        
        # Example: Shortcut for loading an image (Ctrl + O)
        self.load_shortcut = QShortcut(QKeySequence("Ctrl+O"), self)
        self.load_shortcut.activated.connect(self.load_image)
        
        # Example: Shortcut for saving an image (Ctrl + S)
        self.save_shortcut = QShortcut(QKeySequence("Ctrl+S"), self)
        self.save_shortcut.activated.connect(self.save_image)
        
        # Example: Shortcut for resetting the image (Ctrl + R)
        self.reset_shortcut = QShortcut(QKeySequence("Ctrl+R"), self)
        self.reset_shortcut.activated.connect(self.reset_image)
        
        # Example: Shortcut for copying the image to clipboard (Ctrl + C)
        self.copy_shortcut = QShortcut(QKeySequence("Ctrl+C"), self)
        self.copy_shortcut.activated.connect(self.copy_to_clipboard)
        
        # Example: Shortcut for pasting an image from clipboard (Ctrl + V)
        self.paste_shortcut = QShortcut(QKeySequence("Ctrl+V"), self)
        self.paste_shortcut.activated.connect(self.paste_image)
        
        # Example: Shortcut for predicting molecular weight (Ctrl + P)
        self.predict_shortcut = QShortcut(QKeySequence("Ctrl+P"), self)
        self.predict_shortcut.activated.connect(self.predict_molecular_weight)
        
        # Example: Shortcut for clearing the prediction marker (Ctrl + Shift + P)
        self.clear_predict_shortcut = QShortcut(QKeySequence("Ctrl+Shift+P"), self)
        self.clear_predict_shortcut.activated.connect(self.clear_predict_molecular_weight)
        
        # Example: Shortcut for saving SVG image (Ctrl + Shift + S)
        self.save_svg_shortcut = QShortcut(QKeySequence("Ctrl+M"), self)
        self.save_svg_shortcut.activated.connect(self.save_image_svg)
        
        # Example: Shortcut for enabling left marker mode (Ctrl + L)
        self.left_marker_shortcut = QShortcut(QKeySequence("Ctrl+Shift+L"), self)
        self.left_marker_shortcut.activated.connect(self.enable_left_marker_mode)
        
        # Example: Shortcut for enabling right marker mode (Ctrl + R)
        self.right_marker_shortcut = QShortcut(QKeySequence("Ctrl+Shift+R"), self)
        self.right_marker_shortcut.activated.connect(self.enable_right_marker_mode)
        
        # Example: Shortcut for enabling top marker mode (Ctrl + T)
        self.top_marker_shortcut = QShortcut(QKeySequence("Ctrl+Shift+T"), self)
        self.top_marker_shortcut.activated.connect(self.enable_top_marker_mode)
        
        # Example: Shortcut for toggling grid visibility (Ctrl + G)
        self.grid_shortcut = QShortcut(QKeySequence("Ctrl+Shift+G"), self)
        self.grid_shortcut.activated.connect(lambda: self.show_grid_checkbox.setChecked(not self.show_grid_checkbox.isChecked()))
        
        # Example: Shortcut for toggling grid visibility (Ctrl + G)
        self.guidelines_shortcut = QShortcut(QKeySequence("Ctrl+G"), self)
        self.guidelines_shortcut.activated.connect(lambda: self.show_guides_checkbox.setChecked(not self.show_guides_checkbox.isChecked()))
        
        # Example: Shortcut for increasing grid size (Ctrl + Shift + Up)
        self.increase_grid_size_shortcut = QShortcut(QKeySequence("Ctrl+Shift+Up"), self)
        self.increase_grid_size_shortcut.activated.connect(lambda: self.grid_size_input.setValue(self.grid_size_input.value() + 1))
        
        # Example: Shortcut for decreasing grid size (Ctrl + Shift + Down)
        self.decrease_grid_size_shortcut = QShortcut(QKeySequence("Ctrl+Shift+Down"), self)
        self.decrease_grid_size_shortcut.activated.connect(lambda: self.grid_size_input.setValue(self.grid_size_input.value() - 1))
        
        # # Example: Shortcut for increasing font size (Ctrl + Plus)
        # self.increase_font_size_shortcut = QShortcut(QKeySequence("Ctrl+W+Up"), self)
        # self.increase_font_size_shortcut.activated.connect(lambda: self.font_size_spinner.setValue(self.font_size_spinner.value() + 1))
        
        # # Example: Shortcut for decreasing font size (Ctrl + Minus)
        # self.decrease_font_size_shortcut = QShortcut(QKeySequence("Ctrl+W+Down"), self)
        # self.decrease_font_size_shortcut.activated.connect(lambda: self.font_size_spinner.setValue(self.font_size_spinner.value() - 1))
        
        # Example: Shortcut for custom marker left arrow (Ctrl + Left Arrow)
        self.custom_marker_left_arrow_shortcut = QShortcut(QKeySequence("Ctrl+Left"), self)
        self.custom_marker_left_arrow_shortcut.activated.connect(lambda: self.arrow_marker("←"))
        self.custom_marker_left_arrow_shortcut.activated.connect(self.enable_custom_marker_mode)
        
        # Example: Shortcut for custom marker right arrow (Ctrl + Right Arrow)
        self.custom_marker_right_arrow_shortcut = QShortcut(QKeySequence("Ctrl+Right"), self)
        self.custom_marker_right_arrow_shortcut.activated.connect(lambda: self.arrow_marker("→"))
        self.custom_marker_right_arrow_shortcut.activated.connect(self.enable_custom_marker_mode)
        
        # Example: Shortcut for custom marker top arrow (Ctrl + Up Arrow)
        self.custom_marker_top_arrow_shortcut = QShortcut(QKeySequence("Ctrl+Up"), self)
        self.custom_marker_top_arrow_shortcut.activated.connect(lambda: self.arrow_marker("↑"))
        self.custom_marker_top_arrow_shortcut.activated.connect(self.enable_custom_marker_mode)
        
        # Example: Shortcut for custom marker bottom arrow (Ctrl + Down Arrow)
        self.custom_marker_bottom_arrow_shortcut = QShortcut(QKeySequence("Ctrl+Down"), self)
        self.custom_marker_bottom_arrow_shortcut.activated.connect(lambda: self.arrow_marker("↓"))
        self.custom_marker_bottom_arrow_shortcut.activated.connect(self.enable_custom_marker_mode)
        
        # Example: Shortcut for inverting image (Ctrl + T)
        self.invert_shortcut = QShortcut(QKeySequence("Ctrl+I"), self)
        self.invert_shortcut.activated.connect(self.invert_image)
        
        # Example: Shortcut for converting to grayscale (Ctrl + T)
        self.invert_shortcut = QShortcut(QKeySequence("Ctrl+B"), self)
        self.invert_shortcut.activated.connect(self.convert_to_black_and_white)
        
        # Example: Move quickly between tabs (Ctrl + 1,2,3,4)
        self.move_tab_1_shortcut = QShortcut(QKeySequence("Ctrl+1"), self)
        self.move_tab_1_shortcut.activated.connect(lambda: self.move_tab(0))
        
        self.move_tab_2_shortcut = QShortcut(QKeySequence("Ctrl+2"), self)
        self.move_tab_2_shortcut.activated.connect(lambda: self.move_tab(1))
        
        self.move_tab_3_shortcut = QShortcut(QKeySequence("Ctrl+3"), self)
        self.move_tab_3_shortcut.activated.connect(lambda: self.move_tab(2))
        
        self.move_tab_4_shortcut = QShortcut(QKeySequence("Ctrl+4"), self)
        self.move_tab_4_shortcut.activated.connect(lambda: self.move_tab(3))
        
        self.move_tab_5_shortcut = QShortcut(QKeySequence("Ctrl+5"), self)
        self.move_tab_5_shortcut.activated.connect(lambda: self.move_tab(4))
        
        self.move_tab_6_shortcut = QShortcut(QKeySequence("Ctrl+6"), self)
        self.move_tab_6_shortcut.activated.connect(lambda: self.move_tab(5))
        
        self.undo_shortcut = QShortcut(QKeySequence("Ctrl+Z"), self)
        self.undo_shortcut.activated.connect(self.undo_action)
        
        self.redo_shortcut = QShortcut(QKeySequence("Ctrl+Y"), self)
        self.redo_shortcut.activated.connect(self.redo_action)
        self.load_config()
        
    def prompt_save_if_needed(self):
        if not self.is_modified:
            return True # No changes, proceed
        """Checks if modified and prompts user to save. Returns True to proceed, False to cancel."""
        reply = QMessageBox.question(self, 'Unsaved Changes',
                                     "You have unsaved changes. Do you want to save them?",
                                     QMessageBox.Save | QMessageBox.Discard | QMessageBox.Cancel,
                                     QMessageBox.Save) # Default button is Save

        if reply == QMessageBox.Save:
            return self.save_image() # Returns True if saved, False if save cancelled
        elif reply == QMessageBox.Discard:
            return True # Proceed without saving
        else: # Cancelled
            return False # Abort the current action (close/load)

    def closeEvent(self, event):
        """Overrides the default close event to prompt for saving."""
        if self.prompt_save_if_needed():

            event.accept() # Proceed with closing
        else:
            event.ignore() # Abort closing
            
    def qimage_to_numpy(self, qimage: QImage) -> np.ndarray:
        """Converts QImage to NumPy array, preserving format and handling row padding."""
        if qimage.isNull():
            return None
    
        img_format = qimage.format()
        height = qimage.height()
        width = qimage.width()
        bytes_per_line = qimage.bytesPerLine() # Actual bytes per row (includes padding)
    
    
        ptr = qimage.constBits()
        if not ptr:
            return None
    
        # Ensure ptr is treated as a bytes object of the correct size
        # ptr is a sip.voidptr, needs explicit size setting for buffer protocol
        try:
            ptr.setsize(qimage.byteCount())
        except AttributeError: # Handle cases where setsize might not be needed or available depending on Qt/Python version
            pass # Continue, hoping buffer protocol works based on qimage.byteCount()
    
        buffer_data = bytes(ptr) # Create a bytes object from the pointer
        if len(buffer_data) != qimage.byteCount():
             print(f"qimage_to_numpy: Warning - Buffer size mismatch. Expected {qimage.byteCount()}, got {len(buffer_data)}.")
             # Fallback or error? Let's try to continue but warn.
    
        # --- Grayscale Formats ---
        if img_format == QImage.Format_Grayscale16:
            bytes_per_pixel = 2
            dtype = np.uint16
            expected_bytes_per_line = width * bytes_per_pixel
            if bytes_per_line == expected_bytes_per_line:
                # No padding, simple reshape
                arr = np.frombuffer(buffer_data, dtype=dtype).reshape(height, width)
                return arr.copy() # Return a copy
            else:
                # Handle padding
                arr = np.zeros((height, width), dtype=dtype)
                for y in range(height):
                    start = y * bytes_per_line
                    row_data = buffer_data[start : start + expected_bytes_per_line]
                    if len(row_data) == expected_bytes_per_line:
                        arr[y] = np.frombuffer(row_data, dtype=dtype)
                    else:
                        print(f"qimage_to_numpy (Grayscale16): Row data length mismatch for row {y}. Skipping row.") # Error handling
                return arr
    
        elif img_format == QImage.Format_Grayscale8:
            bytes_per_pixel = 1
            dtype = np.uint8
            expected_bytes_per_line = width * bytes_per_pixel
            if bytes_per_line == expected_bytes_per_line:
                arr = np.frombuffer(buffer_data, dtype=dtype).reshape(height, width)
                return arr.copy()
            else:
                arr = np.zeros((height, width), dtype=dtype)
                for y in range(height):
                    start = y * bytes_per_line
                    row_data = buffer_data[start : start + expected_bytes_per_line]
                    if len(row_data) == expected_bytes_per_line:
                        arr[y] = np.frombuffer(row_data, dtype=dtype)
                    else:
                        print(f"qimage_to_numpy (Grayscale8): Row data length mismatch for row {y}. Skipping row.")
                return arr
    
        # --- Color Formats ---
        # ARGB32 (often BGRA in memory) & RGBA8888 (often RGBA in memory)
        elif img_format in (QImage.Format_ARGB32, QImage.Format_RGBA8888, QImage.Format_ARGB32_Premultiplied):
            bytes_per_pixel = 4
            dtype = np.uint8
            expected_bytes_per_line = width * bytes_per_pixel
            if bytes_per_line == expected_bytes_per_line:
                arr = np.frombuffer(buffer_data, dtype=dtype).reshape(height, width, 4)
                # QImage ARGB32 is typically BGRA in memory order for numpy
                # QImage RGBA8888 is typically RGBA in memory order for numpy
                # Let's return the raw order for now. The caller might need to swap if needed.
                # If Format_ARGB32, the array is likely BGRA.
                # If Format_RGBA8888, the array is likely RGBA.
                return arr.copy()
            else:
                arr = np.zeros((height, width, 4), dtype=dtype)
                for y in range(height):
                    start = y * bytes_per_line
                    row_data = buffer_data[start : start + expected_bytes_per_line]
                    if len(row_data) == expected_bytes_per_line:
                        arr[y] = np.frombuffer(row_data, dtype=dtype).reshape(width, 4)
                    else:
                        print(f"qimage_to_numpy ({img_format}): Row data length mismatch for row {y}. Skipping row.")
                return arr
    
        # RGB32 (often BGRX or RGBX in memory) & RGBX8888
        elif img_format in (QImage.Format_RGB32, QImage.Format_RGBX8888):
            bytes_per_pixel = 4 # Stored with an ignored byte
            dtype = np.uint8
            expected_bytes_per_line = width * bytes_per_pixel
            if bytes_per_line == expected_bytes_per_line:
                arr = np.frombuffer(buffer_data, dtype=dtype).reshape(height, width, 4)
                # Memory order is often BGRX (Blue, Green, Red, Ignored Alpha/Padding) for RGB32
                # Let's return the 4 channels for now.
                return arr.copy()
            else:
                arr = np.zeros((height, width, 4), dtype=dtype)
                for y in range(height):
                    start = y * bytes_per_line
                    row_data = buffer_data[start : start + expected_bytes_per_line]
                    if len(row_data) == expected_bytes_per_line:
                        arr[y] = np.frombuffer(row_data, dtype=dtype).reshape(width, 4)
                    else:
                        print(f"qimage_to_numpy ({img_format}): Row data length mismatch for row {y}. Skipping row.")
                return arr
    
        # RGB888 (Tightly packed RGB)
        elif img_format == QImage.Format_RGB888:
            bytes_per_pixel = 3
            dtype = np.uint8
            expected_bytes_per_line = width * bytes_per_pixel
            if bytes_per_line == expected_bytes_per_line:
                arr = np.frombuffer(buffer_data, dtype=dtype).reshape(height, width, 3)
                # QImage RGB888 is RGB order in memory
                return arr.copy()
            else:
                arr = np.zeros((height, width, 3), dtype=dtype)
                for y in range(height):
                    start = y * bytes_per_line
                    row_data = buffer_data[start : start + expected_bytes_per_line]
                    if len(row_data) == expected_bytes_per_line:
                        arr[y] = np.frombuffer(row_data, dtype=dtype).reshape(width, 3)
                    else:
                        print(f"qimage_to_numpy (RGB888): Row data length mismatch for row {y}. Skipping row.")
                return arr
    
        # --- Fallback / Conversion Attempt ---
        else:
            # For other formats, try converting to a known format first
            try:
                # Convert to ARGB32 as a robust intermediate format
                qimage_conv = qimage.convertToFormat(QImage.Format_ARGB32)
                if qimage_conv.isNull():
                    print("qimage_to_numpy: Fallback conversion to ARGB32 failed.")
                    # Last resort: try Grayscale8
                    qimage_conv_gray = qimage.convertToFormat(QImage.Format_Grayscale8)
                    if qimage_conv_gray.isNull():
                         return None
                    else:
                        return self.qimage_to_numpy(qimage_conv_gray) # Recursive call with Grayscale8
    
                return self.qimage_to_numpy(qimage_conv) # Recursive call with ARGB32
            except Exception as e:
                print(f"qimage_to_numpy: Error during fallback conversion: {e}")
                return None

    def numpy_to_qimage(self, array: np.ndarray) -> QImage:
        """Converts NumPy array to QImage, selecting appropriate format."""
        if array is None:
            return QImage() # Return invalid QImage
    
        if not isinstance(array, np.ndarray):
            return QImage()
    
        try:
            if array.ndim == 2: # Grayscale
                height, width = array.shape
                if array.dtype == np.uint16:
                    bytes_per_line = width * 2
                    # Ensure data is contiguous
                    contiguous_array = np.ascontiguousarray(array)
                    # Create QImage directly from contiguous data buffer
                    qimg = QImage(contiguous_array.data, width, height, bytes_per_line, QImage.Format_Grayscale16)
                    # Important: QImage doesn't own the buffer by default. Return a copy.
                    return qimg.copy()
                elif array.dtype == np.uint8:
                    bytes_per_line = width * 1
                    contiguous_array = np.ascontiguousarray(array)
                    qimg = QImage(contiguous_array.data, width, height, bytes_per_line, QImage.Format_Grayscale8)
                    return qimg.copy()
                elif np.issubdtype(array.dtype, np.floating):
                     # Assume float is in 0-1 range, scale to uint8
                     img_norm = np.clip(array * 255.0, 0, 255).astype(np.uint8)
                     bytes_per_line = width * 1
                     contiguous_array = np.ascontiguousarray(img_norm)
                     qimg = QImage(contiguous_array.data, width, height, bytes_per_line, QImage.Format_Grayscale8)
                     return qimg.copy()
                else:
                    raise TypeError(f"Unsupported grayscale NumPy dtype: {array.dtype}")
    
            elif array.ndim == 3: # Color
                height, width, channels = array.shape
                if channels == 3 and array.dtype == np.uint8:
                    # Assume input is BGR (common from OpenCV), convert to RGB for QImage.Format_RGB888
                    # Make a contiguous copy before conversion
                    contiguous_array_bgr = np.ascontiguousarray(array)
                    rgb_image = cv2.cvtColor(contiguous_array_bgr, cv2.COLOR_BGR2RGB)
                    # rgb_image is now contiguous RGB
                    bytes_per_line = width * 3
                    qimg = QImage(rgb_image.data, width, height, bytes_per_line, QImage.Format_RGB888)
                    return qimg.copy()
                elif channels == 4 and array.dtype == np.uint8:
                    # Assume input is BGRA (matches typical QImage.Format_ARGB32 memory layout)
                    contiguous_array_bgra = np.ascontiguousarray(array)
                    bytes_per_line = width * 4
                    qimg = QImage(contiguous_array_bgra.data, width, height, bytes_per_line, QImage.Format_ARGB32)
                    return qimg.copy()
                # Add handling for 16-bit color if needed (less common for display)
                elif channels == 3 and array.dtype == np.uint16:
                     # Downscale to 8-bit for display format
                     array_8bit = (array / 257.0).astype(np.uint8)
                     contiguous_array_bgr = np.ascontiguousarray(array_8bit)
                     rgb_image = cv2.cvtColor(contiguous_array_bgr, cv2.COLOR_BGR2RGB)
                     bytes_per_line = width * 3
                     qimg = QImage(rgb_image.data, width, height, bytes_per_line, QImage.Format_RGB888)
                     return qimg.copy()
                elif channels == 4 and array.dtype == np.uint16:
                     # Downscale color channels, keep alpha potentially?
                     array_8bit = (array / 257.0).astype(np.uint8) # Downscales all channels including alpha
                     contiguous_array_bgra = np.ascontiguousarray(array_8bit)
                     bytes_per_line = width * 4
                     qimg = QImage(contiguous_array_bgra.data, width, height, bytes_per_line, QImage.Format_ARGB32)
                     return qimg.copy()
                else:
                     raise TypeError(f"Unsupported color NumPy dtype/channel combination: {array.dtype} / {channels} channels")
            else:
                raise ValueError(f"Unsupported array dimension: {array.ndim}")
    
        except Exception as e:
            traceback.print_exc()
            return QImage() # Return invalid QImage on error

    def get_image_format(self, image=None):
        """Helper to safely get the format of self.image or a provided image."""
        img_to_check = image if image is not None else self.image
        if img_to_check and not img_to_check.isNull():
            return img_to_check.format()
        return None

    def get_compatible_grayscale_format(self, image=None):
        """Returns Format_Grayscale16 if input is 16-bit, else Format_Grayscale8."""
        current_format = self.get_image_format(image)
        if current_format == QImage.Format_Grayscale16:
            return QImage.Format_Grayscale16
        else: # Treat 8-bit grayscale or color as needing 8-bit target
            return QImage.Format_Grayscale8
    # --- END: Helper Functions ---
        
    def quadrilateral_to_rect(self, image, quad_points):
       # This implementation uses the more robust NumPy/OpenCV approach from Snippet 7
       if not image or image.isNull():
           QMessageBox.warning(self, "Warp Error", "Invalid input image.")
           return None
       if len(quad_points) != 4:
            QMessageBox.warning(self, "Warp Error", "Need exactly 4 points for quadrilateral.")
            return None

       # --- Convert QImage to NumPy array (Preserve bit depth) ---
       try:
           img_array = self.qimage_to_numpy(image)
           if img_array is None: raise ValueError("NumPy Conversion returned None")
       except Exception as e:
           QMessageBox.warning(self, "Warp Error", f"Failed to convert image to NumPy: {e}")
           return None

       # Ensure array is 2D grayscale
       if img_array.ndim == 3:
            print("Warning: Warping input array is 3D. Converting to grayscale.")
            color_code = cv2.COLOR_BGR2GRAY if img_array.shape[2] == 3 else cv2.COLOR_BGRA2GRAY
            img_array = cv2.cvtColor(img_array, color_code)
            # Decide target format - stick to 8-bit if source was color
            # img_array = img_array.astype(np.uint8)


       # --- Transform points from LiveViewLabel space to Image space ---
       # This scaling part remains potentially complex and needs validation
       # based on the exact coordinate system flow through cropping/padding.
       # Assuming direct scaling for now.
       label_width = self.live_view_label.width()
       label_height = self.live_view_label.height()
       current_img_width = image.width() # Use the width of the image *passed in*
       current_img_height = image.height()
       scale_x_disp = current_img_width / label_width if label_width > 0 else 1
       scale_y_disp = current_img_height / label_height if label_height > 0 else 1

       src_points_img = []
       for point in quad_points:
           x_view, y_view = point.x(), point.y()
           if self.live_view_label.zoom_level != 1.0:
               x_view = (x_view - self.live_view_label.pan_offset.x()) / self.live_view_label.zoom_level
               y_view = (y_view - self.live_view_label.pan_offset.y()) / self.live_view_label.zoom_level
           x_image = x_view * scale_x_disp
           y_image = y_view * scale_y_disp
           src_points_img.append([x_image, y_image])
       src_np = np.array(src_points_img, dtype=np.float32)

       # --- Define Destination Rectangle ---
       width_a = np.linalg.norm(src_np[0] - src_np[1])
       width_b = np.linalg.norm(src_np[2] - src_np[3])
       height_a = np.linalg.norm(src_np[0] - src_np[3])
       height_b = np.linalg.norm(src_np[1] - src_np[2])
       max_width = int(max(width_a, width_b))
       max_height = int(max(height_a, height_b))

       if max_width <= 0 or max_height <= 0:
            QMessageBox.warning(self, "Warp Error", "Invalid destination rectangle size.")
            return None

       dst_np = np.array([
           [0, 0], [max_width - 1, 0], [max_width - 1, max_height - 1], [0, max_height - 1]
       ], dtype=np.float32)

       # --- Perform Perspective Warp using OpenCV ---
       try:
           matrix = cv2.getPerspectiveTransform(src_np, dst_np)
           # Warp the original NumPy array (could be uint8 or uint16)
           warped_array = cv2.warpPerspective(img_array, matrix, (max_width, max_height),
                                              flags=cv2.INTER_LINEAR, # Use linear interpolation
                                              borderMode=cv2.BORDER_CONSTANT,
                                              borderValue=0) # Fill borders with black
       except Exception as e:
            QMessageBox.warning(self, "Warp Error", f"OpenCV perspective warp failed: {e}")
            return None

       # --- Convert warped NumPy array back to QImage ---
       try:
           warped_qimage = self.numpy_to_qimage(warped_array) # Handles uint8/uint16
           if warped_qimage.isNull(): raise ValueError("Conversion failed.")
           return warped_qimage
       except Exception as e:
           QMessageBox.warning(self, "Warp Error", f"Failed to convert warped array to QImage: {e}")
           return None
   # --- END: Modified Warping ---
    
    

        
    def create_menu_bar(self):
        # Create the menu bar
        menubar = self.menuBar()

        # Create the "File" menu
        file_menu = menubar.addMenu("File")

        # Add "Load Image" action
        load_action = QAction("Load Image", self)
        load_action.triggered.connect(self.load_image)
        file_menu.addAction(load_action)

        # Add "Save Image" action
        save_action = QAction("Save Image", self)
        save_action.triggered.connect(self.save_image)
        file_menu.addAction(save_action)
        
        # Add "Save Image SVG" action
        save_action_svg = QAction("Save Image as SVG", self)
        save_action_svg.triggered.connect(self.save_image_svg)
        file_menu.addAction(save_action_svg)

        # Add "Reset Image" action
        reset_action = QAction("Reset Image", self)
        reset_action.triggered.connect(self.reset_image)
        file_menu.addAction(reset_action)

        # Add a separator
        file_menu.addSeparator()

        # Add "Exit" action
        exit_action = QAction("Exit", self)
        exit_action.triggered.connect(self.close)
        file_menu.addAction(exit_action)

        # Create the "About" menu
        about_menu = menubar.addMenu("About")

        # Add "GitHub" action
        github_action = QAction("GitHub", self)
        github_action.triggered.connect(self.open_github)
        about_menu.addAction(github_action)
        
        # self.statusBar().showMessage("Ready")
    def open_github(self):
        # Open the GitHub link in the default web browser
        QDesktopServices.openUrl(QUrl("https://github.com/Anindya-Karmaker/Imaging-Assistant"))
        
    def zoom_in(self):
        self.live_view_label.zoom_in()
        self.update_live_view()

    def zoom_out(self):
        self.live_view_label.zoom_out()
        self.update_live_view()
    
    def enable_standard_protein_mode(self):
        """"Enable mode to define standard protein amounts for creating a standard curve."""
        self.measure_quantity_mode = True
        self.live_view_label.measure_quantity_mode = True
        self.live_view_label.setCursor(Qt.CrossCursor)
        self.live_view_label.setMouseTracking(True)  # Ensure mouse events are enabled
        self.setMouseTracking(True)  # Ensure parent also tracks mouse
        # Assign mouse event handlers for bounding box creation
        self.live_view_label.mousePressEvent = lambda event: self.start_bounding_box(event)
        self.live_view_label.mouseReleaseEvent = lambda event: self.end_standard_bounding_box(event)
        self.live_view_label.setCursor(Qt.CrossCursor)
    
    def enable_measure_protein_mode(self):
        """Enable mode to measure protein quantity using the standard curve."""
        if len(self.quantities) < 2:
            QMessageBox.warning(self, "Error", "At least two standard protein amounts are needed to measure quantity.")
        self.measure_quantity_mode = True
        self.live_view_label.measure_quantity_mode = True
        self.live_view_label.setCursor(Qt.CrossCursor)
        self.live_view_label.setMouseTracking(True)  # Ensure mouse events are enabled
        self.setMouseTracking(True)  # Ensure parent also tracks mouse
        self.live_view_label.mousePressEvent = lambda event: self.start_bounding_box(event)
        self.live_view_label.mouseReleaseEvent = lambda event: self.end_measure_bounding_box(event)
        self.live_view_label.setCursor(Qt.CrossCursor)

    def call_live_view(self):
        self.update_live_view()      

    def analyze_bounding_box(self, pil_image_for_dialog, standard): # Input is PIL Image
        peak_area = None # Initialize
        # Clear previous results before new analysis
        self.latest_peak_areas = []
        self.latest_calculated_quantities = []

        if standard:
            quantity, ok = QInputDialog.getText(self, "Enter Standard Quantity", "Enter the known amount (e.g., 1.5):")
            if ok and quantity:
                try:
                    quantity_value = float(quantity.split()[0])
                    # Calculate peak area using the PIL data
                    peak_area = self.calculate_peak_area(pil_image_for_dialog) # Expects PIL
                    if peak_area is not None and len(peak_area) > 0: # Check if list is not empty
                        total_area = sum(peak_area)
                        self.quantities_peak_area_dict[quantity_value] = round(total_area, 3)
                        self.standard_protein_areas_text.setText(str(list(self.quantities_peak_area_dict.values())))
                        self.standard_protein_values.setText(str(list(self.quantities_peak_area_dict.keys())))
                        print(f"Standard Added: Qty={quantity_value}, Area={total_area:.3f}")
                        # Store the areas for potential later export if needed (though usually total area is used for standards)
                        self.latest_peak_areas = [round(a, 3) for a in peak_area] # Store individual areas too
                    else:
                         print("Peak area calculation cancelled or failed for standard.")
                         self.latest_peak_areas = [] # Ensure it's cleared

                except (ValueError, IndexError) as e:
                    QMessageBox.warning(self, "Input Error", f"Please enter a valid number for quantity. Error: {e}")
                    self.latest_peak_areas = []
                except Exception as e:
                     QMessageBox.critical(self, "Analysis Error", f"An error occurred during standard analysis: {e}")
                     self.latest_peak_areas = []
            else:
                print("Standard quantity input cancelled.")
                self.latest_peak_areas = []

        else: # Analyze sample
            # Calculate peak area using the PIL data
            peak_area = self.calculate_peak_area(pil_image_for_dialog) # Expects PIL

            if peak_area is not None and len(peak_area) > 0: # Check if list is not empty
                self.latest_peak_areas = [round(a, 3) for a in peak_area] # Store latest areas
                print(f"Sample Analysis: Calculated Areas = {self.latest_peak_areas}")

                if len(self.quantities_peak_area_dict) >= 2:
                    # Calculate quantities and store them
                    self.latest_calculated_quantities = self.calculate_unknown_quantity(
                        list(self.quantities_peak_area_dict.values()), # Standard Areas (total)
                        list(self.quantities_peak_area_dict.keys()),   # Standard Quantities
                        self.latest_peak_areas                         # Sample Peak Areas (individual)
                    )
                    print(f"Sample Analysis: Calculated Quantities = {self.latest_calculated_quantities}")

                else:
                    self.latest_calculated_quantities = [] # No quantities calculated

                try:
                    # Display the areas in the text box
                    self.target_protein_areas_text.setText(str(self.latest_peak_areas))
                except Exception as e:
                    print(f"Error displaying sample areas: {e}")
                    self.target_protein_areas_text.setText("Error")
            else:
                 print("Peak area calculation cancelled or failed for sample.")
                 self.target_protein_areas_text.setText("N/A")
                 self.latest_peak_areas = []
                 self.latest_calculated_quantities = []


        # --- UI updates after analysis ---
        self.update_live_view() # Update display
    
    def calculate_peak_area(self, pil_image_for_dialog): # Input is EXPECTED to be PIL Image
        """Opens the PeakAreaDialog for interactive adjustment and area calculation."""

        # --- Validate Input ---
        if pil_image_for_dialog is None:
            print("Error: No PIL Image data provided to calculate_peak_area.")
            return None
        if not isinstance(pil_image_for_dialog, Image.Image):
             # This indicates an error in the calling function (process_sample/standard)
             QMessageBox.critical(self, "Internal Error", f"calculate_peak_area expected PIL Image, got {type(pil_image_for_dialog)}")
             return None
        # --- End Validation ---


        # --- Call PeakAreaDialog passing the PIL Image with the 'cropped_data' keyword ---
        # Assumes PeakAreaDialog __init__ expects 'cropped_data' as the first arg (PIL Image type)
        dialog = PeakAreaDialog(
            cropped_data=pil_image_for_dialog, # Pass the received PIL Image
            current_settings=self.peak_dialog_settings,
            persist_checked=self.persist_peak_settings_enabled,
            parent=self
        )

        peak_areas = None
        if dialog.exec_() == QDialog.Accepted:
            peak_areas = dialog.get_final_peak_area()
            if dialog.should_persist_settings():
                self.peak_dialog_settings = dialog.get_current_settings()
                self.persist_peak_settings_enabled = True
            else:
                self.persist_peak_settings_enabled = False
        else:
            print("PeakAreaDialog cancelled.")

        return peak_areas if peak_areas is not None else []
    
    
    def calculate_unknown_quantity(self, standard_total_areas, known_quantities, sample_peak_areas):
        """Calculates unknown quantities based on standards and sample areas. Returns a list of quantities."""
        if not standard_total_areas or not known_quantities or len(standard_total_areas) != len(known_quantities):
            print("Error: Invalid standard data for quantity calculation.")
            return []
        if len(standard_total_areas) < 2:
            print("Error: Need at least 2 standards for regression.")
            return []
        if not sample_peak_areas:
             print("Warning: No sample peak areas provided for quantity calculation.")
             return []

        calculated_quantities = []
        try:
            # Use total standard area vs known quantity for the standard curve
            coefficients = np.polyfit(standard_total_areas, known_quantities, 1)

            # Apply the standard curve to *each individual* sample peak area
            for area in sample_peak_areas:
                val = np.polyval(coefficients, area)
                calculated_quantities.append(round(val, 2))

            # Display the results in a message box (optional, but helpful feedback)
            QMessageBox.information(self, "Protein Quantification", f"Predicted Quantities: {calculated_quantities} units")

        except Exception as e:
             print(f"Error during quantity calculation: {e}")
             QMessageBox.warning(self, "Calculation Error", f"Could not calculate quantities: {e}")
             return [] # Return empty list on error

        return calculated_quantities # <-- Return the list

        
    def draw_quantity_text(self, painter, x, y, quantity, scale_x, scale_y):
        """Draw quantity text at the correct position."""
        text_position = QPoint(int(x * scale_x) + self.x_offset_s, int(y * scale_y) + self.y_offset_s - 5)
        painter.drawText(text_position, str(quantity))
    
    def update_standard_protein_quantities(self):
        self.standard_protein_values.text()
    
    def move_tab(self,tab):
        self.tab_widget.setCurrentIndex(tab)
        
    def save_state(self):
        """Save the current state of the image, markers, and other relevant data."""
        state = {
            "image": self.image.copy() if self.image else None,
            "left_markers": self.left_markers.copy(),
            "right_markers": self.right_markers.copy(),
            "top_markers": self.top_markers.copy(),
            "custom_markers": self.custom_markers.copy() if hasattr(self, "custom_markers") else [],
            "image_before_padding": self.image_before_padding.copy() if self.image_before_padding else None,
            "image_contrasted": self.image_contrasted.copy() if self.image_contrasted else None,
            "image_before_contrast": self.image_before_contrast.copy() if self.image_before_contrast else None,
            "font_family": self.font_family,
            "font_size": self.font_size,
            "font_color": self.font_color,
            "font_rotation": self.font_rotation,
            "left_marker_shift_added": self.left_marker_shift_added,
            "right_marker_shift_added": self.right_marker_shift_added,
            "top_marker_shift_added": self.top_marker_shift_added,
            "quantities_peak_area_dict": self.quantities_peak_area_dict.copy(),
            "quantities": self.quantities.copy(),
            "protein_quantities": self.protein_quantities.copy(),
            "standard_protein_areas": self.standard_protein_areas.copy(),
        }
        self.undo_stack.append(state)
        self.redo_stack.clear()
    
    def undo_action(self):
        """Undo the last action by restoring the previous state."""
        if self.undo_stack:
            # Save the current state to the redo stack
            current_state = {
                "image": self.image.copy() if self.image else None,
                "left_markers": self.left_markers.copy(),
                "right_markers": self.right_markers.copy(),
                "top_markers": self.top_markers.copy(),
                "custom_markers": self.custom_markers.copy() if hasattr(self, "custom_markers") else [],
                "image_before_padding": self.image_before_padding.copy() if self.image_before_padding else None,
                "image_contrasted": self.image_contrasted.copy() if self.image_contrasted else None,
                "image_before_contrast": self.image_before_contrast.copy() if self.image_before_contrast else None,
                "font_family": self.font_family,
                "font_size": self.font_size,
                "font_color": self.font_color,
                "font_rotation": self.font_rotation,
                "left_marker_shift_added": self.left_marker_shift_added,
                "right_marker_shift_added": self.right_marker_shift_added,
                "top_marker_shift_added": self.top_marker_shift_added,
                "quantities_peak_area_dict": self.quantities_peak_area_dict.copy(),
                "quantities": self.quantities.copy(),
                "protein_quantities": self.protein_quantities.copy(),
                "standard_protein_areas": self.standard_protein_areas.copy(),
            }
            self.redo_stack.append(current_state)
            
            # Restore the previous state from the undo stack
            previous_state = self.undo_stack.pop()
            self.image = previous_state["image"]
            self.left_markers = previous_state["left_markers"]
            self.right_markers = previous_state["right_markers"]
            self.top_markers = previous_state["top_markers"]
            self.custom_markers = previous_state["custom_markers"]
            self.image_before_padding = previous_state["image_before_padding"]
            self.image_contrasted = previous_state["image_contrasted"]
            self.image_before_contrast = previous_state["image_before_contrast"]
            self.font_family = previous_state["font_family"]
            self.font_size = previous_state["font_size"]
            self.font_color = previous_state["font_color"]
            self.font_rotation = previous_state["font_rotation"]
            self.left_marker_shift_added = previous_state["left_marker_shift_added"]
            self.right_marker_shift_added = previous_state["right_marker_shift_added"]
            self.top_marker_shift_added = previous_state["top_marker_shift_added"]
            self.quantities_peak_area_dict = previous_state["quantities_peak_area_dict"]
            self.quantities = previous_state["quantities"]
            self.protein_quantities = previous_state["protein_quantities"]
            self.standard_protein_areas = previous_state["standard_protein_areas"]
            try:
                w=self.image.width()
                h=self.image.height()
                # Preview window
                ratio=w/h
                self.label_width=int(self.screen_width * 0.28)
                label_height=int(self.label_width/ratio)
                if label_height>self.label_width:
                    label_height=self.label_width
                    self.label_width=ratio*label_height
                self.live_view_label.setFixedSize(int(self.label_width), int(label_height))
            except:
                pass
            
            self.update_live_view()
            
    
    def redo_action(self):
        """Redo the last undone action by restoring the next state."""
        if self.redo_stack:
            # Save the current state to the undo stack
            current_state = {
                "image": self.image.copy() if self.image else None,
                "left_markers": self.left_markers.copy(),
                "right_markers": self.right_markers.copy(),
                "top_markers": self.top_markers.copy(),
                "custom_markers": self.custom_markers.copy() if hasattr(self, "custom_markers") else [],
                "image_before_padding": self.image_before_padding.copy() if self.image_before_padding else None,
                "image_contrasted": self.image_contrasted.copy() if self.image_contrasted else None,
                "image_before_contrast": self.image_before_contrast.copy() if self.image_before_contrast else None,
                "font_family": self.font_family,
                "font_size": self.font_size,
                "font_color": self.font_color,
                "font_rotation": self.font_rotation,
                "left_marker_shift_added": self.left_marker_shift_added,
                "right_marker_shift_added": self.right_marker_shift_added,
                "top_marker_shift_added": self.top_marker_shift_added,
                "quantities_peak_area_dict": self.quantities_peak_area_dict.copy(),
                "quantities": self.quantities.copy(),
                "protein_quantities": self.protein_quantities.copy(),
                "standard_protein_areas": self.standard_protein_areas.copy(),
            }
            self.undo_stack.append(current_state)
            
            # Restore the next state from the redo stack
            next_state = self.redo_stack.pop()
            self.image = next_state["image"]
            self.left_markers = next_state["left_markers"]
            self.right_markers = next_state["right_markers"]
            self.top_markers = next_state["top_markers"]
            self.custom_markers = next_state["custom_markers"]
            self.image_before_padding = next_state["image_before_padding"]
            self.image_contrasted = next_state["image_contrasted"]
            self.image_before_contrast = next_state["image_before_contrast"]
            self.font_family = next_state["font_family"]
            self.font_size = next_state["font_size"]
            self.font_color = next_state["font_color"]
            self.font_rotation = next_state["font_rotation"]
            self.left_marker_shift_added = next_state["left_marker_shift_added"]
            self.right_marker_shift_added = next_state["right_marker_shift_added"]
            self.top_marker_shift_added = next_state["top_marker_shift_added"]
            self.quantities_peak_area_dict = next_state["quantities_peak_area_dict"]
            self.quantities = next_state["quantities"]
            self.protein_quantities = next_state["protein_quantities"]
            self.standard_protein_areas = next_state["standard_protein_areas"]
            try:
                w=self.image.width()
                h=self.image.height()
                # Preview window
                ratio=w/h
                self.label_width=int(self.screen_width * 0.28)
                label_height=int(self.label_width/ratio)
                if label_height>self.label_width:
                    label_height=self.label_width
                    self.label_width=ratio*label_height
                self.live_view_label.setFixedSize(int(self.label_width), int(label_height))
            except:
                pass
            
            self.update_live_view()
            
    def analysis_tab(self):
        tab = QWidget()
        layout = QVBoxLayout(tab)
        layout.setSpacing(15) # Increase spacing in this tab

        # --- Molecular Weight Prediction ---
        mw_group = QGroupBox("Molecular Weight Prediction")
        mw_group.setStyleSheet("QGroupBox { font-weight: bold; }")
        mw_layout = QVBoxLayout(mw_group)
        mw_layout.setSpacing(8)

        self.predict_button = QPushButton("Predict Molecular Weight")
        self.predict_button.setToolTip("Predicts size based on labeled MWM lane.\nClick marker positions first, then click this button, then click the target band.\nShortcut: Ctrl+P / Cmd+P")
        self.predict_button.setEnabled(False)  # Initially disabled
        self.predict_button.clicked.connect(self.predict_molecular_weight)
        mw_layout.addWidget(self.predict_button)

        layout.addWidget(mw_group)

        # --- Peak Area / Sample Quantification ---
        quant_group = QGroupBox("Peak Area and Sample Quantification")
        quant_group.setStyleSheet("QGroupBox { font-weight: bold; }")
        quant_layout = QVBoxLayout(quant_group)
        quant_layout.setSpacing(8)

        # Area Definition Buttons
        area_def_layout=QHBoxLayout()
        self.btn_define_quad = QPushButton("Define Quad Area")
        self.btn_define_quad.setToolTip(
            "Click 4 corner points to define a region. \n"
            "Use for skewed lanes. The area will be perspective-warped (straightened) before analysis. \n"
            "Results may differ from Rectangle selection due to warping."
        )
        self.btn_define_quad.clicked.connect(self.enable_quad_mode)
        self.btn_define_rec = QPushButton("Define Rectangle Area")
        self.btn_define_rec.setToolTip(
            "Click and drag to define a rectangular region. \n"
            "Use for lanes that are already straight or when simple profile analysis is sufficient. \n"
            "Does not correct for skew."
        )
        self.btn_define_rec.clicked.connect(self.enable_rectangle_mode)
        self.btn_sel_rec = QPushButton("Move Selected Area")
        self.btn_sel_rec.setToolTip("Click and drag the selected Quad or Rectangle to move it.")
        self.btn_sel_rec.clicked.connect(self.enable_move_selection_mode)
        area_def_layout.addWidget(self.btn_define_quad)
        area_def_layout.addWidget(self.btn_define_rec)
        area_def_layout.addWidget(self.btn_sel_rec)
        quant_layout.addLayout(area_def_layout)

        # Standard Processing
        std_proc_layout = QHBoxLayout()
        self.btn_process_std = QPushButton("Process Standard Bands")
        self.btn_process_std.setToolTip("Analyze the defined area as a standard lane.\nYou will be prompted for the known quantity.")
        self.btn_process_std.clicked.connect(self.process_standard)
        std_proc_layout.addWidget(self.btn_process_std)
        quant_layout.addLayout(std_proc_layout)

        # Standard Info Display (Read-only makes more sense)
        std_info_layout = QGridLayout()
        std_info_layout.addWidget(QLabel("Std. Quantities:"), 0, 0)
        self.standard_protein_values = QLineEdit()
        self.standard_protein_values.setPlaceholderText("Known quantities (auto-populated)")
        self.standard_protein_values.setReadOnly(True) # Make read-only
        std_info_layout.addWidget(self.standard_protein_values, 0, 1)

        std_info_layout.addWidget(QLabel("Std. Areas:"), 1, 0)
        self.standard_protein_areas_text = QLineEdit()
        self.standard_protein_areas_text.setPlaceholderText("Calculated total areas (auto-populated)")
        self.standard_protein_areas_text.setReadOnly(True) # Make read-only
        std_info_layout.addWidget(self.standard_protein_areas_text, 1, 1)
        quant_layout.addLayout(std_info_layout)

        quant_layout.addWidget(self.create_separator()) # Add visual separator

        # Sample Processing
        sample_proc_layout = QHBoxLayout()
        self.btn_analyze_sample = QPushButton("Analyze Sample Bands")
        self.btn_analyze_sample.setToolTip("Analyze the defined area as a sample lane using the standard curve.")
        self.btn_analyze_sample.clicked.connect(self.process_sample)
        sample_proc_layout.addWidget(self.btn_analyze_sample)
        quant_layout.addLayout(sample_proc_layout)


        # Sample Info Display
        sample_info_layout = QGridLayout()
        sample_info_layout.addWidget(QLabel("Sample Areas:"), 0, 0)
        self.target_protein_areas_text = QLineEdit()
        self.target_protein_areas_text.setPlaceholderText("Calculated peak areas (auto-populated)")
        self.target_protein_areas_text.setReadOnly(True)
        sample_info_layout.addWidget(self.target_protein_areas_text, 0, 1)

        # Add Table Export button next to sample areas
        self.table_export_button = QPushButton("Export Results Table")
        self.table_export_button.setToolTip("Export the analysis results (areas, percentages, quantities) to Excel.")
        self.table_export_button.clicked.connect(self.open_table_window)
        sample_info_layout.addWidget(self.table_export_button, 1, 0, 1, 2) # Span button across columns
        quant_layout.addLayout(sample_info_layout)

        layout.addWidget(quant_group)

        # --- Clear Button ---
        clear_layout = QHBoxLayout() # Layout to center button maybe
        clear_layout.addStretch()
        self.clear_predict_button = QPushButton("Clear Analysis Markers")
        self.clear_predict_button.setToolTip("Clears MW prediction line and analysis regions.\nShortcut: Ctrl+Shift+P / Cmd+Shift+P")
        self.clear_predict_button.clicked.connect(self.clear_predict_molecular_weight)
        layout.addWidget(self.clear_predict_button)
        clear_layout.addStretch()
        layout.addLayout(clear_layout)


        layout.addStretch()
        return tab
    
    def enable_move_selection_mode(self):
        """Enable mode to move the selected quadrilateral or rectangle."""
        self.live_view_label.mode = "move"
        self.live_view_label.setCursor(Qt.SizeAllCursor)  # Change cursor to indicate move mode
        self.live_view_label.mousePressEvent = self.start_move_selection
        self.live_view_label.mouseReleaseEvent = self.end_move_selection
        self.update_live_view()
        
    def start_move_selection(self, event):
        """Start moving the selection when the mouse is pressed."""
        if self.live_view_label.mode == "move":
            self.live_view_label.draw_edges=False
            # Check if the mouse is near the quadrilateral or rectangle
            if self.live_view_label.quad_points:
                self.live_view_label.selected_point = self.get_nearest_point(event.pos(), self.live_view_label.quad_points)
            elif self.live_view_label.bounding_box_preview:
                self.live_view_label.selected_point = self.get_nearest_point(event.pos(), [
                    QPointF(self.live_view_label.bounding_box_preview[0], self.live_view_label.bounding_box_preview[1])
                ])
            self.live_view_label.drag_start_pos = event.pos()
            self.update_live_view()
        self.live_view_label.mouseMoveEvent = self.move_selection
    
    def move_selection(self, event):
        """Move the selection while the mouse is being dragged."""
        if self.live_view_label.mode == "move" and self.live_view_label.selected_point is not None:
            # Calculate the offset
            offset = event.pos() - self.live_view_label.drag_start_pos
            self.live_view_label.drag_start_pos = event.pos()
            
            # Move the quadrilateral or rectangle
            if self.live_view_label.quad_points:
                for i in range(len(self.live_view_label.quad_points)):
                    self.live_view_label.quad_points[i] += offset
            elif self.live_view_label.bounding_box_preview:
                self.live_view_label.bounding_box_preview = (
                    self.live_view_label.bounding_box_preview[0] + offset.x(),
                    self.live_view_label.bounding_box_preview[1] + offset.y(),
                    self.live_view_label.bounding_box_preview[2] + offset.x(),
                    self.live_view_label.bounding_box_preview[3] + offset.y(),
                )
            self.update_live_view()
    
    def end_move_selection(self, event):
        """End moving the selection when the mouse is released."""
        if self.live_view_label.mode == "move":
            self.live_view_label.selected_point = -1            
            self.update_live_view()
            self.live_view_label.draw_edges=True
        self.live_view_label.mouseMoveEvent = None
        
        
             
    def get_nearest_point(self, mouse_pos, points):
        """Get the nearest point to the mouse position."""
        min_distance = float('inf')
        nearest_point = None
        for point in points:
            distance = (mouse_pos - point).manhattanLength()
            if distance < min_distance:
                min_distance = distance
                nearest_point = point
        return nearest_point
    
    def open_table_window(self):
        # Use the latest calculated peak areas
        peak_areas_to_export = self.latest_peak_areas
        calculated_quantities_to_export = self.latest_calculated_quantities

        if not peak_areas_to_export:
            QMessageBox.information(self, "No Data", "No analysis results available to export.")
            return

        standard_dict = self.quantities_peak_area_dict
        # Determine if quantities should be calculated/displayed based on standards
        standard_flag = len(standard_dict) >= 2

        # Open the table window with the latest data
        # Pass the pre-calculated quantities if available
        self.table_window = TableWindow(
            peak_areas_to_export,
            standard_dict,
            standard_flag,
            calculated_quantities_to_export, # Pass the calculated quantities
            self
        )
        self.table_window.show()
    
    def enable_quad_mode(self):
        """Enable mode to define a quadrilateral area."""
        self.live_view_label.bounding_box_preview = []
        self.live_view_label.quad_points = []
        self.live_view_label.bounding_box_complete = False
        self.live_view_label.measure_quantity_mode = True
        self.live_view_label.mode = "quad"
        self.live_view_label.setCursor(Qt.CrossCursor)
        
        # # Reset mouse event handlers
        self.live_view_label.mousePressEvent = None
        self.live_view_label.mouseMoveEvent = None
        self.live_view_label.mouseReleaseEvent = None
        
        # Update the live view
        self.update_live_view()
    
    def enable_rectangle_mode(self):
        """Enable mode to define a rectangle area."""
        self.live_view_label.bounding_box_preview = []
        self.live_view_label.quad_points = []
        self.live_view_label.bounding_box_complete = False
        self.live_view_label.measure_quantity_mode = True
        self.live_view_label.mode = "rectangle"
        self.live_view_label.rectangle_points = []  # Clear previous rectangle points
        self.live_view_label.rectangle_start = None  # Reset rectangle start
        self.live_view_label.rectangle_end = None    # Reset rectangle end
        self.live_view_label.bounding_box_preview = None  # Reset bounding box preview
        self.live_view_label.setCursor(Qt.CrossCursor)
        
        # Set mouse event handlers for rectangle mode
        self.live_view_label.mousePressEvent = self.start_rectangle
        self.live_view_label.mouseMoveEvent = self.update_rectangle_preview
        self.live_view_label.mouseReleaseEvent = self.finalize_rectangle
        
        # Update the live view
        self.update_live_view()

    
    def start_rectangle(self, event):
        """Record the start position of the rectangle."""
        if self.live_view_label.mode == "rectangle":
            self.live_view_label.rectangle_start = self.live_view_label.transform_point(event.pos())
            self.live_view_label.rectangle_points = [self.live_view_label.rectangle_start]
            self.live_view_label.bounding_box_preview = None  # Reset bounding box preview
            self.update_live_view()
    
    def update_rectangle_preview(self, event):
        """Update the rectangle preview as the mouse moves."""
        if self.live_view_label.mode == "rectangle" and self.live_view_label.rectangle_start:
            self.live_view_label.rectangle_end = self.live_view_label.transform_point(event.pos())
            self.live_view_label.rectangle_points = [self.live_view_label.rectangle_start, self.live_view_label.rectangle_end]
            
            # Update bounding_box_preview with the rectangle coordinates
            self.live_view_label.bounding_box_preview = (
                self.live_view_label.rectangle_start.x(),
                self.live_view_label.rectangle_start.y(),
                self.live_view_label.rectangle_end.x(),
                self.live_view_label.rectangle_end.y(),
            )
            self.update_live_view()
    
    def finalize_rectangle(self, event):
        """Finalize the rectangle when the mouse is released."""
        if self.live_view_label.mode == "rectangle" and self.live_view_label.rectangle_start:
            self.live_view_label.rectangle_end = self.live_view_label.transform_point(event.pos())
            self.live_view_label.rectangle_points = [self.live_view_label.rectangle_start, self.live_view_label.rectangle_end]
            
            # Save the final bounding box preview
            self.live_view_label.bounding_box_preview = (
                self.live_view_label.rectangle_start.x(),
                self.live_view_label.rectangle_start.y(),
                self.live_view_label.rectangle_end.x(),
                self.live_view_label.rectangle_end.y(),
            )
            
            self.live_view_label.mode = None  # Exit rectangle mode
            self.live_view_label.setCursor(Qt.ArrowCursor)
            self.update_live_view()
        
    
    def process_standard(self):
        """Processes the defined region (quad or rect) as a standard."""
        extracted_qimage = None # Will hold the QImage of the extracted region

        if len(self.live_view_label.quad_points) == 4:
            # --- Quadrilateral Logic ---
            print("Processing Standard: Quadrilateral")
            # Pass the *current* self.image (could be color or gray)
            extracted_qimage = self.quadrilateral_to_rect(self.image, self.live_view_label.quad_points)
            if not extracted_qimage or extracted_qimage.isNull():
                QMessageBox.warning(self, "Error", "Quadrilateral warping failed.")
                return

        elif self.live_view_label.bounding_box_preview is not None and len(self.live_view_label.bounding_box_preview) == 4:
            # --- Rectangle Logic ---
            print("Processing Standard: Rectangle")
            try:
                # (Coordinate Transformation Logic - KEEP AS IS from previous fix)
                start_x_view, start_y_view, end_x_view, end_y_view = self.live_view_label.bounding_box_preview
                # ... (rest of coordinate transformation logic) ...
                zoom = self.live_view_label.zoom_level
                offset_x, offset_y = self.live_view_label.pan_offset.x(), self.live_view_label.pan_offset.y()
                start_x_unzoomed = (start_x_view - offset_x) / zoom
                start_y_unzoomed = (start_y_view - offset_y) / zoom
                end_x_unzoomed = (end_x_view - offset_x) / zoom
                end_y_unzoomed = (end_y_view - offset_y) / zoom
                if not self.image or self.image.isNull(): raise ValueError("Base image invalid.")
                img_w, img_h = self.image.width(), self.image.height()
                label_w, label_h = self.live_view_label.width(), self.live_view_label.height()
                scale_factor = min(label_w / img_w, label_h / img_h) if img_w > 0 and img_h > 0 else 1
                display_offset_x = (label_w - img_w * scale_factor) / 2
                display_offset_y = (label_h - img_h * scale_factor) / 2
                start_x_img = (start_x_unzoomed - display_offset_x) / scale_factor
                start_y_img = (start_y_unzoomed - display_offset_y) / scale_factor
                end_x_img = (end_x_unzoomed - display_offset_x) / scale_factor
                end_y_img = (end_y_unzoomed - display_offset_y) / scale_factor
                x, y = int(min(start_x_img, end_x_img)), int(min(start_y_img, end_y_img))
                w, h = int(abs(end_x_img - start_x_img)), int(abs(end_y_img - start_y_img))
                if w <= 0 or h <= 0: raise ValueError(f"Invalid calculated rectangle dimensions (w={w}, h={h}).")
                x_clamped, y_clamped = max(0, x), max(0, y)
                w_clamped, h_clamped = max(0, min(w, img_w - x_clamped)), max(0, min(h, img_h - y_clamped))
                if w_clamped <= 0 or h_clamped <= 0: raise ValueError(f"Clamped rectangle dimensions invalid (w={w_clamped}, h={h_clamped}).")
                # --- End Coordinate Transformation ---

                extracted_qimage = self.image.copy(x_clamped, y_clamped, w_clamped, h_clamped)

                if extracted_qimage.isNull():
                    raise ValueError("QImage.copy failed for rectangle.")

            except Exception as e:
                 print(f"Error processing rectangle region for standard: {e}")
                 QMessageBox.warning(self, "Error", "Could not process rectangular region.")
                 return
        else:
            QMessageBox.warning(self, "Input Error", "Please define a Quadrilateral or Rectangle area first.")
            return

        # --- Convert extracted region to Grayscale PIL for analysis ---
        if extracted_qimage and not extracted_qimage.isNull():
            processed_data_pil = self.convert_qimage_to_grayscale_pil(extracted_qimage)
            if processed_data_pil:
                # Call analyze_bounding_box with the Grayscale PIL image
                self.analyze_bounding_box(processed_data_pil, standard=True)
            else:
                QMessageBox.warning(self, "Error", "Could not convert extracted region to grayscale for analysis.")
        # No else needed, errors handled above
    
    def process_sample(self):
        """Processes the defined region (quad or rect) as a sample."""
        extracted_qimage = None # Will hold the QImage of the extracted region

        if len(self.live_view_label.quad_points) == 4:
            # --- Quadrilateral Logic ---
            print("Processing Sample: Quadrilateral")
            extracted_qimage = self.quadrilateral_to_rect(self.image, self.live_view_label.quad_points)
            if not extracted_qimage or extracted_qimage.isNull():
                QMessageBox.warning(self, "Error", "Quadrilateral warping failed.")
                return

        elif self.live_view_label.bounding_box_preview is not None and len(self.live_view_label.bounding_box_preview) == 4:
            # --- Rectangle Logic ---
            print("Processing Sample: Rectangle")
            try:
                # (Coordinate Transformation Logic - KEEP AS IS)
                start_x_view, start_y_view, end_x_view, end_y_view = self.live_view_label.bounding_box_preview
                # ... (rest of coordinate transformation logic) ...
                zoom = self.live_view_label.zoom_level
                offset_x, offset_y = self.live_view_label.pan_offset.x(), self.live_view_label.pan_offset.y()
                start_x_unzoomed = (start_x_view - offset_x) / zoom
                start_y_unzoomed = (start_y_view - offset_y) / zoom
                end_x_unzoomed = (end_x_view - offset_x) / zoom
                end_y_unzoomed = (end_y_view - offset_y) / zoom
                if not self.image or self.image.isNull(): raise ValueError("Base image invalid.")
                img_w, img_h = self.image.width(), self.image.height()
                label_w, label_h = self.live_view_label.width(), self.live_view_label.height()
                scale_factor = min(label_w / img_w, label_h / img_h) if img_w > 0 and img_h > 0 else 1
                display_offset_x = (label_w - img_w * scale_factor) / 2
                display_offset_y = (label_h - img_h * scale_factor) / 2
                start_x_img = (start_x_unzoomed - display_offset_x) / scale_factor
                start_y_img = (start_y_unzoomed - display_offset_y) / scale_factor
                end_x_img = (end_x_unzoomed - display_offset_x) / scale_factor
                end_y_img = (end_y_unzoomed - display_offset_y) / scale_factor
                x, y = int(min(start_x_img, end_x_img)), int(min(start_y_img, end_y_img))
                w, h = int(abs(end_x_img - start_x_img)), int(abs(end_y_img - start_y_img))
                if w <= 0 or h <= 0: raise ValueError(f"Invalid calculated rectangle dimensions (w={w}, h={h}).")
                x_clamped, y_clamped = max(0, x), max(0, y)
                w_clamped, h_clamped = max(0, min(w, img_w - x_clamped)), max(0, min(h, img_h - y_clamped))
                if w_clamped <= 0 or h_clamped <= 0: raise ValueError(f"Clamped rectangle dimensions invalid (w={w_clamped}, h={h_clamped}).")
                # --- End Coordinate Transformation ---

                extracted_qimage = self.image.copy(x_clamped, y_clamped, w_clamped, h_clamped)

                if extracted_qimage.isNull():
                    raise ValueError("QImage.copy failed for rectangle.")

            except Exception as e:
                 print(f"Error processing rectangle region for sample: {e}")
                 QMessageBox.warning(self, "Error", "Could not process rectangular region.")
                 return
        else:
            QMessageBox.warning(self, "Input Error", "Please define a Quadrilateral or Rectangle area first.")
            return

        # --- Convert extracted region to Grayscale PIL for analysis ---
        if extracted_qimage and not extracted_qimage.isNull():
            processed_data_pil = self.convert_qimage_to_grayscale_pil(extracted_qimage)
            if processed_data_pil:
                # Call analyze_bounding_box with the Grayscale PIL image
                self.analyze_bounding_box(processed_data_pil, standard=False)
            else:
                QMessageBox.warning(self, "Error", "Could not convert extracted region to grayscale for analysis.")


    def convert_qimage_to_grayscale_pil(self, qimg):
        """
        Converts a QImage (any format) to a suitable Grayscale PIL Image ('L' or 'I;16')
        for use with PeakAreaDialog. Returns None on failure.
        """
        if not qimg or qimg.isNull():
            return None

        fmt = qimg.format()
        pil_img = None

        try:
            # Already grayscale? Convert directly if possible.
            if fmt == QImage.Format_Grayscale16:
                print("Converting Grayscale16 QImage to PIL 'I;16'")
                np_array = self.qimage_to_numpy(qimg)
                if np_array is not None and np_array.dtype == np.uint16:
                    try: pil_img = Image.fromarray(np_array, mode='I;16')
                    except ValueError: pil_img = Image.fromarray(np_array, mode='I')
                else: raise ValueError("Failed NumPy conversion for Grayscale16")
            elif fmt == QImage.Format_Grayscale8:
                print("Converting Grayscale8 QImage to PIL 'L'")
                # Try direct conversion first
                try:
                    pil_img = ImageQt.fromqimage(qimg).convert('L')
                    if pil_img is None: raise ValueError("Direct QImage->PIL(L) failed.")
                except Exception as e_direct:
                    print(f"Direct QImage->PIL(L) failed ({e_direct}), falling back via NumPy.")
                    np_array = self.qimage_to_numpy(qimg)
                    if np_array is not None and np_array.dtype == np.uint8:
                        pil_img = Image.fromarray(np_array, mode='L')
                    else: raise ValueError("Failed NumPy conversion for Grayscale8")
            else: # Color or other format
                print(f"Converting format {fmt} QImage to Grayscale PIL")
                # Use NumPy for robust conversion to 16-bit grayscale intermediate
                np_img = self.qimage_to_numpy(qimg)
                if np_img is None: raise ValueError("NumPy conversion failed for color.")
                if np_img.ndim == 3:
                    gray_np = cv2.cvtColor(np_img[...,:3], cv2.COLOR_BGR2GRAY) # Assume BGR/BGRA input
                    # Convert to 16-bit PIL
                    gray_np_16bit = (gray_np / 255.0 * 65535.0).astype(np.uint16)
                    try: pil_img = Image.fromarray(gray_np_16bit, mode='I;16')
                    except ValueError: pil_img = Image.fromarray(gray_np_16bit, mode='I')
                elif np_img.ndim == 2: # Should have been caught by Grayscale checks, but handle anyway
                     if np_img.dtype == np.uint16:
                         try: pil_img = Image.fromarray(np_img, mode='I;16')
                         except ValueError: pil_img = Image.fromarray(np_img, mode='I')
                     else: # Assume uint8 or other, convert to L
                         pil_img = Image.fromarray(np_img).convert('L')
                else:
                     raise ValueError(f"Unsupported array dimensions: {np_img.ndim}")

            if pil_img is None:
                raise ValueError("PIL Image creation failed.")

            print(f"  Successfully converted to PIL Image: mode={pil_img.mode}, size={pil_img.size}")
            return pil_img

        except Exception as e:
            print(f"Error in convert_qimage_to_grayscale_pil: {e}")
            traceback.print_exc()
            return None            
        
        
            
    def combine_image_tab(self):
        tab = QWidget()
        layout = QVBoxLayout(tab)
        layout.setSpacing(10)

        # Calculate initial slider ranges based on current view size
        render_scale=3
        # Use a sensible default or calculate from screen if view isn't ready
        initial_width = self.live_view_label.width() if self.live_view_label.width() > 0 else 500
        initial_height = self.live_view_label.height() if self.live_view_label.height() > 0 else 500
        render_width = initial_width * render_scale
        render_height = initial_height * render_scale

        # --- Image 1 Group ---
        image1_group = QGroupBox("Image 1 Overlay")
        image1_group.setStyleSheet("QGroupBox { font-weight: bold; }")
        image1_layout = QGridLayout(image1_group) # Use grid for better control layout
        image1_layout.setSpacing(8)

        # Load/Place/Remove buttons
        copy_image1_button = QPushButton("Copy Current Image")
        copy_image1_button.setToolTip("Copies the main image to the Image 1 buffer.")
        copy_image1_button.clicked.connect(self.save_image1) # Keep original name save_image1
        place_image1_button = QPushButton("Place Image 1")
        place_image1_button.setToolTip("Positions Image 1 based on sliders.")
        place_image1_button.clicked.connect(self.place_image1)
        remove_image1_button = QPushButton("Remove Image 1")
        remove_image1_button.setToolTip("Removes Image 1 from the overlay.")
        remove_image1_button.clicked.connect(self.remove_image1)

        image1_layout.addWidget(copy_image1_button, 0, 0)
        image1_layout.addWidget(place_image1_button, 0, 1)
        image1_layout.addWidget(remove_image1_button, 0, 2)

        # Position Sliders
        image1_layout.addWidget(QLabel("Horizontal Pos:"), 1, 0)
        self.image1_left_slider = QSlider(Qt.Horizontal)
        self.image1_left_slider.setRange(-render_width, render_width) # Wider range
        self.image1_left_slider.setValue(0)
        self.image1_left_slider.valueChanged.connect(self.update_live_view)
        image1_layout.addWidget(self.image1_left_slider, 1, 1, 1, 2) # Span 2 columns

        image1_layout.addWidget(QLabel("Vertical Pos:"), 2, 0)
        self.image1_top_slider = QSlider(Qt.Horizontal)
        self.image1_top_slider.setRange(-render_height, render_height) # Wider range
        self.image1_top_slider.setValue(0)
        self.image1_top_slider.valueChanged.connect(self.update_live_view)
        image1_layout.addWidget(self.image1_top_slider, 2, 1, 1, 2) # Span 2 columns

        # Resize Slider
        image1_layout.addWidget(QLabel("Resize (%):"), 3, 0)
        self.image1_resize_slider = QSlider(Qt.Horizontal)
        self.image1_resize_slider.setRange(10, 300)  # Range 10% to 300%
        self.image1_resize_slider.setValue(100)
        self.image1_resize_slider.valueChanged.connect(self.update_live_view)
        self.image1_resize_label = QLabel("100%") # Show current percentage
        self.image1_resize_slider.valueChanged.connect(lambda val, lbl=self.image1_resize_label: lbl.setText(f"{val}%"))
        image1_layout.addWidget(self.image1_resize_slider, 3, 1)
        image1_layout.addWidget(self.image1_resize_label, 3, 2)


        layout.addWidget(image1_group)

        # --- Image 2 Group --- (Similar structure to Image 1)
        image2_group = QGroupBox("Image 2 Overlay")
        image2_group.setStyleSheet("QGroupBox { font-weight: bold; }")
        image2_layout = QGridLayout(image2_group)
        image2_layout.setSpacing(8)

        copy_image2_button = QPushButton("Copy Current Image")
        copy_image2_button.clicked.connect(self.save_image2)
        place_image2_button = QPushButton("Place Image 2")
        place_image2_button.clicked.connect(self.place_image2)
        remove_image2_button = QPushButton("Remove Image 2")
        remove_image2_button.clicked.connect(self.remove_image2)
        image2_layout.addWidget(copy_image2_button, 0, 0)
        image2_layout.addWidget(place_image2_button, 0, 1)
        image2_layout.addWidget(remove_image2_button, 0, 2)

        image2_layout.addWidget(QLabel("Horizontal Pos:"), 1, 0)
        self.image2_left_slider = QSlider(Qt.Horizontal)
        self.image2_left_slider.setRange(-render_width, render_width)
        self.image2_left_slider.setValue(0)
        self.image2_left_slider.valueChanged.connect(self.update_live_view)
        image2_layout.addWidget(self.image2_left_slider, 1, 1, 1, 2)

        image2_layout.addWidget(QLabel("Vertical Pos:"), 2, 0)
        self.image2_top_slider = QSlider(Qt.Horizontal)
        self.image2_top_slider.setRange(-render_height, render_height)
        self.image2_top_slider.setValue(0)
        self.image2_top_slider.valueChanged.connect(self.update_live_view)
        image2_layout.addWidget(self.image2_top_slider, 2, 1, 1, 2)

        image2_layout.addWidget(QLabel("Resize (%):"), 3, 0)
        self.image2_resize_slider = QSlider(Qt.Horizontal)
        self.image2_resize_slider.setRange(10, 300)
        self.image2_resize_slider.setValue(100)
        self.image2_resize_slider.valueChanged.connect(self.update_live_view)
        self.image2_resize_label = QLabel("100%")
        self.image2_resize_slider.valueChanged.connect(lambda val, lbl=self.image2_resize_label: lbl.setText(f"{val}%"))
        image2_layout.addWidget(self.image2_resize_slider, 3, 1)
        image2_layout.addWidget(self.image2_resize_label, 3, 2)

        layout.addWidget(image2_group)

        # --- Finalize Button ---
        finalize_layout = QHBoxLayout()
        finalize_layout.addStretch()
        finalize_button = QPushButton("Rasterize Image")
        finalize_button.setToolTip("Permanently merges the placed overlays onto the main image. This action cannot be undone easily.")
        finalize_button.clicked.connect(self.finalize_combined_image)
        layout.addWidget(finalize_button)
        finalize_layout.addStretch()
        layout.addLayout(finalize_layout)

        layout.addStretch() # Push content up
        return tab
    
    def remove_image1(self):
        """Remove Image 1 and reset its sliders."""
        if hasattr(self, 'image1'):
            # del self.image1
            # del self.image1_original
            try:
                del self.image1_position
            except:
                pass
            self.image1_left_slider.setValue(0)
            self.image1_top_slider.setValue(0)
            self.image1_resize_slider.setValue(100)
            self.update_live_view()
    
    def remove_image2(self):
        """Remove Image 2 and reset its sliders."""
        if hasattr(self, 'image2'):
            # del self.image2
            # del self.image2_original
            try:
                del self.image2_position
            except:
                pass
            self.image2_left_slider.setValue(0)
            self.image2_top_slider.setValue(0)
            self.image2_resize_slider.setValue(100)
            self.update_live_view()
    
    def save_image1(self):
        if self.image:
            self.image1 = self.image.copy()
            self.image1_original = self.image1.copy()  # Save the original image for resizing
            QMessageBox.information(self, "Success", "Image 1 copied.")
    
    def save_image2(self):
        if self.image:
            self.image2 = self.image.copy()
            self.image2_original = self.image2.copy()  # Save the original image for resizing
            QMessageBox.information(self, "Success", "Image 2 copied.")
    
    def place_image1(self):
        if hasattr(self, 'image1'):
            self.image1_position = (self.image1_left_slider.value(), self.image1_top_slider.value())
            self.update_live_view()
    
    def place_image2(self):
        if hasattr(self, 'image2'):
            self.image2_position = (self.image2_left_slider.value(), self.image2_top_slider.value())
            self.update_live_view()
    
    # def update_combined_image(self):
    #     if not hasattr(self, 'image1') and not hasattr(self, 'image2'):
    #         return
    
    #     # Create a copy of the original image to avoid modifying it
    #     combined_image = self.image.copy()
    
    #     painter = QPainter(combined_image)
    
    #     # Draw Image 1 if it exists
    #     if hasattr(self, 'image1') and hasattr(self, 'image1_position'):
    #         # Resize Image 1 based on the slider value
    #         scale_factor = self.image1_resize_slider.value() / 100.0
    #         width = int(self.image1_original.width() * scale_factor)
    #         height = int(self.image1_original.height() * scale_factor)
    #         resized_image1 = self.image1_original.scaled(width, height, Qt.KeepAspectRatio, Qt.SmoothTransformation)
    #         painter.drawImage(self.image1_position[0], self.image1_position[1], resized_image1)
    
    #     # Draw Image 2 if it exists
    #     if hasattr(self, 'image2') and hasattr(self, 'image2_position'):
    #         # Resize Image 2 based on the slider value
    #         scale_factor = self.image2_resize_slider.value() / 100.0
    #         width = int(self.image2_original.width() * scale_factor)
    #         height = int(self.image2_original.height() * scale_factor)
    #         resized_image2 = self.image2_original.scaled(width, height, Qt.KeepAspectRatio, Qt.SmoothTransformation)
    #         painter.drawImage(self.image2_position[0], self.image2_position[1], resized_image2)
    
    #     painter.end()
    
    #     # Update the live view with the combined image
    #     self.live_view_label.setPixmap(QPixmap.fromImage(combined_image))
    
    def finalize_combined_image(self):
        """Overlap image1 and image2 on top of self.image and save the result as self.image."""
        # if not hasattr(self, 'image1') and not hasattr(self, 'image2'):
        #     QMessageBox.warning(self, "Warning", "No images to overlap.")
        #     return
        
        # Define cropping boundaries
        x_start_percent = self.crop_x_start_slider.value() / 100
        y_start_percent = self.crop_y_start_slider.value() / 100
        x_start = int(self.image.width() * x_start_percent)
        y_start = int(self.image.height() * y_start_percent)
        # Create a high-resolution canvas for the final image
        render_scale = 3  # Scale factor for rendering resolution
        render_width = self.live_view_label.width() * render_scale
        render_height = self.live_view_label.height() * render_scale
    
        # Create a blank high-resolution image with a white background
        high_res_canvas = QImage(render_width, render_height, QImage.Format_RGB888)
        high_res_canvas.fill(Qt.white)
    
        # Create a QPainter to draw on the high-resolution canvas
        painter = QPainter(high_res_canvas)
    
        # Draw the base image (self.image) onto the high-resolution canvas
        scaled_base_image = self.image.scaled(render_width, render_height, Qt.KeepAspectRatio, Qt.SmoothTransformation)
        painter.drawImage(0, 0, scaled_base_image)
    
        # Draw Image 1 if it exists
        if hasattr(self, 'image1') and hasattr(self, 'image1_position'):
            scale_factor = self.image1_resize_slider.value() / 100.0
            width = int(self.image1_original.width() * scale_factor)
            height = int(self.image1_original.height() * scale_factor)
            resized_image1 = self.image1_original.scaled(width, height, Qt.KeepAspectRatio, Qt.SmoothTransformation)
            painter.drawImage(
                int(self.image1_position[0] + self.x_offset_s),
                int(self.image1_position[1] + self.y_offset_s),
                resized_image1
            )
    
        # Draw Image 2 if it exists
        if hasattr(self, 'image2') and hasattr(self, 'image2_position'):
            scale_factor = self.image2_resize_slider.value() / 100.0
            width = int(self.image2_original.width() * scale_factor)
            height = int(self.image2_original.height() * scale_factor)
            resized_image2 = self.image2_original.scaled(width, height, Qt.KeepAspectRatio, Qt.SmoothTransformation)
            painter.drawImage(
                int(self.image2_position[0] + self.x_offset_s),
                int(self.image2_position[1] + self.y_offset_s),
                resized_image2
            )
    
        # End painting
        painter.end()
        
        
        self.render_image_on_canvas(
                high_res_canvas, scaled_base_image, x_start, y_start, render_scale, draw_guides=False
            )
        
        self.image=high_res_canvas.copy()
        self.image_before_padding=self.image.copy()
        self.image_before_contrast=self.image.copy()
        self.update_live_view()
    
        # Remove Image 1 and Image 2 after finalizing
        self.remove_image1()
        self.remove_image2()
        
        self.left_markers.clear()  # Clear left markers
        self.right_markers.clear()  # Clear right markers
        self.top_markers.clear()
        self.custom_markers.clear()
        self.remove_custom_marker_mode()
        self.clear_predict_molecular_weight()
    
        QMessageBox.information(self, "Success", "The images have been overlapped and saved in memory.")
    
        
    


    def font_and_image_tab(self):
        tab = QWidget()
        layout = QVBoxLayout(tab)
        layout.setSpacing(15) # Add spacing between groups

        # --- Font Options Group ---
        font_options_group = QGroupBox("Marker and Label Font") # Renamed for clarity
        font_options_group.setStyleSheet("QGroupBox { font-weight: bold; }")
        font_options_layout = QGridLayout(font_options_group)
        font_options_layout.setSpacing(8)

        # Font type
        font_type_label = QLabel("Font Family:")
        self.font_combo_box = QFontComboBox()
        self.font_combo_box.setEditable(False)
        self.font_combo_box.setCurrentFont(QFont(self.font_family)) # Use initialized value
        self.font_combo_box.currentFontChanged.connect(self.update_font) # Connect signal
        font_options_layout.addWidget(font_type_label, 0, 0)
        font_options_layout.addWidget(self.font_combo_box, 0, 1, 1, 2) # Span 2 columns

        # Font size
        font_size_label = QLabel("Font Size:")
        self.font_size_spinner = QSpinBox()
        self.font_size_spinner.setRange(6, 72)  # Adjusted range
        self.font_size_spinner.setValue(self.font_size) # Use initialized value
        self.font_size_spinner.valueChanged.connect(self.update_font) # Connect signal
        font_options_layout.addWidget(font_size_label, 1, 0)
        font_options_layout.addWidget(self.font_size_spinner, 1, 1)

        # Font color
        self.font_color_button = QPushButton("Font Color")
        self.font_color_button.setToolTip("Select color for Left, Right, Top markers.")
        self.font_color_button.clicked.connect(self.select_font_color)
        self._update_color_button_style(self.font_color_button, self.font_color) # Set initial button color
        font_options_layout.addWidget(self.font_color_button, 1, 2)

        # Font rotation (Top/Bottom)
        font_rotation_label = QLabel("Top Label Rotation:") # Specific label
        self.font_rotation_input = QSpinBox()
        self.font_rotation_input.setRange(-180, 180)
        self.font_rotation_input.setValue(self.font_rotation) # Use initialized value
        self.font_rotation_input.setSuffix(" °") # Add degree symbol
        self.font_rotation_input.valueChanged.connect(self.update_font) # Connect signal
        font_options_layout.addWidget(font_rotation_label, 2, 0)
        font_options_layout.addWidget(self.font_rotation_input, 2, 1, 1, 2) # Span 2 columns

        layout.addWidget(font_options_group)


        # --- Image Adjustments Group ---
        img_adjust_group = QGroupBox("Image Adjustments")
        img_adjust_group.setStyleSheet("QGroupBox { font-weight: bold; }")
        img_adjust_layout = QGridLayout(img_adjust_group)
        img_adjust_layout.setSpacing(8)

        # Contrast Sliders with Value Labels
        high_contrast_label = QLabel("Brightness:") # Renamed for clarity
        self.high_slider = QSlider(Qt.Horizontal)
        self.high_slider.setRange(0, 200) # Range 0 to 200%
        self.high_slider.setValue(100) # Default 100%
        self.high_slider.valueChanged.connect(self.update_image_contrast)
        self.high_value_label = QLabel("1.00") # Display factor
        self.high_slider.valueChanged.connect(lambda val, lbl=self.high_value_label: lbl.setText(f"{val/100.0:.2f}"))
        img_adjust_layout.addWidget(high_contrast_label, 0, 0)
        img_adjust_layout.addWidget(self.high_slider, 0, 1)
        img_adjust_layout.addWidget(self.high_value_label, 0, 2)

        low_contrast_label = QLabel("Contrast:") # Renamed for clarity
        self.low_slider = QSlider(Qt.Horizontal)
        self.low_slider.setRange(0, 200) # Range 0 to 100%
        self.low_slider.setValue(100) # Default 100%
        self.low_slider.valueChanged.connect(self.update_image_contrast)
        self.low_value_label = QLabel("1.00") # Display factor
        self.low_slider.valueChanged.connect(lambda val, lbl=self.low_value_label: lbl.setText(f"{val/100.0:.2f}"))
        img_adjust_layout.addWidget(low_contrast_label, 1, 0)
        img_adjust_layout.addWidget(self.low_slider, 1, 1)
        img_adjust_layout.addWidget(self.low_value_label, 1, 2)


        # Gamma Adjustment Slider
        gamma_label = QLabel("Gamma:")
        self.gamma_slider = QSlider(Qt.Horizontal)
        self.gamma_slider.setRange(10, 300)  # Range 0.1 to 3.0
        self.gamma_slider.setValue(100)  # Default 1.0 (100%)
        self.gamma_slider.valueChanged.connect(self.update_image_gamma)
        self.gamma_value_label = QLabel("1.00") # Display factor
        self.gamma_slider.valueChanged.connect(lambda val, lbl=self.gamma_value_label: lbl.setText(f"{val/100.0:.2f}"))
        img_adjust_layout.addWidget(gamma_label, 2, 0)
        img_adjust_layout.addWidget(self.gamma_slider, 2, 1)
        img_adjust_layout.addWidget(self.gamma_value_label, 2, 2)


        # Separator
        img_adjust_layout.addWidget(self.create_separator(), 3, 0, 1, 3) # Span across columns

        # Action Buttons
        btn_layout = QHBoxLayout()
        self.bw_button = QPushButton("Grayscale")
        self.bw_button.setToolTip("Convert the image to grayscale.\nShortcut: Ctrl+B / Cmd+B")
        self.bw_button.clicked.connect(self.convert_to_black_and_white)
        invert_button = QPushButton("Invert")
        invert_button.setToolTip("Invert image colors.\nShortcut: Ctrl+I / Cmd+I")
        invert_button.clicked.connect(self.invert_image)
        reset_button = QPushButton("Reset Adjustments")
        reset_button.setToolTip("Reset Brightness, Contrast, and Gamma sliders to default.\n(Does not undo Grayscale or Invert)")
        reset_button.clicked.connect(self.reset_gamma_contrast)
        btn_layout.addWidget(self.bw_button)
        btn_layout.addWidget(invert_button)
        btn_layout.addStretch() # Push reset button to the right
        btn_layout.addWidget(reset_button)

        img_adjust_layout.addLayout(btn_layout, 4, 0, 1, 3) # Add button layout

        layout.addWidget(img_adjust_group)
        layout.addStretch() # Push groups up
        return tab

    
    def _update_color_button_style(self, button, color):
        """ Helper to update button background color preview """
        if color.isValid():
            button.setStyleSheet(f"QPushButton {{ background-color: {color.name()}; color: {'black' if color.lightness() > 128 else 'white'}; }}")
        else:
            button.setStyleSheet("") # Reset to default stylesheet
    
    def create_separator(self):
        """Creates a horizontal separator line."""
        line = QFrame()
        line.setFrameShape(QFrame.HLine)
        line.setFrameShadow(QFrame.Sunken)
        return line
    
    
    def create_cropping_tab(self):
        """Create the Cropping tab."""
        tab = QWidget()
        layout = QVBoxLayout(tab)
    
        # Group Box for Alignment Options
        alignment_params_group = QGroupBox("Alignment Options")
        alignment_params_group.setStyleSheet("QGroupBox { font-weight: bold; }")
        alignment_layout = QVBoxLayout()
    
        
        #Guide Lines
        self.show_guides_label = QLabel("Show Guide Lines")
        self.show_guides_checkbox = QCheckBox("", self)
        self.show_guides_checkbox.setChecked(False)
        self.show_guides_checkbox.stateChanged.connect(self.update_live_view)
        
        # Rotation Angle 
        rotation_layout = QHBoxLayout()
        self.orientation_label = QLabel("Rotation Angle (0.00°)")
        self.orientation_label.setFixedWidth(150)
        self.orientation_slider = QSlider(Qt.Horizontal)
        self.orientation_slider.setRange(-3600, 3600)  # Scale by 10 to allow decimals
        self.orientation_slider.setValue(0)
        self.orientation_slider.setSingleStep(1)
        self.orientation_slider.valueChanged.connect(self.update_live_view)
        # Align Button
        self.align_button = QPushButton("Apply Rotation")
        self.align_button.clicked.connect(self.align_image)
        
        self.reset_align_button = QPushButton("Reset Rotation")
        self.reset_align_button.clicked.connect(self.reset_align_image)
        
        rotation_layout.addWidget(self.show_guides_label)
        rotation_layout.addWidget(self.show_guides_checkbox)
        rotation_layout.addWidget(self.orientation_label)
        rotation_layout.addWidget(self.orientation_slider)
        rotation_layout.addWidget(self.align_button)
        rotation_layout.addWidget(self.reset_align_button)      
        
        
    
        
        
        # Flip Vertical Button
        self.flip_vertical_button = QPushButton("Flip Vertical")
        self.flip_vertical_button.clicked.connect(self.flip_vertical)
    
        # Flip Horizontal Button
        self.flip_horizontal_button = QPushButton("Flip Horizontal")
        self.flip_horizontal_button.clicked.connect(self.flip_horizontal)
    
        alignment_layout.addLayout(rotation_layout)
        
        alignment_layout.addWidget(self.flip_vertical_button)  
        alignment_layout.addWidget(self.flip_horizontal_button)  
        alignment_params_group.setLayout(alignment_layout)
        
        
       # Add Tapering Skew Fix Group
        taper_skew_group = QGroupBox("Skew Fix")
        taper_skew_group.setStyleSheet("QGroupBox { font-weight: bold; }")
        taper_skew_layout = QHBoxLayout()
    
        # Taper Skew Slider
        self.taper_skew_label = QLabel("Tapering Skew (0.00)")
        self.taper_skew_label.setFixedWidth(150)
        self.taper_skew_slider = QSlider(Qt.Horizontal)
        self.taper_skew_slider.setRange(-70, 70)  # Adjust as needed
        self.taper_skew_slider.setValue(0)
        self.taper_skew_slider.valueChanged.connect(self.update_live_view)
        
        # Align Button
        self.skew_button = QPushButton("Apply Skew")
        self.skew_button.clicked.connect(self.update_skew)
    
        # Add widgets to taper skew layout
        taper_skew_layout.addWidget(self.taper_skew_label)
        taper_skew_layout.addWidget(self.taper_skew_slider)
        taper_skew_layout.addWidget(self.skew_button)
        taper_skew_group.setLayout(taper_skew_layout)
        
        # Group Box for Cropping Options
        cropping_params_group = QGroupBox("Cropping Options")
        cropping_params_group.setStyleSheet("QGroupBox { font-weight: bold; }")
        cropping_layout = QVBoxLayout()
    
        # Cropping Sliders
        crop_slider_layout = QGridLayout()
    
        crop_x_start_label = QLabel("Crop Left (%)")
        self.crop_x_start_slider = QSlider(Qt.Horizontal)
        self.crop_x_start_slider.setRange(0, 100)
        self.crop_x_start_slider.setValue(0)
        self.crop_x_start_slider.valueChanged.connect(self.update_live_view)
        crop_slider_layout.addWidget(crop_x_start_label, 0, 0)
        crop_slider_layout.addWidget(self.crop_x_start_slider, 0, 1)
    
        crop_x_end_label = QLabel("Crop Right (%)")
        self.crop_x_end_slider = QSlider(Qt.Horizontal)
        self.crop_x_end_slider.setRange(0, 100)
        self.crop_x_end_slider.setValue(100)
        self.crop_x_end_slider.valueChanged.connect(self.update_live_view)
        crop_slider_layout.addWidget(crop_x_end_label, 0, 2)
        crop_slider_layout.addWidget(self.crop_x_end_slider, 0, 3)
    
        crop_y_start_label = QLabel("Crop Top (%)")
        self.crop_y_start_slider = QSlider(Qt.Horizontal)
        self.crop_y_start_slider.setRange(0, 100)
        self.crop_y_start_slider.setValue(0)
        self.crop_y_start_slider.valueChanged.connect(self.update_live_view)
        crop_slider_layout.addWidget(crop_y_start_label, 1, 0)
        crop_slider_layout.addWidget(self.crop_y_start_slider, 1, 1)
    
        crop_y_end_label = QLabel("Crop Bottom (%)")
        self.crop_y_end_slider = QSlider(Qt.Horizontal)
        self.crop_y_end_slider.setRange(0, 100)
        self.crop_y_end_slider.setValue(100)
        self.crop_y_end_slider.valueChanged.connect(self.update_live_view)
        crop_slider_layout.addWidget(crop_y_end_label, 1, 2)
        crop_slider_layout.addWidget(self.crop_y_end_slider, 1, 3)
    
        cropping_layout.addLayout(crop_slider_layout)
    
        # Crop Update Button
        crop_button = QPushButton("Update Crop")
        crop_button.clicked.connect(self.update_crop)
        cropping_layout.addWidget(crop_button)
    
        cropping_params_group.setLayout(cropping_layout)
    
        # Add both group boxes to the main layout
        layout.addWidget(alignment_params_group)
        layout.addWidget(cropping_params_group)
        layout.addWidget(taper_skew_group)
        layout.addStretch()
    
        return tab
    
    def create_white_space_tab(self):
        tab = QWidget()
        layout = QVBoxLayout(tab)
        layout.setSpacing(15)

        # --- Padding Group ---
        padding_params_group = QGroupBox("Add White Space (Padding)")
        padding_params_group.setStyleSheet("QGroupBox { font-weight: bold; }")
        padding_layout = QGridLayout(padding_params_group)
        padding_layout.setSpacing(8)


        # Input fields with validation (optional but good)
        from PyQt5.QtGui import QIntValidator
        int_validator = QIntValidator(0, 5000, self) # Allow padding up to 5000px

        # Left Padding
        left_padding_label = QLabel("Left Padding (px):")
        self.left_padding_input = QLineEdit("100") # Default
        self.left_padding_input.setValidator(int_validator)
        self.left_padding_input.setToolTip("Pixels to add to the left.")
        padding_layout.addWidget(left_padding_label, 0, 0)
        padding_layout.addWidget(self.left_padding_input, 0, 1)

        # Right Padding
        right_padding_label = QLabel("Right Padding (px):")
        self.right_padding_input = QLineEdit("100") # Default
        self.right_padding_input.setValidator(int_validator)
        self.right_padding_input.setToolTip("Pixels to add to the right.")
        padding_layout.addWidget(right_padding_label, 0, 2)
        padding_layout.addWidget(self.right_padding_input, 0, 3)

        # Top Padding
        top_padding_label = QLabel("Top Padding (px):")
        self.top_padding_input = QLineEdit("100") # Default
        self.top_padding_input.setValidator(int_validator)
        self.top_padding_input.setToolTip("Pixels to add to the top.")
        padding_layout.addWidget(top_padding_label, 1, 0)
        padding_layout.addWidget(self.top_padding_input, 1, 1)

        # Bottom Padding
        bottom_padding_label = QLabel("Bottom Padding (px):")
        self.bottom_padding_input = QLineEdit("0") # Default
        self.bottom_padding_input.setValidator(int_validator)
        self.bottom_padding_input.setToolTip("Pixels to add to the bottom.")
        padding_layout.addWidget(bottom_padding_label, 1, 2)
        padding_layout.addWidget(self.bottom_padding_input, 1, 3)

        layout.addWidget(padding_params_group)

        # --- Buttons Layout ---
        button_layout = QHBoxLayout()
        self.recommend_button = QPushButton("Set Recommended Values")
        self.recommend_button.setToolTip("Auto-fill padding values based on image size (approx. 10-15%).")
        self.recommend_button.clicked.connect(self.recommended_values)

        self.clear_padding_button = QPushButton("Clear Values")
        self.clear_padding_button.setToolTip("Set all padding values to 0.")
        self.clear_padding_button.clicked.connect(self.clear_padding_values)

        self.finalize_button = QPushButton("Apply Padding")
        self.finalize_button.setToolTip("Permanently add the specified padding to the image.")
        self.finalize_button.clicked.connect(self.finalize_image)

        button_layout.addWidget(self.recommend_button)
        button_layout.addWidget(self.clear_padding_button)
        button_layout.addStretch()
        button_layout.addWidget(self.finalize_button)

        layout.addLayout(button_layout)
        layout.addStretch() # Push content up

        return tab
    def clear_padding_values(self):
        self.bottom_padding_input.setText("0")
        self.top_padding_input.setText("0")
        self.right_padding_input.setText("0")
        self.left_padding_input.setText("0")
        
    def recommended_values(self):
        render_scale = 3  # Scale factor for rendering resolution
        render_width = self.live_view_label.width() * render_scale
        render_height = self.live_view_label.height() * render_scale
        self.left_slider_range=[-100,int(render_width)+100]
        self.left_padding_slider.setRange(self.left_slider_range[0],self.left_slider_range[1])
        self.right_slider_range=[-100,int(render_width)+100]
        self.right_padding_slider.setRange(self.right_slider_range[0],self.right_slider_range[1])
        self.top_slider_range=[-100,int(render_height)+100]
        self.top_padding_slider.setRange(self.top_slider_range[0],self.top_slider_range[1])
        self.left_padding_input.setText(str(int(self.image.width()*0.1)))
        self.right_padding_input.setText(str(int(self.image.width()*0.1)))
        self.top_padding_input.setText(str(int(self.image.height()*0.15)))
        self.update_live_view()
        
        
        
    def create_markers_tab(self):
        """Create the Markers tab."""
        tab = QWidget()
        layout = QVBoxLayout(tab)
        
        marker_options_layout = QHBoxLayout()
    
        # Left/Right Marker Options
        left_right_marker_group = QGroupBox("Left/Right Marker Options")
        left_right_marker_group.setStyleSheet("QGroupBox { font-weight: bold; }")
        left_right_marker_layout = QVBoxLayout()
        
        # Combo box for marker presets or custom option
        self.combo_box = QComboBox(self)
        self.combo_box.addItems(self.marker_values_dict.keys())
        self.combo_box.addItem("Custom")
        self.combo_box.setCurrentText("Precision Plus All Blue/Unstained")
        self.combo_box.currentTextChanged.connect(self.on_combobox_changed)
        
        # Textbox to allow modification of marker values (shown when "Custom" is selected)
        self.marker_values_textbox = QLineEdit(self)
        self.marker_values_textbox.setPlaceholderText("Enter custom values as comma-separated list")
        self.marker_values_textbox.setEnabled(False)  # Disable by default
        
        # Rename input for custom option
        self.rename_input = QLineEdit(self)
        self.rename_input.setPlaceholderText("Enter new name for Custom")
        self.rename_input.setEnabled(False)
        
        # Save button for the current configuration
        self.save_button = QPushButton("Save Config", self)
        self.save_button.clicked.connect(self.save_config)
        
        # Delete button 
        self.remove_config_button = QPushButton("Remove Config", self)
        self.remove_config_button.clicked.connect(self.remove_config)
        
        # Add widgets to the Left/Right Marker Options layout
        left_right_marker_layout.addWidget(self.combo_box)
        left_right_marker_layout.addWidget(self.marker_values_textbox)
        left_right_marker_layout.addWidget(self.rename_input)
        left_right_marker_layout.addWidget(self.save_button)
        left_right_marker_layout.addWidget(self.remove_config_button)
        
        # Set layout for the Left/Right Marker Options group
        left_right_marker_group.setLayout(left_right_marker_layout)
        
        # Top Marker Options
        top_marker_group = QGroupBox("Top/Bottom Marker Options")
        top_marker_group.setStyleSheet("QGroupBox { font-weight: bold;}")
        
        # Vertical layout for top marker group
        top_marker_layout = QVBoxLayout()
        
        # Text input for Top Marker Labels (multi-column support)
        self.top_marker_input = QTextEdit(self)
        self.top_marker_input.setText(", ".join(self.top_label))  # Populate with initial values
        self.top_marker_input.setMinimumHeight(50)  # Increase height for better visibility
        self.top_marker_input.setMaximumHeight(120)
        self.top_marker_input.setPlaceholderText("Enter labels for each column, separated by commas. Use new lines for multiple columns.")
        
        # # Button to add a new column
        # self.add_column_button = QPushButton("Add Column")
        # self.add_column_button.clicked.connect(self.add_column)
        
        # # Button to remove the last column
        # self.remove_column_button = QPushButton("Remove Last Column")
        # self.remove_column_button.clicked.connect(self.remove_column)
        
        # # Button to update all labels
        self.update_top_labels_button = QPushButton("Update All Labels")
        self.update_top_labels_button.clicked.connect(self.update_all_labels)
        
        # Layout for column management buttons
        # column_buttons_layout = QHBoxLayout()
        # column_buttons_layout.addWidget(self.add_column_button)
        # column_buttons_layout.addWidget(self.remove_column_button)
        
        # Add widgets to the top marker layout
        top_marker_layout.addWidget(self.top_marker_input)
        # top_marker_layout.addLayout(column_buttons_layout)
        top_marker_layout.addWidget(self.update_top_labels_button)
        
        # Set the layout for the Top Marker Group
        top_marker_group.setLayout(top_marker_layout)
        
        # Add both groups to the horizontal layout
        marker_options_layout.addWidget(left_right_marker_group)
        marker_options_layout.addWidget(top_marker_group)
        
        # Add the horizontal layout to the main layout
        layout.addLayout(marker_options_layout)
    
        # Marker padding sliders - Group box for marker distance adjustment
        padding_params_group = QGroupBox("Marker Placement and Offsets")
        padding_params_group.setStyleSheet("QGroupBox { font-weight: bold; }")
        
        # Grid layout for the marker group box
        padding_layout = QGridLayout()
        
        # Left marker: Button, slider, reset, and duplicate in the same row
        left_marker_button = QPushButton("Place Left Markers")
        left_marker_button.setToolTip("Places the left markers at the exact location of the mouse pointer on the left. Shortcut: Ctrl+Shift+L or CMD+Shift+L")
        left_marker_button.clicked.connect(self.enable_left_marker_mode)
        self.left_padding_slider = QSlider(Qt.Horizontal)
        self.left_padding_slider.setRange(self.left_slider_range[0], self.left_slider_range[1])
        self.left_padding_slider.setValue(0)
        self.left_padding_slider.valueChanged.connect(self.update_left_padding)
        
        remove_left_button = QPushButton("Remove Last")
        remove_left_button.clicked.connect(lambda: self.reset_marker('left','remove'))
        
        reset_left_button = QPushButton("Reset")
        reset_left_button.clicked.connect(lambda: self.reset_marker('left','reset'))
        
        duplicate_left_button = QPushButton("Duplicate Right")
        duplicate_left_button.clicked.connect(lambda: self.duplicate_marker('left'))
        
        # Add left marker widgets to the grid layout
        padding_layout.addWidget(left_marker_button, 0, 0)
        padding_layout.addWidget(remove_left_button, 0, 1)
        padding_layout.addWidget(reset_left_button, 0, 2)
        padding_layout.addWidget(self.left_padding_slider, 0, 3,1,2)
        padding_layout.addWidget(duplicate_left_button, 0, 5)
        
        
        # Right marker: Button, slider, reset, and duplicate in the same row
        right_marker_button = QPushButton("Place Right Markers")
        right_marker_button.setToolTip("Places the right markers at the exact location of the mouse pointer on the right. Shortcut: Ctrl+Shift+R or CMD+Shift+R")
        right_marker_button.clicked.connect(self.enable_right_marker_mode)
        self.right_padding_slider = QSlider(Qt.Horizontal)
        self.right_padding_slider.setRange(self.right_slider_range[0], self.right_slider_range[1])
        self.right_padding_slider.setValue(0)
        self.right_padding_slider.valueChanged.connect(self.update_right_padding)
        
        remove_right_button = QPushButton("Remove Last")
        remove_right_button.clicked.connect(lambda: self.reset_marker('right','remove'))
        
        reset_right_button = QPushButton("Reset")
        reset_right_button.clicked.connect(lambda: self.reset_marker('right','reset'))
        
        duplicate_right_button = QPushButton("Duplicate Left")
        duplicate_right_button.clicked.connect(lambda: self.duplicate_marker('right'))
        
        # Add right marker widgets to the grid layout
        padding_layout.addWidget(right_marker_button, 1, 0)
        padding_layout.addWidget(remove_right_button, 1, 1)
        padding_layout.addWidget(reset_right_button, 1, 2)        
        padding_layout.addWidget(self.right_padding_slider, 1, 3,1,2)
        padding_layout.addWidget(duplicate_right_button, 1, 5)
        
        # Top marker: Button, slider, and reset in the same row
        top_marker_button = QPushButton("Place Top Markers")
        top_marker_button.setToolTip("Places the top markers at the exact location of the mouse pointer on the top. Shortcut: Ctrl+Shift+T or CMD+Shift+T")
        top_marker_button.clicked.connect(self.enable_top_marker_mode)
        self.top_padding_slider = QSlider(Qt.Horizontal)
        self.top_padding_slider.setRange(self.top_slider_range[0], self.top_slider_range[1])
        self.top_padding_slider.setValue(0)
        self.top_padding_slider.valueChanged.connect(self.update_top_padding)
        
        remove_top_button = QPushButton("Remove Last")
        remove_top_button.clicked.connect(lambda: self.reset_marker('top','remove'))
        
        reset_top_button = QPushButton("Reset")
        reset_top_button.clicked.connect(lambda: self.reset_marker('top','reset'))
        
        # Add top marker widgets to the grid layout
        padding_layout.addWidget(top_marker_button, 2, 0)
        padding_layout.addWidget(remove_top_button, 2, 1)
        padding_layout.addWidget(reset_top_button, 2, 2)
        padding_layout.addWidget(self.top_padding_slider, 2, 3, 1, 2)  # Slider spans 2 columns for better alignment
        
        for i in range(6):  # Assuming 6 columns in the grid
            padding_layout.setColumnStretch(i, 1)
        
        # Add button and QLineEdit for the custom marker
        self.custom_marker_button = QPushButton("Place Custom Marker", self)
        self.custom_marker_button.setToolTip("Places custom markers at the middle of the mouse pointer")
    
        self.custom_marker_button.clicked.connect(self.enable_custom_marker_mode)
        
        self.custom_marker_button_left_arrow = QPushButton("←", self)
        
        self.custom_marker_button_right_arrow = QPushButton("→", self)
        
        self.custom_marker_button_top_arrow = QPushButton("↑", self)
        
        self.custom_marker_button_bottom_arrow = QPushButton("↓", self)
        
        self.custom_marker_text_entry = QLineEdit(self)        
        self.custom_marker_text_entry.setPlaceholderText("Enter custom marker text")
        
        self.remove_custom_marker_button = QPushButton("Remove Last", self)
        self.remove_custom_marker_button.clicked.connect(self.remove_custom_marker_mode)
        
        self.reset_custom_marker_button = QPushButton("Reset", self)
        self.reset_custom_marker_button.clicked.connect(self.reset_custom_marker_mode)
        
        # Add color selection button for custom markers
        self.custom_marker_color_button = QPushButton("Custom Marker Color")
        self.custom_marker_color_button.clicked.connect(self.select_custom_marker_color)
        self._update_color_button_style(self.custom_marker_color_button, self.custom_marker_color)
        
        marker_buttons_layout = QHBoxLayout()
        
        # Add the arrow buttons with fixed sizes to the marker buttons layout
        self.custom_marker_button_left_arrow.setFixedSize(30, 30)
    
        marker_buttons_layout.addWidget(self.custom_marker_button_left_arrow)
        
        self.custom_marker_button_right_arrow.setFixedSize(30, 30)
        marker_buttons_layout.addWidget(self.custom_marker_button_right_arrow)
        
        self.custom_marker_button_top_arrow.setFixedSize(30, 30)
        marker_buttons_layout.addWidget(self.custom_marker_button_top_arrow)
        
        self.custom_marker_button_bottom_arrow.setFixedSize(30, 30)
        marker_buttons_layout.addWidget(self.custom_marker_button_bottom_arrow)
        
        #Assign functions to the buttons
        self.custom_marker_button_left_arrow.clicked.connect(lambda: self.arrow_marker("←"))
        self.custom_marker_button_right_arrow.clicked.connect(lambda: self.arrow_marker("→"))
        self.custom_marker_button_top_arrow.clicked.connect(lambda: self.arrow_marker("↑"))
        self.custom_marker_button_bottom_arrow.clicked.connect(lambda: self.arrow_marker("↓"))
    
        
        # Create a QWidget to hold the QHBoxLayout
        marker_buttons_widget = QWidget()
        marker_buttons_widget.setLayout(marker_buttons_layout)
        
        # Add the custom marker button
        padding_layout.addWidget(self.custom_marker_button, 3, 0)
        
        # Add the text entry for the custom marker
        padding_layout.addWidget(self.custom_marker_text_entry, 3, 1,1,1)
        
        # Add the marker buttons widget to the layout
        padding_layout.addWidget(marker_buttons_widget, 3, 2) 
        
        # Add the remove button
        padding_layout.addWidget(self.remove_custom_marker_button, 3, 3)
        
        # Add the reset button
        padding_layout.addWidget(self.reset_custom_marker_button, 3, 4)
        
        # Add the color button
        padding_layout.addWidget(self.custom_marker_color_button, 3, 5)
        
        self.custom_font_type_label = QLabel("Custom Marker Font:", self)
        self.custom_font_type_dropdown = QFontComboBox()
        self.custom_font_type_dropdown.setCurrentFont(QFont("Arial"))
        self.custom_font_type_dropdown.currentFontChanged.connect(self.update_marker_text_font)
        
        # Font size selector
        self.custom_font_size_label = QLabel("Custom Marker Size:", self)
        self.custom_font_size_spinbox = QSpinBox(self)
        self.custom_font_size_spinbox.setRange(2, 150)  # Allow font sizes from 8 to 72
        self.custom_font_size_spinbox.setValue(12)  # Default font size
        
        # Grid checkbox
        self.show_grid_checkbox = QCheckBox("Show Snap Grid", self)
        self.show_grid_checkbox.setToolTip("Places a snapping grid and the text or marker will be places at the center of the grid. Shortcut: Ctrl+Shift+G. To increase or decrease the grid size: Ctrl+Shift+Up or Down arrow or CMD+Shift+Up or Down arrow ")
        self.show_grid_checkbox.setChecked(False)  # Default: Grid is off
        self.show_grid_checkbox.stateChanged.connect(self.update_live_view)
        
        # Grid size input (optional
        self.grid_size_input = QSpinBox(self)
        self.grid_size_input.setRange(5, 100)  # Grid cell size in pixels
        self.grid_size_input.setValue(20)  # Default grid size
        self.grid_size_input.setPrefix("Grid Size (px): ")
        self.grid_size_input.valueChanged.connect(self.update_live_view)
        
        # Add font type and size widgets to the layout
        padding_layout.addWidget(self.custom_font_type_label, 4, 0,)
        padding_layout.addWidget(self.custom_font_type_dropdown, 4, 1,1,1)  # Span 2 columns
        
        padding_layout.addWidget(self.custom_font_size_label, 4, 2)
        padding_layout.addWidget(self.custom_font_size_spinbox, 4, 3,1,1)
        padding_layout.addWidget(self.show_grid_checkbox,4,4)
        padding_layout.addWidget(self.grid_size_input,4,5)
    
        # Set the layout for the marker group box
        padding_params_group.setLayout(padding_layout)
        
        # Add the font options group box to the main layout
        layout.addWidget(padding_params_group)
        
        layout.addStretch()
        return tab
    
    def add_column(self):
        """Add a new column to the top marker labels."""
        current_text = self.top_marker_input.toPlainText()
        if current_text.strip():
            self.top_marker_input.append("")  # Add a new line for a new column
        else:
            self.top_marker_input.setPlainText("")  # Start with an empty line if no text exists
    
    def remove_column(self):
        """Remove the last column from the top marker labels."""
        current_text = self.top_marker_input.toPlainText()
        lines = current_text.split("\n")
        if len(lines) > 1:
            lines.pop()  # Remove the last line
            self.top_marker_input.setPlainText("\n".join(lines))
        else:
            self.top_marker_input.clear()  # Clear the text if only one line exists

    
    def flip_vertical(self):
        self.save_state()
        """Flip the image vertically."""
        if self.image:
            self.image = self.image.mirrored(vertical=True, horizontal=False)
            self.image_before_contrast=self.image.copy()
            self.image_before_padding=self.image.copy()
            self.image_contrasted=self.image.copy()
            self.update_live_view()
    
    def flip_horizontal(self):
        self.save_state()
        """Flip the image horizontally."""
        if self.image:
            self.image = self.image.mirrored(vertical=False, horizontal=True)
            self.image_before_contrast=self.image.copy()
            self.image_before_padding=self.image.copy()
            self.image_contrasted=self.image.copy()
            self.update_live_view()
    
    def convert_to_black_and_white(self):
        """Explicitly converts the current self.image to grayscale."""
        self.save_state()
        if not self.image or self.image.isNull():
             QMessageBox.warning(self, "Error", "No image loaded.")
             return

        current_format = self.image.format()

        # Check if already grayscale
        if current_format in [QImage.Format_Grayscale8, QImage.Format_Grayscale16]:
            QMessageBox.information(self, "Info", "Image is already grayscale.")
            return

        print(f"Converting image from format {current_format} to grayscale.")

        # Preferred target format for grayscale conversion
        # Let's aim for 16-bit if the source might have high dynamic range (like color)
        # If source was already 8-bit (though format check above prevents this), stick to 8-bit.
        target_format = QImage.Format_Grayscale16 # Default target
        converted_image = None

        try:
            np_img = self.qimage_to_numpy(self.image)
            if np_img is None: raise ValueError("NumPy conversion failed.")

            if np_img.ndim == 3: # Color image
                # Use weighted average for luminance (standard RGB/BGR -> Gray)
                # OpenCV handles BGR/BGRA automatically in cvtColor
                # color_code = cv2.COLOR_BGR2GRAY if np_img.shape[2] == 3 else cv2.COLOR_BGRA2GRAY
                # gray_np = cv2.cvtColor(np_img, color_code)
                # Use Pillow for potentially better color space handling? Let's stick to OpenCV for now.
                gray_np = cv2.cvtColor(np_img[..., :3], cv2.COLOR_BGR2GRAY) # Use first 3 channels

                # Scale to target bit depth (16-bit)
                gray_np_target = (gray_np / 255.0 * 65535.0).astype(np.uint16)
                converted_image = self.numpy_to_qimage(gray_np_target)
            elif np_img.ndim == 2: # Already grayscale (but maybe different format code)
                 print("Warning: Image is NumPy 2D but not Format_Grayscale. Attempting direct conversion.")
                 # Try to convert to the target format
                 if target_format == QImage.Format_Grayscale16 and np_img.dtype == np.uint8:
                     converted_image = self.numpy_to_qimage((np_img * 257.0).astype(np.uint16)) # Scale up
                 elif target_format == QImage.Format_Grayscale8 and np_img.dtype == np.uint16:
                     converted_image = self.numpy_to_qimage((np_img / 257.0).astype(np.uint8)) # Scale down
                 else:
                     converted_image = self.numpy_to_qimage(np_img.astype(np.uint16 if target_format == QImage.Format_Grayscale16 else np.uint8)) # Just ensure dtype
            else:
                raise ValueError(f"Unsupported NumPy array dimension: {np_img.ndim}")

            if converted_image is None or converted_image.isNull():
                raise ValueError("Grayscale conversion via NumPy failed.")

        except Exception as e:
            print(f"Error during NumPy conversion for grayscale: {e}. Falling back to QImage.convertToFormat.")
            # Fallback to simple QImage conversion (likely Grayscale8)
            converted_image = self.image.convertToFormat(QImage.Format_Grayscale8) # Fallback target

        if not converted_image.isNull():
             print(f"Image converted to grayscale format: {converted_image.format()}")
             self.image = converted_image
             # Update backups consistently
             self.image_before_contrast = self.image.copy()
             self.image_contrasted = self.image.copy()
             # Reset padding state if format changes implicitly? Let's assume padding needs re-application.
             if self.image_padded:
                 print("Warning: Grayscale conversion applied after padding. Padding might need re-application.")
                 self.image_before_padding = None # Invalidate padding backup
                 self.image_padded = False
             else:
                 self.image_before_padding = self.image.copy()

             # Reset contrast/gamma sliders as appearance changed significantly
             self.reset_gamma_contrast() # Resets sliders and updates view
             self.update_live_view() # Ensure view updates even if reset_gamma_contrast fails
        else:
              QMessageBox.warning(self, "Conversion Failed", "Could not convert image to the target grayscale format.")


    def invert_image(self):
        self.save_state()
        if self.image:
            inverted_image = self.image.copy()
            inverted_image.invertPixels()
            self.image = inverted_image
            self.update_live_view()
        self.image_before_contrast=self.image.copy()
        self.image_before_padding=self.image.copy()
        self.image_contrasted=self.image.copy()

    def keyPressEvent(self, event):
        if event.key() == Qt.Key_Escape:
            self.live_view_label.setCursor(Qt.ArrowCursor)
            if self.live_view_label.preview_marker_enabled:
                self.live_view_label.preview_marker_enabled = False  # Turn off the preview
            self.live_view_label.measure_quantity_mode = False
            self.live_view_label.bounding_box_complete==False
            self.live_view_label.counter=0
            self.live_view_label.quad_points=[]
            self.live_view_label.bounding_box_preview = None
            self.live_view_label.rectangle_points = []
            self.live_view_label.mousePressEvent=None
            self.live_view_label.mode=None

        if self.live_view_label.zoom_level != 1.0:  # Only allow panning when zoomed in
            step = 20  # Adjust the panning step size as needed
            if event.key() == Qt.Key_Left:
                self.live_view_label.pan_offset.setX(self.live_view_label.pan_offset.x() - step)
            elif event.key() == Qt.Key_Right:
                self.live_view_label.pan_offset.setX(self.live_view_label.pan_offset.x() + step)
            elif event.key() == Qt.Key_Up:
                self.live_view_label.pan_offset.setY(self.live_view_label.pan_offset.y() - step)
            elif event.key() == Qt.Key_Down:
                self.live_view_label.pan_offset.setY(self.live_view_label.pan_offset.y() + step)
        self.update_live_view()
        super().keyPressEvent(event)
    
    def update_marker_text_font(self, font: QFont):
        """
        Updates the font of self.custom_marker_text_entry based on the selected font
        from self.custom_font_type_dropdown.
    
        :param font: QFont object representing the selected font from the combobox.
        """
        # Get the name of the selected font
        selected_font_name = font.family()
        
        # Set the font of the QLineEdit
        self.custom_marker_text_entry.setFont(QFont(selected_font_name))
    
    def arrow_marker(self, text: str):
        self.custom_marker_text_entry.clear()
        self.custom_marker_text_entry.setText(text)
    
    def enable_custom_marker_mode(self):
        """Enable the custom marker mode and set the mouse event."""
        self.live_view_label.setCursor(Qt.ArrowCursor)
        custom_text = self.custom_marker_text_entry.text().strip()
    
        self.live_view_label.preview_marker_enabled = True
        self.live_view_label.preview_marker_text = custom_text
        self.live_view_label.marker_font_type=self.custom_font_type_dropdown.currentText()
        self.live_view_label.marker_font_size=self.custom_font_size_spinbox.value()
        self.live_view_label.marker_color=self.custom_marker_color
        
        self.live_view_label.setFocus()
        self.live_view_label.update()
        
        self.marker_mode = "custom"  # Indicate custom marker mode
        self.live_view_label.mousePressEvent = lambda event: self.place_custom_marker(event, custom_text)
        
    def remove_custom_marker_mode(self):
        if hasattr(self, "custom_markers") and isinstance(self.custom_markers, list) and self.custom_markers:
            self.custom_markers.pop()  # Remove the last entry from the list           
        self.update_live_view()  # Update the display
        
    def reset_custom_marker_mode(self):
        if hasattr(self, "custom_markers") and isinstance(self.custom_markers, list) and self.custom_markers:
            self.custom_markers=[]  # Remove the last entry from the list           
        self.update_live_view()  # Update the display
        
    
    def place_custom_marker(self, event, custom_text):
        """Place a custom marker at the cursor location."""
        self.save_state()
        # Get cursor position from the event
        pos = event.pos()
        cursor_x, cursor_y = pos.x(), pos.y()
        
        if self.live_view_label.zoom_level != 1.0:
            cursor_x = (cursor_x - self.live_view_label.pan_offset.x()) / self.live_view_label.zoom_level
            cursor_y = (cursor_y - self.live_view_label.pan_offset.y()) / self.live_view_label.zoom_level
            
        # Adjust for snapping to grid
        if self.show_grid_checkbox.isChecked():
            grid_size = self.grid_size_input.value()
            cursor_x = round(cursor_x / grid_size) * grid_size
            cursor_y = round(cursor_y / grid_size) * grid_size
    
        # Dimensions of the displayed image
        displayed_width = self.live_view_label.width()
        displayed_height = self.live_view_label.height()
    
        # Dimensions of the actual image
        image_width = self.image.width()
        image_height = self.image.height()
    
        # Calculate scaling factors
        scale = min(displayed_width / image_width, displayed_height / image_height)
    
        # Calculate offsets
        x_offset = (displayed_width - image_width * scale) / 2
        y_offset = (displayed_height - image_height * scale) / 2
    
        # Transform cursor position to image space
        image_x = (cursor_x - x_offset) / scale
        image_y = (cursor_y - y_offset) / scale
        
            
        # Store the custom marker's position and text
        self.custom_markers = getattr(self, "custom_markers", [])
        self.custom_markers.append((image_x, image_y, custom_text, self.custom_marker_color, self.custom_font_type_dropdown.currentText(), self.custom_font_size_spinbox.value()))
        # print("CUSTOM_MARKER: ",self.custom_markers)
        # Update the live view to render the custom marker
        self.update_live_view()
        
    def select_custom_marker_color(self):
        """Open a color picker dialog to select the color for custom markers."""
        color = QColorDialog.getColor(self.custom_marker_color, self, "Select Custom Marker Color")
        if color.isValid():
            self.custom_marker_color = color  # Update the custom marker color
        self._update_color_button_style(self.custom_marker_color_button, self.custom_marker_color)
    
    def update_all_labels(self):
        """Update all labels from the QTextEdit, supporting multiple columns."""
        self.marker_values=[int(num) if num.strip().isdigit() else num.strip() for num in self.marker_values_textbox.text().strip("[]").split(",")]
        input_text = self.top_marker_input.toPlainText()
        
        # Split the text into a list by commas and strip whitespace
        self.top_label= [label.strip() for label in input_text.split(",") if label.strip()]
        

        # if len(self.top_label) < len(self.top_markers):
        #     self.top_markers = self.top_markers[:len(self.top_label)]
        for i in range(0, len(self.top_markers)):
            try:
                self.top_markers[i] = (self.top_markers[i][0], self.top_label[i])
            except:
                self.top_markers[i] = (self.top_markers[i][0], str(""))

        # if len(self.marker_values) < len(self.left_markers):
        #     self.left_markers = self.left_markers[:len(self.marker_values)]
        for i in range(0, len(self.left_markers)):
            try:
                self.left_markers[i] = (self.left_markers[i][0], self.marker_values[i])
            except:
                self.left_markers[i] = (self.left_markers[i][0], str(""))
                

        # if len(self.marker_values) < len(self.right_markers):
        #     self.right_markers = self.right_markers[:len(self.marker_values)]
        for i in range(0, len(self.right_markers)):
            try:
                self.right_markers[i] = (self.right_markers[i][0], self.marker_values[i])
            except:
                self.right_markers[i] = (self.right_markers[i][0], str(""))
                

        
        # Trigger a refresh of the live view
        self.update_live_view()
        
    def reset_marker(self, marker_type, param):       
        if marker_type == 'left':
            if param == 'remove' and len(self.left_markers)!=0:
                self.left_markers.pop()  
                if self.current_left_marker_index > 0:
                    self.current_left_marker_index -= 1
            elif param == 'reset':
                self.left_markers.clear()
                self.current_left_marker_index = 0  

             
        elif marker_type == 'right' and len(self.right_markers)!=0:
            if param == 'remove':
                self.right_markers.pop()  
                if self.current_right_marker_index > 0:
                    self.current_right_marker_index -= 1
            elif param == 'reset':
                self.right_markers.clear()
                self.current_right_marker_index = 0

        elif marker_type == 'top' and len(self.top_markers)!=0:
            if param == 'remove':
                self.top_markers.pop() 
                if self.current_top_label_index > 0:
                    self.current_top_label_index -= 1
            elif param == 'reset':
                self.top_markers.clear()
                self.current_top_label_index = 0
    
        # Call update live view after resetting markers
        self.update_live_view()
        
    def duplicate_marker(self, marker_type):
        if marker_type == 'left' and self.right_markers:
            self.left_markers = self.right_markers.copy()  # Copy right markers to left
        elif marker_type == 'right' and self.left_markers:
            self.right_markers = self.left_markers.copy()  # Copy left markers to right

            # self.right_padding = self.image_width*0.9
    
        # Call update live view after duplicating markers
        self.update_live_view()
        
    def on_combobox_changed(self):
        
        text=self.combo_box.currentText()
        """Handle the ComboBox change event to update marker values."""
        if text == "Custom":
            # Enable the textbox for custom values when "Custom" is selected
            self.marker_values_textbox.setEnabled(True)
            self.rename_input.setEnabled(True)
            
            # self.marker_values_textbox.setText()
        else:
            # Update marker values based on the preset option
            self.marker_values = self.marker_values_dict.get(text, [])
            self.marker_values_textbox.setEnabled(False)  # Disable the textbox
            self.rename_input.setEnabled(False)
            self.top_label = self.top_label_dict.get(text, [])
            self.top_label = [str(item) if not isinstance(item, str) else item for item in self.top_label]
            self.top_marker_input.setText(", ".join(self.top_label))
            try:
                
                # Ensure that the top_markers list only updates the top_label values serially
                for i in range(0, len(self.top_markers)):
                    self.top_markers[i] = (self.top_markers[i][0], self.top_label[i])
                
                # # If self.top_label has more entries than current top_markers, add them
                # if len(self.top_label) > len(self.top_markers):
                #     additional_markers = [(self.top_markers[-1][0] + 50 * (i + 1), label) 
                #                           for i, label in enumerate(self.top_label[len(self.top_markers):])]
                #     self.top_markers.extend(additional_markers)
                
                # If self.top_label has fewer entries, truncate the list
                if len(self.top_label) < len(self.top_markers):
                    self.top_markers = self.top_markers[:len(self.top_label)]
                    
                
                
            except:
                pass
            try:
                self.marker_values_textbox.setText(str(self.marker_values_dict[self.combo_box.currentText()]))
            except:
                pass

    # Functions for updating contrast and gamma
    
    def reset_gamma_contrast(self):
        try:
            if self.image_before_contrast==None:
                self.image_before_contrast=self.image_master.copy()
            self.image_contrasted = self.image_before_contrast.copy()  # Update the contrasted image
            self.image_before_padding = self.image_before_contrast.copy()  # Ensure padding resets use the correct base
            self.high_slider.setValue(100)  # Reset contrast to default
            self.low_slider.setValue(100)  # Reset contrast to default
            self.gamma_slider.setValue(100)  # Reset gamma to default
            self.update_live_view()
        except:
            pass

    
    def update_image_contrast(self):
        try:
            if self.contrast_applied==False:
                self.image_before_contrast=self.image.copy()
                self.contrast_applied=True
            
            if self.image:
                high_contrast_factor = self.high_slider.value() / 100.0
                low_contrast_factor = self.low_slider.value() / 100.0
                gamma_factor = self.gamma_slider.value() / 100.0
                self.image = self.apply_contrast_gamma(self.image_contrasted, high_contrast_factor, low_contrast_factor, gamma=gamma_factor)  
                self.update_live_view()
        except:
            pass
    
    def update_image_gamma(self):
        try:
            if self.contrast_applied==False:
                self.image_before_contrast=self.image.copy()
                self.contrast_applied=True
                
            if self.image:
                high_contrast_factor = self.high_slider.value() / 100.0
                low_contrast_factor = self.low_slider.value() / 100.0
                gamma_factor = self.gamma_slider.value() / 100.0
                self.image = self.apply_contrast_gamma(self.image_contrasted, high_contrast_factor, low_contrast_factor, gamma=gamma_factor)            
                self.update_live_view()
        except:
            pass
    
    def apply_contrast_gamma(self, qimage, high_factor, low_factor, gamma):
        """
        Applies brightness (high), contrast (low), and gamma adjustments to a QImage,
        preserving the original format (including color and bit depth) where possible.
        Uses NumPy/OpenCV for calculations. Applies adjustments independently to color channels.
        """
        if not qimage or qimage.isNull():
            return qimage

        original_format = qimage.format()
        try:
            img_array = self.qimage_to_numpy(qimage)
            if img_array is None: raise ValueError("NumPy conversion failed.")

            # Work with float64 for calculations to avoid precision issues
            img_array_float = img_array.astype(np.float64)

            # Determine max value based on original data type
            if img_array.dtype == np.uint16:
                max_val = 65535.0
            elif img_array.dtype == np.uint8:
                max_val = 255.0
            else: # Default for unexpected types (e.g., float input?)
                 max_val = np.max(img_array_float) if np.any(img_array_float) else 1.0
                 print(f"Warning: Unknown dtype {img_array.dtype}, using max_val={max_val}")

            # --- Apply adjustments ---
            if img_array.ndim == 3: # Color Image (e.g., RGB, RGBA, BGR, BGRA)
                num_channels = img_array.shape[2]
                adjusted_channels = []

                # Process only the color channels (first 3 usually)
                channels_to_process = min(num_channels, 3)
                for i in range(channels_to_process):
                    channel = img_array_float[:, :, i]
                    # Normalize to 0-1 range
                    channel_norm = channel / max_val

                    # Apply brightness (high_factor): Multiply
                    channel_norm = channel_norm * high_factor

                    # Apply contrast (low_factor): Scale difference from mid-grey (0.5)
                    mid_grey = 0.5
                    contrast_factor = max(0.01, low_factor) # Prevent zero/negative
                    channel_norm = mid_grey + contrast_factor * (channel_norm - mid_grey)

                    # Clip to 0-1 range after contrast/brightness
                    channel_norm = np.clip(channel_norm, 0.0, 1.0)

                    # Apply gamma correction
                    safe_gamma = max(0.01, gamma)
                    channel_norm = np.power(channel_norm, safe_gamma)

                    # Clip again after gamma
                    channel_norm_clipped = np.clip(channel_norm, 0.0, 1.0)

                    # Scale back to original range
                    adjusted_channels.append(channel_norm_clipped * max_val)

                # Reconstruct the image array
                img_array_final_float = np.stack(adjusted_channels, axis=2)

                # Keep the alpha channel (if present) untouched
                if num_channels == 4:
                    alpha_channel = img_array_float[:, :, 3] # Get original alpha
                    img_array_final_float = np.dstack((img_array_final_float, alpha_channel))

            elif img_array.ndim == 2: # Grayscale Image
                print("  Processing Grayscale Image")
                # Normalize to 0-1 range
                img_array_norm = img_array_float / max_val

                # Apply brightness
                img_array_norm = img_array_norm * high_factor

                # Apply contrast
                mid_grey = 0.5
                contrast_factor = max(0.01, low_factor)
                img_array_norm = mid_grey + contrast_factor * (img_array_norm - mid_grey)

                # Clip
                img_array_norm = np.clip(img_array_norm, 0.0, 1.0)

                # Apply gamma
                safe_gamma = max(0.01, gamma)
                img_array_norm = np.power(img_array_norm, safe_gamma)

                # Clip again
                img_array_norm_clipped = np.clip(img_array_norm, 0.0, 1.0)

                # Scale back
                img_array_final_float = img_array_norm_clipped * max_val
            else:
                print(f"Warning: Unsupported array dimension {img_array.ndim} in apply_contrast_gamma.")
                return qimage # Return original if unsupported dimensions

            # Convert back to original data type
            img_array_final = img_array_final_float.astype(img_array.dtype)

            # Convert back to QImage using the helper function
            result_qimage = self.numpy_to_qimage(img_array_final)
            if result_qimage.isNull():
                raise ValueError("Conversion back to QImage failed.")

            # numpy_to_qimage should infer the correct format (e.g., ARGB32 for 4 channels)
            return result_qimage

        except Exception as e:
            traceback.print_exc() # Print detailed traceback
            return qimage # Return original QImage on error
    

    def save_contrast_options(self):
        if self.image:
            self.image_contrasted = self.image.copy()  # Save the current image as the contrasted image
            self.image_before_padding = self.image.copy()  # Ensure the pre-padding state is also updated
        else:
            QMessageBox.warning(self, "Error", "No image is loaded to save contrast options.")

    def remove_config(self):
        try:
            # Get the currently selected marker label
            selected_marker = self.combo_box.currentText()
    
            # Ensure the selected marker is not "Custom" before deleting
            if selected_marker == "Custom":
                QMessageBox.warning(self, "Error", "Cannot remove the 'Custom' marker.")
                return
            
            elif selected_marker == "Precision Plus All Blue/Unstained":
                QMessageBox.warning(self, "Error", "Cannot remove the 'Inbuilt' marker.")
                return
            
            elif selected_marker == "1 kB Plus":
                QMessageBox.warning(self, "Error", "Cannot remove the 'Inbuilt' marker.")
                return
    
            # Remove the marker label and top_label if they exist
            if selected_marker in self.marker_values_dict:
                del self.marker_values_dict[selected_marker]
            if selected_marker in self.top_label_dict:
                del self.top_label_dict[selected_marker]
    
            # Save the updated configuration
            with open("Imaging_assistant_config.txt", "w") as f:
                config = {
                    "marker_values": self.marker_values_dict,
                    "top_label": self.top_label_dict
                }
                json.dump(config, f)
    
            # Remove from the ComboBox and reset UI
            self.combo_box.removeItem(self.combo_box.currentIndex())
            self.top_marker_input.clear()
    
            QMessageBox.information(self, "Success", f"Configuration '{selected_marker}' removed.")
    
        except Exception as e:
            QMessageBox.warning(self, "Error", f"Error removing config: {e}")

    def save_config(self):
        """Rename the 'Custom' marker values option and save the configuration."""
        new_name = self.rename_input.text().strip()
        
        # Ensure that the correct dictionary (self.marker_values_dict) is being used
        if self.rename_input.text() != "Enter new name for Custom" and self.rename_input.text() != "":  # Correct condition
            # Save marker values list
            self.marker_values_dict[new_name] = [int(num) if num.strip().isdigit() else num.strip() for num in self.marker_values_textbox.text().strip("[]").split(",")]
            
            # Save top_label list under the new_name key
            self.top_label_dict[new_name] = [int(num) if num.strip().isdigit() else num.strip() for num in self.top_marker_input.toPlainText().strip("[]").split(",")]

            try:
                # Save both the marker values and top label (under new_name) to the config file
                with open("Imaging_assistant_config.txt", "w") as f:
                    config = {
                        "marker_values": self.marker_values_dict,
                        "top_label": self.top_label_dict  # Save top_label_dict as a dictionary with new_name as key
                    }
                    json.dump(config, f)  # Save both marker_values and top_label_dict
            except Exception as e:
                print(f"Error saving config: {e}")

        self.combo_box.setCurrentText(new_name)
        self.load_config()  # Reload the configuration after saving
    
    
    def load_config(self):
        """Load the configuration from the file."""
        try:
            with open("Imaging_assistant_config.txt", "r") as f:
                config = json.load(f)
                
                # Load marker values and top label from the file
                self.marker_values_dict = config.get("marker_values", {})
                
                # Retrieve top_label list from the dictionary using the new_name key
                new_name = self.rename_input.text().strip()  # Assuming `new_name` is defined here; otherwise, set it manually
                self.top_label_dict = config.get("top_label", {})  # Default if not found
                # self.top_label = self.top_label_dict.get(new_name, ["MWM", "S1", "S2", "S3", "S4", "S5", "S6", "S7", "S8", "S9", "MWM"])  # Default if not found
                
        except FileNotFoundError:
            # Set default marker values and top_label if config is not found
            self.marker_values_dict = {
                "Precision Plus All Blue/Unstained": [250, 150, 100, 75, 50, 37, 25, 20, 15, 10],
                "1 kB Plus": [15000, 10000, 8000, 7000, 6000, 5000, 4000, 3000, 2000, 1500, 1000, 850, 650, 500, 400, 300, 200, 100],
            }
            self.top_label_dict = {
                "Precision Plus All Blue/Unstained": ["MWM", "S1", "S2", "S3", "S4", "S5", "S6", "S7", "S8", "S9", "MWM"],
                "1 kB Plus": ["MWM", "S1", "S2", "S3", "S4", "S5", "S6", "S7", "S8", "S9", "MWM"],
            }  # Default top labels
            self.top_label = self.top_label_dict.get("Precision Plus All Blue/Unstained", ["MWM", "S1", "S2", "S3", "S4", "S5", "S6", "S7", "S8", "S9", "MWM"])
        except Exception as e:
            print(f"Error loading config: {e}")
            # Fallback to default values if there is an error
            self.marker_values_dict = {
                "Precision Plus All Blue/Unstained": [250, 150, 100, 75, 50, 37, 25, 20, 15, 10],
                "1 kB Plus": [15000, 10000, 8000, 7000, 6000, 5000, 4000, 3000, 2000, 1500, 1000, 850, 650, 500, 400, 300, 200, 100],
            }
            self.top_label_dict = {
                "Precision Plus All Blue/Unstained": ["MWM", "S1", "S2", "S3", "S4", "S5", "S6", "S7", "S8", "S9", "MWM"],
                "1 kB Plus": ["MWM", "S1", "S2", "S3", "S4", "S5", "S6", "S7", "S8", "S9", "MWM"],
            }
            self.top_label = self.top_label_dict.get("Precision Plus All Blue/Unstained", ["MWM", "S1", "S2", "S3", "S4", "S5", "S6", "S7", "S8", "S9", "MWM"])
    
        # Update combo box options with loaded marker values
        self.combo_box.clear()
        self.combo_box.addItems(self.marker_values_dict.keys())
        self.combo_box.addItem("Custom")
    
    def paste_image(self):
        """Handle pasting image from clipboard."""
        self.is_modified = True
        self.reset_image() # Clear previous state first
        self.load_image_from_clipboard()
        # UI updates (label size, sliders) should happen within load_image_from_clipboard
        self.update_live_view()
        self.save_state() # Save state after pasting
    
    def load_image_from_clipboard(self):
        """Load an image from the clipboard into self.image, preserving format."""
        clipboard = QApplication.clipboard()
        mime_data = clipboard.mimeData()
        loaded_image = None
        source_info = "Clipboard" # Default source

        if mime_data.hasImage():
            loaded_image = clipboard.image()  # Get QImage
            if not loaded_image.isNull():
                source_info = "Clipboard Image"
            else: # Sometimes clipboard.image() fails, try Pillow as backup
                try:
                    pil_image = ImageGrab.grabclipboard()
                    if isinstance(pil_image, Image.Image):
                        loaded_image = ImageQt.ImageQt(pil_image)
                        source_info = "Clipboard Image (Pillow Grab)"
                    else:
                        loaded_image = None # Reset if Pillow didn't get an image
                except Exception as e:
                    print(f"Pillow clipboard grab failed: {e}")
                    loaded_image = None

        elif mime_data.hasUrls():
            urls = mime_data.urls()
            if urls:
                file_path = urls[0].toLocalFile()
                if file_path.lower().endswith(('.png', '.jpg', '.jpeg', '.bmp', '.tif', '.tiff')):
                    loaded_image = QImage(file_path)
                    source_info = file_path
                    if loaded_image.isNull():
                         # Try Pillow fallback for files too
                        try:
                            pil_image = Image.open(file_path)
                            loaded_image = ImageQt.ImageQt(pil_image)
                            if loaded_image.isNull(): raise ValueError("Pillow failed for file.")
                            source_info = f"{file_path} (Pillow)"
                            print(f"Loaded pasted file using Pillow fallback, format: {loaded_image.format()}")
                        except Exception as e:
                            QMessageBox.warning(self, "Error", f"Failed to load pasted file '{os.path.basename(file_path)}': {e}")
                            return
                else:
                     QMessageBox.warning(self, "Paste Error", "Pasted file is not a supported image type.")
                     return

        if not loaded_image or loaded_image.isNull():
            QMessageBox.warning(self, "Paste Error", "No valid image found on clipboard or in pasted file.")
            return

        # --- Keep the loaded image format ---
        self.image = loaded_image
        print(f"Pasted image format: {self.image.format()}")

        # --- Initialize backups ---
        if not self.image.isNull():
            self.original_image = self.image.copy() # Keep original pasted format
            self.image_master = self.image.copy()
            self.image_before_padding = None
            self.image_contrasted = self.image.copy()
            self.image_before_contrast = self.image.copy()
            self.image_padded = False
            self.setWindowTitle(f"{self.window_title}::IMAGE SIZE:{self.image.width()}x{self.image.height()}:{source_info}")

            # --- Update UI Elements (Label size, sliders) ---
            try:
                # (UI update logic remains the same)
                w=self.image.width()
                h=self.image.height()
                ratio=w/h if h > 0 else 1
                self.label_width=int(self.screen_width * 0.28)
                label_height=int(self.label_width/ratio)
                if label_height>self.label_width:
                    label_height=self.label_width
                    self.label_width=int(ratio*label_height)
                self.live_view_label.setFixedSize(int(self.label_width), int(label_height))

                render_scale = 3
                render_width = self.live_view_label.width() * render_scale
                render_height = self.live_view_label.height() * render_scale
                self.left_slider_range=[-100,int(render_width)+100]
                self.left_padding_slider.setRange(self.left_slider_range[0],self.left_slider_range[1])
                self.right_slider_range=[-100,int(render_width)+100]
                self.right_padding_slider.setRange(self.right_slider_range[0],self.right_slider_range[1])
                self.top_slider_range=[-100,int(render_height)+100]
                self.top_padding_slider.setRange(self.top_slider_range[0],self.top_slider_range[1])

                self.left_padding_input.setText(str(int(self.image.width()*0.1)))
                self.right_padding_input.setText(str(int(self.image.width()*0.1)))
                self.top_padding_input.setText(str(int(self.image.height()*0.15)))
                self.bottom_padding_input.setText("0")
            except Exception as e:
                print(f"Error setting up UI after paste: {e}")
            # --- End UI Element Update ---
        else:
             QMessageBox.critical(self, "Paste Error", "Failed to initialize image object after pasting.")
    # --- END: Modified Loading / Pasting ---
        
    def update_font(self):
        """Update the font settings based on UI inputs"""
        # Update font family from the combo box
        self.font_family = self.font_combo_box.currentFont().family()
        
        # Update font size from the spin box
        self.font_size = self.font_size_spinner.value()
        
        self.font_rotation = int(self.font_rotation_input.value())
        
    
        # Once font settings are updated, update the live view immediately
        self.update_live_view()
    
    def select_font_color(self):
        color = QColorDialog.getColor()
        if color.isValid():
            self.font_color = color  # Store the selected color   
            self.update_font()
        self._update_color_button_style(self.font_color_button, self.font_color)
            
    def load_image(self):
        self.is_modified = True # Mark as modified when loading new image
        self.undo_stack = []
        self.redo_stack = []
        self.reset_image() # Clear previous state

        options = QFileDialog.Options()
        file_path, _ = QFileDialog.getOpenFileName(
            self, "Open Image File", "", "Image Files (*.png *.jpg *.bmp *.tif *.tiff)", options=options
        )
        if file_path:
            self.image_path = file_path
            loaded_image = QImage(self.image_path)

            if loaded_image.isNull():
                # Try loading with Pillow as fallback
                try:
                    pil_image = Image.open(self.image_path)
                    # Use ImageQt for reliable conversion, preserves most formats
                    loaded_image = ImageQt.ImageQt(pil_image)
                    if loaded_image.isNull():
                        raise ValueError("Pillow could not convert to QImage.")
                    print(f"Loaded '{os.path.basename(file_path)}' using Pillow fallback, format: {loaded_image.format()}")

                except Exception as e:
                    QMessageBox.warning(self, "Error", f"Failed to load image '{os.path.basename(file_path)}': {e}")
                    self.image_path = None
                    return

            # --- Keep the loaded image format ---
            self.image = loaded_image
            print(f"Loaded image format: {self.image.format()}")

            # --- Initialize backups with the loaded format ---
            if not self.image.isNull():
                self.original_image = self.image.copy() # Keep a pristine copy of the initially loaded image
                self.image_master = self.image.copy()   # Master copy for resets
                self.image_before_padding = None        # Reset padding state
                self.image_contrasted = self.image.copy() # Backup for contrast
                self.image_before_contrast = self.image.copy() # Backup for contrast
                self.image_padded = False               # Reset flag

                self.setWindowTitle(f"{self.window_title}::IMAGE SIZE:{self.image.width()}x{self.image.height()}:{self.image_path}")

                # --- Load Associated Config File (Logic remains the same) ---
                self.base_name = os.path.splitext(os.path.basename(file_path))[0]
                config_name = ""
                if self.base_name.endswith("_original"):
                    config_name = self.base_name.replace("_original", "_config.txt")
                elif self.base_name.endswith("_modified"):
                    config_name = self.base_name.replace("_modified", "_config.txt")
                else:
                    config_name = self.base_name + "_config.txt"

                config_path = os.path.join(os.path.dirname(file_path), config_name)
                print(f"Looking for config file at: {config_path}")

                if os.path.exists(config_path):
                    try:
                        print(f"Found config file. Loading...")
                        with open(config_path, "r") as config_file:
                            config_data = json.load(config_file)
                        self.apply_config(config_data) # Apply loaded settings
                        print("Config file applied.")
                    except Exception as e:
                        QMessageBox.warning(self, "Config Load Error", f"Failed to load or apply config file '{config_name}': {e}")
                else:
                    print("No associated config file found.")
                # --- End Config File Loading ---

            else:
                 QMessageBox.critical(self, "Load Error", "Failed to initialize image object after loading.")
                 return

            # --- Update UI Elements (Label size, sliders) ---
            if self.image and not self.image.isNull():
                try:
                    # (UI update logic remains the same as before)
                    w=self.image.width()
                    h=self.image.height()
                    ratio=w/h if h > 0 else 1
                    self.label_width=int(self.screen_width * 0.28)
                    label_height=int(self.label_width/ratio)
                    if label_height>self.label_width:
                        label_height=self.label_width
                        self.label_width=int(ratio*label_height)
                    self.live_view_label.setFixedSize(int(self.label_width), int(label_height))

                    render_scale = 3
                    render_width = self.live_view_label.width() * render_scale
                    render_height = self.live_view_label.height() * render_scale
                    self.left_slider_range=[-100,int(render_width)+100]
                    self.left_padding_slider.setRange(self.left_slider_range[0],self.left_slider_range[1])
                    self.right_slider_range=[-100,int(render_width)+100]
                    self.right_padding_slider.setRange(self.right_slider_range[0],self.right_slider_range[1])
                    self.top_slider_range=[-100,int(render_height)+100]
                    self.top_padding_slider.setRange(self.top_slider_range[0],self.top_slider_range[1])

                    if not os.path.exists(config_path):
                        self.left_padding_input.setText(str(int(self.image.width()*0.1)))
                        self.right_padding_input.setText(str(int(self.image.width()*0.1)))
                        self.top_padding_input.setText(str(int(self.image.height()*0.15)))
                        self.bottom_padding_input.setText("0")

                except Exception as e:
                    print(f"Error setting up UI after load: {e}")
            # --- End UI Element Update ---

            self.update_live_view() # Render the loaded image
            self.save_state() # Save initial loaded state
    
    def apply_config(self, config_data):
        self.left_padding_input.setText(config_data["adding_white_space"]["left"])
        self.right_padding_input.setText(config_data["adding_white_space"]["right"])
        self.top_padding_input.setText(config_data["adding_white_space"]["top"])
        try:
            self.bottom_padding_input.setText(config_data["adding_white_space"]["bottom"])
        except:
            pass
    
        self.crop_x_start_slider.setValue(config_data["cropping_parameters"]["x_start_percent"])
        self.crop_x_end_slider.setValue(config_data["cropping_parameters"]["x_end_percent"])
        self.crop_y_start_slider.setValue(config_data["cropping_parameters"]["y_start_percent"])
        self.crop_y_end_slider.setValue(config_data["cropping_parameters"]["y_end_percent"])
    
        try:
            self.left_markers = [(float(pos), str(label)) for pos, label in config_data["marker_positions"]["left"]]
            self.right_markers = [(float(pos), str(label)) for pos, label in config_data["marker_positions"]["right"]]
            self.top_markers = [(float(pos), str(label)) for pos, label in config_data["marker_positions"]["top"]]
        except (KeyError, ValueError) as e:
            # QMessageBox.warning(self, "Error", f"Invalid marker data in config: {e}")
            pass
        try:
            # print("TOP LABELS: ",config_data["marker_labels"]["top"])
            self.top_label = [str(label) for label in config_data["marker_labels"]["top"]]
            self.top_marker_input.setText(", ".join(self.top_label))
        except KeyError as e:
            # QMessageBox.warning(self, "Error", f"Invalid marker labels in config: {e}")
            pass
    
        self.font_family = config_data["font_options"]["font_family"]
        self.font_size = config_data["font_options"]["font_size"]
        self.font_rotation = config_data["font_options"]["font_rotation"]
        self.font_color = QColor(config_data["font_options"]["font_color"])
    
        self.top_padding_slider.setValue(config_data["marker_padding"]["top"])
        self.left_padding_slider.setValue(config_data["marker_padding"]["left"])
        self.right_padding_slider.setValue(config_data["marker_padding"]["right"])
        
        try:
            try:
                x = list(map(int, config_data["marker_labels"]["left"]))
                self.marker_values_textbox.setText(str(x))
            except:
                x = list(map(int, config_data["marker_labels"]["right"]))
                self.marker_values_textbox.setText(str(x))
            if self.marker_values_textbox.text!="":
                self.combo_box.setCurrentText("Custom")
                self.marker_values_textbox.setEnabled(True)                
            else:
                self.combo_box.setCurrentText("Precision Plus All Blue/Unstained")
                self.marker_values_textbox.setEnabled(False)
                
        except:
            # print("ERROR IN LEFT/RIGHT MARKER DATA")
            pass
    
        try:
            self.custom_markers = [
                (marker["x"], marker["y"], marker["text"], QColor(marker["color"]), marker["font"], marker["font_size"])
                for marker in config_data.get("custom_markers", [])
            ]
                
                
                
        except:
            pass
        # Set slider ranges from config_data
        try:
            self.left_padding_slider.setRange(
                int(config_data["slider_ranges"]["left"][0]), int(config_data["slider_ranges"]["left"][1])
            )
            self.right_padding_slider.setRange(
                int(config_data["slider_ranges"]["right"][0]), int(config_data["slider_ranges"]["right"][1])
            )
            self.top_padding_slider.setRange(
                int(config_data["slider_ranges"]["top"][0]), int(config_data["slider_ranges"]["top"][1])
            )
        except KeyError:
            # Handle missing or incomplete slider_ranges data
            # print("Error: Slider ranges not found in config_data.")
            pass
        
        try:
            self.left_marker_shift_added=int(config_data["added_shift"]["left"])
            self.right_marker_shift_added=int(config_data["added_shift"]["right"])
            self.top_marker_shift_added=int(config_data["added_shift"]["top"])
            
        except KeyError:
            # Handle missing or incomplete slider_ranges data
            # print("Error: Added Shift not found in config_data.");
            pass
            
        # Apply font settings

        
        #DO NOT KNOW WHY THIS WORKS BUT DIRECT VARIABLE ASSIGNING DOES NOT WORK
        
        font_size_new=self.font_size
        font_rotation_new=self.font_rotation
        
        self.font_combo_box.setCurrentFont(QFont(self.font_family))
        self.font_size_spinner.setValue(font_size_new)
        self.font_rotation_input.setValue(font_rotation_new)

        self.update_live_view()
        
    def get_current_config(self):
        config = {
            "adding_white_space": {
                "left": self.left_padding_input.text(),
                "right": self.right_padding_input.text(),
                "top": self.top_padding_input.text(),
                "bottom": self.bottom_padding_input.text(),
            },
            "cropping_parameters": {
                "x_start_percent": self.crop_x_start_slider.value(),
                "x_end_percent": self.crop_x_end_slider.value(),
                "y_start_percent": self.crop_y_start_slider.value(),
                "y_end_percent": self.crop_y_end_slider.value(),
            },
            "marker_positions": {
                "left": self.left_markers,
                "right": self.right_markers,
                "top": self.top_markers,
            },
            "marker_labels": {
                "top": self.top_label,
                "left": [marker[1] for marker in self.left_markers],
                "right": [marker[1] for marker in self.right_markers],
            },
            "marker_padding": {
                "top": self.top_padding_slider.value(),
                "left": self.left_padding_slider.value(),
                "right": self.right_padding_slider.value(),
            },
            "font_options": {
                "font_family": self.font_family,
                "font_size": self.font_size,
                "font_rotation": self.font_rotation,
                "font_color": self.font_color.name(),
            },
        }
    
        try:
            # Add custom markers with font and font size
            config["custom_markers"] = [
                {"x": x, "y": y, "text": text, "color": color.name(), "font": font, "font_size": font_size}
                for x, y, text, color, font, font_size in self.custom_markers
            ]
        except AttributeError:
            # Handle the case where self.custom_markers is not defined or invalid
            config["custom_markers"] = []
        try:
            config["slider_ranges"] = {
                    "left": self.left_slider_range,
                    "right": self.right_slider_range,
                    "top": self.top_slider_range,
                }
            
        except AttributeError:
            # Handle the case where slider ranges are not defined
            config["slider_ranges"] = []
            
        try:
            config["added_shift"] = {
                    "left": self.left_marker_shift_added,
                    "right": self.right_marker_shift_added,
                    "top": self.top_marker_shift_added,
                }
            
        except AttributeError:
            # Handle the case where slider ranges are not defined
            config["added_shift"] = []
    
        return config
    
    def add_band(self, event):
        # --- Ensure internal lists match UI input BEFORE adding band ---
        # This call ensures self.marker_values and self.top_label are
        # updated based on the current text in the UI boxes.
        self.update_all_labels()
        # -------------------------------------------------------------

        self.save_state()
        # Ensure there's an image loaded and marker mode is active
        self.live_view_label.preview_marker_enabled = False
        self.live_view_label.preview_marker_text = ""
        if not self.image or not self.marker_mode:
            return

        # --- Get Coordinates and Scaling (Same as previous version) ---
        pos = event.pos()
        cursor_x, cursor_y = pos.x(), pos.y()
        if self.live_view_label.zoom_level != 1.0:
            cursor_x = (cursor_x - self.live_view_label.pan_offset.x()) / self.live_view_label.zoom_level
            cursor_y = (cursor_y - self.live_view_label.pan_offset.y()) / self.live_view_label.zoom_level

        if self.show_grid_checkbox.isChecked():
            grid_size = self.grid_size_input.value()
            cursor_x = round(cursor_x / grid_size) * grid_size
            cursor_y = round(cursor_y / grid_size) * grid_size

        displayed_width = self.live_view_label.width()
        displayed_height = self.live_view_label.height()
        image_width = self.image.width() if self.image.width() > 0 else 1
        image_height = self.image.height() if self.image.height() > 0 else 1
        uniform_scale = min(displayed_width / image_width, displayed_height / image_height) if image_width > 0 and image_height > 0 else 1
        offset_x = (displayed_width - image_width * uniform_scale) / 2
        offset_y = (displayed_height - image_height * uniform_scale) / 2
        image_x = (cursor_x - offset_x) / uniform_scale if uniform_scale != 0 else 0
        image_y = (cursor_y - offset_y) / uniform_scale if uniform_scale != 0 else 0
        
        x_start_percent = self.crop_x_start_slider.value() / 100
        x_end_percent = self.crop_x_end_slider.value() / 100
        y_start_percent = self.crop_y_start_slider.value() / 100
        y_end_percent = self.crop_y_end_slider.value() / 100
    
        # Calculate the crop boundaries based on the percentages
        x_start = int(self.image.width() * x_start_percent)
        x_end = int(self.image.width() * x_end_percent)
        y_start = int(self.image.height() * y_start_percent)
        y_end = int(self.image.height() * y_end_percent)

        # --- Get Render Info for Slider Positioning (Same as previous version) ---
        render_scale = 3
        render_width = self.live_view_label.width() * render_scale
        render_height = self.live_view_label.height() * render_scale

        # --- Validate Coordinates (Same as previous version) ---
        current_image_width = self.image.width()
        current_image_height = self.image.height()
        if not (0 <= image_y <= current_image_height) and self.marker_mode in ["left", "right"]:
             return
        if not (0 <= image_x <= current_image_width) and self.marker_mode == "top":
             return
        # --- End Coordinate/Scaling/Validation ---

        try:
            # --- Left Marker Logic ---
            if self.marker_mode == "left":
                # Determine label based on current count and *now updated* self.marker_values
                current_marker_count = len(self.left_markers)
                is_first_marker = (current_marker_count == 0)

                # Use the self.marker_values list, which was just updated
                if current_marker_count < len(self.marker_values):
                    marker_value_to_add = self.marker_values[current_marker_count]
                else:
                    marker_value_to_add = ""
                    if current_marker_count == len(self.marker_values): # Print warning only once
                        print(f"Warning: Adding left marker {current_marker_count + 1} beyond preset count. Using empty label.")

                self.left_markers.append((image_y, marker_value_to_add))
                self.current_left_marker_index += 1 # Still increment conceptual index

                # Set slider position only for the *very first* marker placed
                if is_first_marker:
                    padding_value=int((image_x - x_start) * (render_width / self.image.width()))
                    self.left_padding_slider.setValue(0)
                    self.left_slider_range=[-100,int(render_width)+100]
                    self.left_padding_slider.setRange(self.left_slider_range[0],self.left_slider_range[1])
                    self.left_padding_slider.setValue(padding_value)
                    self.left_marker_shift_added = self.left_padding_slider.value()    

            # --- Right Marker Logic ---
            elif self.marker_mode == "right":
                current_marker_count = len(self.right_markers)
                is_first_marker = (current_marker_count == 0)

                # Use the self.marker_values list, which was just updated
                if current_marker_count < len(self.marker_values):
                    marker_value_to_add = self.marker_values[current_marker_count]
                else:
                    marker_value_to_add = ""
                    if current_marker_count == len(self.marker_values): # Print warning only once
                        print(f"Warning: Adding right marker {current_marker_count + 1} beyond preset count. Using empty label.")

                self.right_markers.append((image_y, marker_value_to_add))
                self.current_right_marker_index += 1

                if is_first_marker:
                    padding_value=int((image_x - x_start) * (render_width / self.image.width()))
                    self.right_padding_slider.setValue(0)
                    self.right_slider_range=[-100,int(render_width)+100]
                    self.right_padding_slider.setRange(self.right_slider_range[0],self.right_slider_range[1])
                    self.right_padding_slider.setValue(padding_value)
                    self.right_marker_shift_added = self.right_padding_slider.value()

            # --- Top Marker Logic ---
            elif self.marker_mode == "top":
                current_marker_count = len(self.top_markers)
                is_first_marker = (current_marker_count == 0)

                # Use the self.top_label list, which was just updated
                if current_marker_count < len(self.top_label):
                    label_to_add = self.top_label[current_marker_count]
                else:
                    label_to_add = ""
                    if current_marker_count == len(self.top_label): # Print warning only once
                         print(f"Warning: Adding top marker {current_marker_count + 1} beyond preset count. Using empty label.")

                self.top_markers.append((image_x, label_to_add))
                self.current_top_label_index += 1

                if is_first_marker:
                    padding_value=int((image_y - y_start) * (render_height / self.image.height()))
                    self.top_padding_slider.setValue(0)
                    self.top_slider_range=[-100,int(render_height)+100]
                    self.top_padding_slider.setRange(self.top_slider_range[0],self.top_slider_range[1])
                    self.top_padding_slider.setValue(padding_value)
                    self.top_marker_shift_added = self.top_padding_slider.value()

        except Exception as e:
             # Catch other potential errors
             print(f"ERROR ADDING BANDS: An unexpected error occurred - {type(e).__name__}: {e}")
             import traceback
             traceback.print_exc()
             QMessageBox.critical(self, "Error", f"An unexpected error occurred while adding the marker:\n{e}")

        # Update the live view to render the newly added marker
        self.update_live_view()
        

        
    def enable_left_marker_mode(self):
        self.marker_mode = "left"
        self.current_left_marker_index = 0
        self.live_view_label.mousePressEvent = self.add_band
        self.live_view_label.setCursor(Qt.CrossCursor)

    def enable_right_marker_mode(self):
        self.marker_mode = "right"
        self.current_right_marker_index = 0
        self.live_view_label.mousePressEvent = self.add_band
        self.live_view_label.setCursor(Qt.CrossCursor)
    
    def enable_top_marker_mode(self):
        self.marker_mode = "top"
        self.current_top_label_index
        self.live_view_label.mousePressEvent = self.add_band
        self.live_view_label.setCursor(Qt.CrossCursor)
        
    # def remove_padding(self):
    #     if self.image_before_padding!=None:
    #         self.image = self.image_before_padding.copy()  # Revert to the image before padding
    #     self.image_contrasted = self.image.copy()  # Sync the contrasted image
    #     self.image_padded = False  # Reset the padding state
    #     w=self.image.width()
    #     h=self.image.height()
    #     # Preview window
    #     ratio=w/h
    #     self.label_width = 540
    #     label_height=int(self.label_width/ratio)
    #     if label_height>self.label_width:
    #         label_height=540
    #         self.label_width=ratio*label_height
    #     self.live_view_label.setFixedSize(int(self.label_width), int(label_height))
    #     self.update_live_view()
        
    def finalize_image(self): # Padding
        self.save_state()
        if not self.image or self.image.isNull():
            QMessageBox.warning(self, "Error", "No image loaded to apply padding.")
            return

        try:
            padding_left = abs(int(self.left_padding_input.text()))
            padding_right = abs(int(self.right_padding_input.text()))
            padding_top = abs(int(self.top_padding_input.text()))
            padding_bottom = abs(int(self.bottom_padding_input.text()))
        except ValueError:
            QMessageBox.warning(self, "Error", "Please enter valid integers for padding.")
            return

        # Don't rely on get_compatible_grayscale_format, check the actual image
        try:
            np_img = self.qimage_to_numpy(self.image)
            if np_img is None: raise ValueError("NumPy conversion failed.")

            # Determine fill value based on dtype and dimensions
            fill_value = None
            if np_img.ndim == 2: # Grayscale
                fill_value = 65535 if np_img.dtype == np.uint16 else 255
                print(f"Padding grayscale image with fill value: {fill_value}")
            elif np_img.ndim == 3: # Color (assume BGR/BGRA)
                num_channels = np_img.shape[2]
                base_val = 65535 if np_img.dtype == np.uint16 else 255
                if num_channels == 3: # BGR
                    fill_value = (base_val, base_val, base_val)
                elif num_channels == 4: # BGRA - white opaque
                    fill_value = (base_val, base_val, base_val, base_val) # White, Alpha=max
                else:
                     raise ValueError(f"Unsupported number of color channels: {num_channels}")
                print(f"Padding color image ({num_channels} channels) with fill value: {fill_value}")
            else:
                 raise ValueError(f"Unsupported image dimensions: {np_img.ndim}")

            # Adjust markers BEFORE creating the new image
            self.adjust_markers_for_padding(padding_left, padding_right, padding_top, padding_bottom)

            # Pad using OpenCV's copyMakeBorder
            padded_np = cv2.copyMakeBorder(np_img, padding_top, padding_bottom,
                                           padding_left, padding_right,
                                           cv2.BORDER_CONSTANT, value=fill_value) # Use color/gray fill_value

            # Convert back to QImage
            padded_image = self.numpy_to_qimage(padded_np) # Should handle color correctly now
            if padded_image.isNull():
                 raise ValueError("Conversion back to QImage failed.")

            # Update main image and backups
            self.image_before_padding = self.image.copy() if not self.image_padded else self.image_before_padding
            self.image = padded_image
            self.image_padded = True # Indicate padding applied
            # Update backups consistently
            self.image_contrasted = self.image.copy()
            self.image_before_contrast = self.image.copy()

            # Update UI (label size, slider ranges, live view)
            # (UI update logic remains the same)
            w = self.image.width()
            h = self.image.height()
            ratio = w / h if h > 0 else 1
            self.label_width = int(self.screen_width * 0.28)
            label_height = int(self.label_width / ratio)
            if label_height > self.label_width:
                label_height = self.label_width
                self.label_width = int(ratio * label_height)
            self.live_view_label.setFixedSize(int(self.label_width), int(label_height))

            render_scale = 3
            render_width = self.live_view_label.width() * render_scale
            render_height = self.live_view_label.height() * render_scale
            self.left_slider_range = [-100, int(render_width) + 100]
            self.left_padding_slider.setRange(self.left_slider_range[0], self.left_slider_range[1])
            self.right_slider_range = [-100, int(render_width) + 100]
            self.right_padding_slider.setRange(self.right_slider_range[0], self.right_slider_range[1])
            self.top_slider_range = [-100, int(render_height) + 100]
            self.top_padding_slider.setRange(self.top_slider_range[0], self.top_slider_range[1])


            self.update_live_view()
            print(f"Padding applied. New image format: {self.image.format()}")

        except Exception as e:
            QMessageBox.critical(self, "Padding Error", f"Failed to apply padding: {e}")
            traceback.print_exc()
    # --- END: Modified Image Operations ---
    
    def adjust_markers_for_padding(self, padding_left, padding_right, padding_top, padding_bottom):
        """Adjust marker positions based on padding."""
        # Adjust left markers
        self.left_markers = [(y + padding_top, label) for y, label in self.left_markers]
        # Adjust right markers
        self.right_markers = [(y + padding_top, label) for y, label in self.right_markers]
        # Adjust top markers
        self.top_markers = [(x + padding_left, label) for x, label in self.top_markers]        
        
    
    def update_left_padding(self):
        # Update left padding when slider value changes
        self.left_marker_shift_added = self.left_padding_slider.value()
        self.update_live_view()

    def update_right_padding(self):
        # Update right padding when slider value changes
        self.right_marker_shift_added = self.right_padding_slider.value()
        self.update_live_view()
        
    def update_top_padding(self):
        # Update top padding when slider value changes
        self.top_marker_shift_added = self.top_padding_slider.value()
        self.update_live_view()

    def update_live_view(self):
        if not self.image:
            return
    
        # Enable the "Predict Molecular Weight" button if markers are present
        if self.left_markers or self.right_markers:
            self.predict_button.setEnabled(True)
        else:
            self.predict_button.setEnabled(False)
    
        # Define a higher resolution for processing (e.g., 2x or 3x label size)
        render_scale = 3  # Scale factor for rendering resolution
        render_width = self.live_view_label.width() * render_scale
        render_height = self.live_view_label.height() * render_scale
    
        # Calculate scaling factors and offsets
        scale_x = self.image.width() / render_width
        scale_y = self.image.height() / render_height
    
        # Get the crop percentage values from sliders
        x_start_percent = self.crop_x_start_slider.value() / 100
        x_end_percent = self.crop_x_end_slider.value() / 100
        y_start_percent = self.crop_y_start_slider.value() / 100
        y_end_percent = self.crop_y_end_slider.value() / 100
    
        # Calculate the crop boundaries based on the percentages
        x_start = int(self.image.width() * x_start_percent)
        x_end = int(self.image.width() * x_end_percent)
        y_start = int(self.image.height() * y_start_percent)
        y_end = int(self.image.height() * y_end_percent)
    
        # Ensure the cropping boundaries are valid
        if x_start >= x_end or y_start >= y_end:
            QMessageBox.warning(self, "Warning", "Invalid cropping values.")
            self.crop_x_start_slider.setValue(0)
            self.crop_x_end_slider.setValue(100)
            self.crop_y_start_slider.setValue(0)
            self.crop_y_end_slider.setValue(100)
            return
    
        # Crop the image based on the defined boundaries
        cropped_image = self.image.copy(x_start, y_start, x_end - x_start, y_end - y_start)
    
        # Get the orientation value from the slider
        orientation = float(self.orientation_slider.value() / 20)  # Orientation slider value
        self.orientation_label.setText(f"Rotation Angle ({orientation:.2f}°)")
    
        # Apply the rotation to the cropped image
        rotated_image = cropped_image.transformed(QTransform().rotate(orientation))
    
        taper_value = self.taper_skew_slider.value() / 100  # Normalize taper value to a range of -1 to 1
        self.taper_skew_label.setText(f"Tapering Skew ({taper_value:.2f}) ")
    
        width = self.image.width()
        height = self.image.height()
    
        # Define corner points for perspective transformation
        source_corners = QPolygonF([QPointF(0, 0), QPointF(width, 0), QPointF(0, height), QPointF(width, height)])
    
        # Initialize destination corners as a copy of source corners
        destination_corners = QPolygonF(source_corners)
    
        # Adjust perspective based on taper value
        if taper_value > 0:
            # Narrower at the top, wider at the bottom
            destination_corners[0].setX(width * taper_value / 2)  # Top-left
            destination_corners[1].setX(width * (1 - taper_value / 2))  # Top-right
        elif taper_value < 0:
            # Wider at the top, narrower at the bottom
            destination_corners[2].setX(width * (-taper_value / 2))  # Bottom-left
            destination_corners[3].setX(width * (1 + taper_value / 2))  # Bottom-right
    
        # Create a perspective transformation using quadToQuad
        transform = QTransform()
        if not QTransform.quadToQuad(source_corners, destination_corners, transform):
            print("Failed to create transformation matrix")
            return
    
        # Apply the transformation
        skewed_image = rotated_image.transformed(transform, Qt.SmoothTransformation)
    
        # Scale the rotated image to the rendering resolution
        scaled_image = skewed_image.scaled(
            render_width,
            render_height,
            Qt.KeepAspectRatio,
            Qt.SmoothTransformation,
        )
    
        # Render on a high-resolution canvas
        canvas = QImage(render_width, render_height, QImage.Format_ARGB32)
        canvas.fill(Qt.transparent)  # Transparent background
    
        # Render the base image and overlays
        self.render_image_on_canvas(canvas, scaled_image, x_start, y_start, render_scale)
    
        # Scale the high-resolution canvas down to the label's size for display
        pixmap = QPixmap.fromImage(canvas).scaled(
            self.live_view_label.size(),
            Qt.KeepAspectRatio,
            Qt.SmoothTransformation,
        )
    
        painter = QPainter(pixmap)
        pen = QPen(Qt.red)
        pen.setWidth(1)
        painter.setPen(pen)
    
        scale_x = self.live_view_label.width() / self.image.width()
        scale_y = self.live_view_label.height() / self.image.height()
    
        # Draw the quadrilateral or rectangle if in move mode
        if self.live_view_label.mode == "move":
            pen = QPen(Qt.red, 2)
            painter.setPen(pen)
            if self.live_view_label.quad_points:
                # Draw the quadrilateral
                painter.drawPolygon(QPolygonF(self.live_view_label.quad_points))
            elif self.live_view_label.bounding_box_preview:
                # Draw the rectangle
                start_x, start_y, end_x, end_y = self.live_view_label.bounding_box_preview
                rect = QRectF(start_x, start_y, end_x - start_x, end_y - start_y)
                painter.drawRect(rect)
    
        painter.end()
    
        # Apply zoom and pan transformations to the final pixmap
        if self.live_view_label.zoom_level != 1.0:
            # Create a new pixmap to apply zoom and pan
            zoomed_pixmap = QPixmap(pixmap.size())
            zoomed_pixmap.fill(Qt.transparent)
            zoom_painter = QPainter(zoomed_pixmap)
            zoom_painter.translate(self.live_view_label.pan_offset)
            zoom_painter.scale(self.live_view_label.zoom_level, self.live_view_label.zoom_level)
            zoom_painter.drawPixmap(0, 0, pixmap)
            zoom_painter.end()  # Properly end the QPainter
            pixmap = zoomed_pixmap
    
        # Set the final pixmap to the live view label
        self.live_view_label.setPixmap(pixmap)
    
    def render_image_on_canvas(self, canvas, scaled_image, x_start, y_start, render_scale, draw_guides=True):
        painter = QPainter(canvas)
        x_offset = (canvas.width() - scaled_image.width()) // 2
        y_offset = (canvas.height() - scaled_image.height()) // 2
        
        self.x_offset_s=x_offset
        self.y_offset_s=y_offset
    
        # Draw the base image
        painter.drawImage(x_offset, y_offset, scaled_image)
    
        # Draw Image 1 if it exists
        if hasattr(self, 'image1') and hasattr(self, 'image1_position'):
            self.image1_position = (self.image1_left_slider.value(), self.image1_top_slider.value())
            # Resize Image 1 based on the slider value
            scale_factor = self.image1_resize_slider.value() / 100.0
            width = int(self.image1_original.width() * scale_factor)
            height = int(self.image1_original.height() * scale_factor)
            resized_image1 = self.image1_original.scaled(width, height, Qt.KeepAspectRatio, Qt.SmoothTransformation)
    
            # Calculate the position of Image 1
            image1_x = x_offset + self.image1_position[0]
            image1_y = y_offset + self.image1_position[1]
            painter.drawImage(image1_x, image1_y, resized_image1)
    
        # Draw Image 2 if it exists
        if hasattr(self, 'image2') and hasattr(self, 'image2_position'):
            self.image2_position = (self.image2_left_slider.value(), self.image2_top_slider.value())
            # Resize Image 2 based on the slider value
            scale_factor = self.image2_resize_slider.value() / 100.0
            width = int(self.image2_original.width() * scale_factor)
            height = int(self.image2_original.height() * scale_factor)
            resized_image2 = self.image2_original.scaled(width, height, Qt.KeepAspectRatio, Qt.SmoothTransformation)
    
            # Calculate the position of Image 2
            image2_x = x_offset + self.image2_position[0]
            image2_y = y_offset + self.image2_position[1]
            painter.drawImage(image2_x, image2_y, resized_image2)
    
        # Get the selected font settings
        font = QFont(self.font_combo_box.currentFont().family(), self.font_size_spinner.value() * render_scale)
        font_color = self.font_color if hasattr(self, 'font_color') else QColor(0, 0, 0)  # Default to black if no color selected
    
        painter.setFont(font)
        painter.setPen(font_color)  # Set the font color
    
        # Measure text height for vertical alignment
        font_metrics = painter.fontMetrics()
        text_height = font_metrics.height()
        line_padding = 5 * render_scale  # Space between text and line
        
        y_offset_global = text_height / 4
    
        # Draw the left markers (aligned right)
        for y_pos, marker_value in self.left_markers:
            y_pos_cropped = (y_pos - y_start) * (scaled_image.height() / self.image.height())
            if 0 <= y_pos_cropped <= scaled_image.height():
                text = f"{marker_value} ⎯ "  ##CHANGE HERE IF YOU WANT TO REMOVE THE "-"
                text_width = font_metrics.horizontalAdvance(text)  # Get text width
                painter.drawText(
                    int(x_offset + self.left_marker_shift + self.left_marker_shift_added - text_width),
                    int(y_offset + y_pos_cropped + y_offset_global),  # Adjust for proper text placement
                    text,
                )
        
        
        # Draw the right markers (aligned left)
        for y_pos, marker_value in self.right_markers:
            y_pos_cropped = (y_pos - y_start) * (scaled_image.height() / self.image.height())
            if 0 <= y_pos_cropped <= scaled_image.height():
                text = f" ⎯ {marker_value}" ##CHANGE HERE IF YOU WANT TO REMOVE THE "-"
                text_width = font_metrics.horizontalAdvance(text)  # Get text width
                painter.drawText(
                    int(x_offset + self.right_marker_shift_added),# + line_padding),
                    int(y_offset + y_pos_cropped + y_offset_global),  # Adjust for proper text placement
                    text,
                )
                
    
        # Draw the top markers (if needed)
        for x_pos, top_label in self.top_markers:
            x_pos_cropped = (x_pos - x_start) * (scaled_image.width() / self.image.width())
            if 0 <= x_pos_cropped <= scaled_image.width():
                text = f"{top_label}"
                painter.save()
                text_width = font_metrics.horizontalAdvance(text)
                text_height= font_metrics.height()
                label_x = x_offset + x_pos_cropped 
                label_y = y_offset + self.top_marker_shift + self.top_marker_shift_added 
                painter.translate(int(label_x), int(label_y))
                painter.rotate(self.font_rotation)
                painter.drawText(0,int(y_offset_global), f"{top_label}")
                painter.restore()
    
        # Draw guide lines
        if draw_guides and self.show_guides_checkbox.isChecked():
            pen = QPen(Qt.red, 2 * render_scale)
            painter.setPen(pen)
            center_x = canvas.width() // 2
            center_y = canvas.height() // 2
            painter.drawLine(center_x, 0, center_x, canvas.height())  # Vertical line
            painter.drawLine(0, center_y, canvas.width(), center_y)  # Horizontal line
            
        
        
        
        # Draw the protein location marker (*)
        if hasattr(self, "protein_location") and self.run_predict_MW == False:
            x, y = self.protein_location
            text = "⎯⎯"
            text_width = font_metrics.horizontalAdvance(text)
            text_height= font_metrics.height()
            painter.drawText(
                int(x * render_scale - text_width / 2),  #Currently left edge # FOR Center horizontally use int(x * render_scale - text_width / 2)
                int(y * render_scale + text_height / 4),  # Center vertically
                text
            )
        if hasattr(self, "custom_markers"):
            
            # Get default font type and size
            default_font_type = QFont(self.custom_font_type_dropdown.currentText())
            default_font_size = int(self.custom_font_size_spinbox.value())
        
            for x_pos, y_pos, marker_text, color, *optional in self.custom_markers:
                
                   
                # Use provided font type and size if available, otherwise use defaults
                marker_font_type = optional[0] if len(optional) > 0 else default_font_type
                marker_font_size = optional[1] if len(optional) > 1 else default_font_size
        
                # Ensure marker_font_type is a QFont instance
                font = QFont(marker_font_type) if isinstance(marker_font_type, str) else QFont(marker_font_type)
                font.setPointSize(marker_font_size * render_scale)  # Adjust font size for rendering scale
        
                # Apply the font and pen settings
                painter.setFont(font)
                painter.setPen(color)
        
                # Correct scaling and offsets
                x_pos_cropped = (x_pos - x_start) * (scaled_image.width() / self.image.width()) 
                y_pos_cropped = (y_pos - y_start) * (scaled_image.height() / self.image.height()) 
        
                # Only draw markers if they fall within the visible scaled image area
                if 0 <= x_pos_cropped <= scaled_image.width() and 0 <= y_pos_cropped <= scaled_image.height():
                    # Calculate text dimensions for alignment
                    font_metrics = painter.fontMetrics()
                    text_width = font_metrics.horizontalAdvance(marker_text)
                    text_height = font_metrics.height()
        
                    # Draw text centered horizontally and vertically
                    painter.drawText(
                        int(x_pos_cropped + x_offset- text_width / 2),  # Center horizontally
                        int(y_pos_cropped + y_offset+ text_height / 4),  # Center vertically
                        marker_text
                    )
            
            
    

        # Draw the grid (if enabled)
        if self.show_grid_checkbox.isChecked():
            grid_size = self.grid_size_input.value() * render_scale
            pen = QPen(Qt.red)
            pen.setStyle(Qt.DashLine)
            painter.setPen(pen)
    
            # Draw vertical grid lines
            for x in range(0, canvas.width(), grid_size):
                painter.drawLine(x, 0, x, canvas.height())
    
            # Draw horizontal grid lines
            for y in range(0, canvas.height(), grid_size):
                painter.drawLine(0, y, canvas.width(), y)
                
        painter.end()


     
    def crop_image(self):
        """Function to crop the current image."""
        if not self.image:
            return None
    
        # Get crop percentage from sliders
        x_start_percent = self.crop_x_start_slider.value() / 100
        x_end_percent = self.crop_x_end_slider.value() / 100
        y_start_percent = self.crop_y_start_slider.value() / 100
        y_end_percent = self.crop_y_end_slider.value() / 100
    
        # Calculate crop boundaries
        x_start = int(self.image.width() * x_start_percent)
        x_end = int(self.image.width() * x_end_percent)
        y_start = int(self.image.height() * y_start_percent)
        y_end = int(self.image.height() * y_end_percent)
    
        # Ensure cropping is valid
        if x_start >= x_end or y_start >= y_end:
            QMessageBox.warning(self, "Warning", "Invalid cropping values.")
            return None
    
        # Crop the image
        cropped_image = self.image.copy(x_start, y_start, x_end - x_start, y_end - y_start)
        return cropped_image
    
    # Modify align_image and update_crop to preserve settings
    def reset_align_image(self):
        self.orientation_slider.setValue(0)
        
    def align_image(self):
        self.save_state()
        if not self.image or self.image.isNull(): return

        angle = float(self.orientation_slider.value() / 20.0) # Corrected division
        if abs(angle) < 0.01: # No significant rotation needed
            return

        target_grayscale_format = self.get_compatible_grayscale_format() # Grayscale16 or Grayscale8
        is_16bit = (target_grayscale_format == QImage.Format_Grayscale16)
        fill_color = 65535 if is_16bit else 255 # White background value

        # Use NumPy and OpenCV for rotation to better handle formats and prevent clipping
        try:
            np_img = self.qimage_to_numpy(self.image)
            if np_img is None: raise ValueError("NumPy conversion failed.")

            # Ensure grayscale for rotation calculation if needed
            if np_img.ndim == 3:
                 print("Warning: Aligning color image. Converting to grayscale first.")
                 color_code = cv2.COLOR_BGR2GRAY if np_img.shape[2] == 3 else cv2.COLOR_BGRA2GRAY
                 np_img_gray = cv2.cvtColor(np_img, color_code)
            else:
                 np_img_gray = np_img

            # Get rotation matrix
            (h, w) = np_img_gray.shape[:2]
            center = (w // 2, h // 2)
            M = cv2.getRotationMatrix2D(center, angle, 1.0)

            # Calculate new bounding box size
            cos = np.abs(M[0, 0])
            sin = np.abs(M[0, 1])
            new_w = int((h * sin) + (w * cos))
            new_h = int((h * cos) + (w * sin))

            # Adjust rotation matrix for translation
            M[0, 2] += (new_w / 2) - center[0]
            M[1, 2] += (new_h / 2) - center[1]

            # Perform the rotation using OpenCV warpAffine
            # Use the appropriate interpolation, e.g., INTER_LINEAR or INTER_CUBIC
            # Use the original numpy array (could be color or gray)
            rotated_np = cv2.warpAffine(np_img, M, (new_w, new_h),
                                        flags=cv2.INTER_LINEAR, # Or INTER_CUBIC
                                        borderMode=cv2.BORDER_CONSTANT,
                                        borderValue=fill_color) # Use appropriate fill color


            # Convert back to QImage
            aligned_canvas = self.numpy_to_qimage(rotated_np)

            if aligned_canvas.isNull():
                raise ValueError("Conversion back to QImage failed.")

            # Update the main image and backups
            self.image = aligned_canvas
            self.image_before_padding = self.image.copy()
            self.image_contrasted = self.image.copy()
            self.image_before_contrast = self.image.copy()

            # Reset slider and update view
            self.orientation_slider.setValue(0)
            self.update_live_view()

        except Exception as e:
            QMessageBox.critical(self, "Alignment Error", f"Failed to align image: {e}")
            # Optionally restore previous state if needed
            # self.undo_action() # If you want undo on error
    
    
    def update_crop(self):
        self.save_state()
        """Update the image based on current crop sliders. First align, then crop the image."""
        self.show_grid_checkbox.setChecked(False)
        self.update_live_view() # Update once before operations

        # Get crop parameters *before* aligning/cropping
        x_start_percent = self.crop_x_start_slider.value() / 100
        x_end_percent = self.crop_x_end_slider.value() / 100
        y_start_percent = self.crop_y_start_slider.value() / 100
        y_end_percent = self.crop_y_end_slider.value() / 100

        # Calculate the crop boundaries based on the *current* image dimensions before cropping
        if not self.image:
            return # Should not happen if called from UI, but safety check
        current_width = self.image.width()
        current_height = self.image.height()
        x_start = int(current_width * x_start_percent)
        x_end = int(current_width * x_end_percent)
        y_start = int(current_height * y_start_percent)
        y_end = int(current_height * y_end_percent)

        # Ensure cropping boundaries are valid relative to current image
        if x_start >= x_end or y_start >= y_end:
            QMessageBox.warning(self, "Warning", "Invalid cropping values based on current image size.")
            # Optionally reset sliders here if needed
            return


        # Align the image first (rotate it) - align_image modifies self.image
        # We align *before* cropping based on the original logic flow provided.
        # Note: If alignment changes dimensions significantly, this might need rethinking,
        # but typically rotation keeps content centered.
        # self.align_image() # Call align_image *if* it should happen before crop


        # Now apply cropping to the *current* self.image
        cropped_image = self.crop_image() # crop_image uses current self.image state

        if cropped_image:
            # --- Adjust marker positions relative to the crop ---
            new_left_markers = []
            for y, label in self.left_markers:
                if y_start <= y < y_end: # Check if marker was within vertical crop bounds
                    new_y = y - y_start
                    new_left_markers.append((new_y, label))
            self.left_markers = new_left_markers

            new_right_markers = []
            for y, label in self.right_markers:
                if y_start <= y < y_end:
                    new_y = y - y_start
                    new_right_markers.append((new_y, label))
            self.right_markers = new_right_markers

            new_top_markers = []
            for x, label in self.top_markers:
                if x_start <= x < x_end: # Check if marker was within horizontal crop bounds
                    new_x = x - x_start
                    new_top_markers.append((new_x, label))
            self.top_markers = new_top_markers

            new_custom_markers = []
            if hasattr(self, "custom_markers"):
                for x, y, text, color, font, font_size in self.custom_markers:
                    if x_start <= x < x_end and y_start <= y < y_end:
                        new_x = x - x_start
                        new_y = y - y_start
                        new_custom_markers.append((new_x, new_y, text, color, font, font_size))
            self.custom_markers = new_custom_markers
            # -----------------------------------------------------

            # Update the main image and related states
            self.image = cropped_image
            # Ensure these backups reflect the *newly cropped* state
            self.image_before_padding = self.image.copy()
            self.image_contrasted = self.image.copy()
            self.image_before_contrast = self.image.copy()

        # Reset sliders after applying the crop
        self.crop_x_start_slider.setValue(0)
        self.crop_x_end_slider.setValue(100)
        self.crop_y_start_slider.setValue(0)
        self.crop_y_end_slider.setValue(100)

        # Update live view label size based on the *new* image dimensions
        try:
            if self.image: # Check if image exists after cropping
                w = self.image.width()
                h = self.image.height()
                # Preview window
                ratio = w / h if h > 0 else 1 # Avoid division by zero
                self.label_width = int(self.screen_width * 0.28)
                label_height = int(self.label_width / ratio)
                if label_height > self.label_width:
                    label_height = self.label_width
                    self.label_width = int(ratio * label_height) # Ensure integer width
                self.live_view_label.setFixedSize(int(self.label_width), int(label_height))
        except Exception as e:
            print(f"Error resizing label after crop: {e}")
            # Fallback size?
            self.live_view_label.setFixedSize(int(self.screen_width * 0.28), int(self.screen_width * 0.28))


        self.update_live_view() # Final update with corrected markers and image
        
    def update_skew(self):
        self.save_state()
        taper_value = self.taper_skew_slider.value() / 100  # Normalize taper value to a range of -1 to 1

        width = self.image.width()
        height = self.image.height()
    
        # Define corner points for perspective transformation
        source_corners = QPolygonF([QPointF(0, 0), QPointF(width, 0), QPointF(0, height), QPointF(width, height)])
    
        # Initialize destination corners as a copy of source corners
        destination_corners = QPolygonF(source_corners)
    
        # Adjust perspective based on taper value
        if taper_value > 0:
            # Narrower at the top, wider at the bottom
            destination_corners[0].setX(width * taper_value / 2)  # Top-left
            destination_corners[1].setX(width * (1 - taper_value / 2))  # Top-right
        elif taper_value < 0:
            # Wider at the top, narrower at the bottom
            destination_corners[2].setX(width * (-taper_value / 2))  # Bottom-left
            destination_corners[3].setX(width * (1 + taper_value / 2))  # Bottom-right
    
        # Create a perspective transformation using quadToQuad
        transform = QTransform()
        if not QTransform.quadToQuad(source_corners, destination_corners, transform):
            print("Failed to create transformation matrix")
            return
    
        # Apply the transformation
        # self.image = self.image.transformed(transform, Qt.SmoothTransformation)

 
        self.image = self.image.transformed(transform, Qt.SmoothTransformation)
        self.taper_skew_slider.setValue(0)

    
        
    def save_image(self):
        """Saves the original image, the modified image (rendered view), and configuration."""
        if not self.image_master: # Check if an initial image was ever loaded/pasted
             QMessageBox.warning(self, "Error", "No image data to save.")
             return False # Indicate save failed or was aborted

        self.is_modified = False # Mark as saved

        options = QFileDialog.Options()
        # Suggest a filename based on the loaded image or a default
        suggested_name = ""
        if self.image_path:
            base = os.path.splitext(os.path.basename(self.image_path))[0]
            # Remove common suffixes if they exist from previous saves
            base = base.replace("_original", "").replace("_modified", "")
            suggested_name = f"{base}_edited.png" # Suggest PNG for modified view
        elif hasattr(self, 'base_name') and self.base_name:
            suggested_name = f"{self.base_name}_edited.png"
        else:
            suggested_name = "untitled_image.png"

        save_dir = os.path.dirname(self.image_path) if self.image_path else "" # Suggest save in original dir

        # --- Get the base save path from the user ---
        base_save_path, selected_filter = QFileDialog.getSaveFileName(
            self, "Save Image Base Name", os.path.join(save_dir, suggested_name),
            "PNG Files (*.png);;TIFF Files (*.tif *.tiff);;JPEG Files (*.jpg *.jpeg);;BMP Files (*.bmp);;All Files (*)",
            options=options
        )
        # --- End getting base save path ---

        if not base_save_path: # User cancelled
            self.is_modified = True # Revert saved status if cancelled
            return False # Indicate save cancelled

        # Determine paths based on the base path
        base_name_nosuffix = os.path.splitext(base_save_path)[0]
        # Determine the desired suffix based on the selected filter or default to .png
        if "png" in selected_filter.lower():
             suffix = ".png"
        elif "tif" in selected_filter.lower():
             suffix = ".tif"
        elif "jpg" in selected_filter.lower() or "jpeg" in selected_filter.lower():
             suffix = ".jpg"
        elif "bmp" in selected_filter.lower():
             suffix = ".bmp"
        else: # Default or All Files
             suffix = os.path.splitext(base_save_path)[1] # Keep user's suffix if provided
             if not suffix: suffix = ".png" # Default to png if no suffix given

        original_save_path = f"{base_name_nosuffix}_original{suffix}"
        modified_save_path = f"{base_name_nosuffix}_modified{suffix}"
        config_save_path = f"{base_name_nosuffix}_config.txt"

        # --- Save original image (using self.image_master) ---
        img_to_save_orig = self.image_master # Use the pristine master copy
        if img_to_save_orig and not img_to_save_orig.isNull():
            save_format_orig = suffix.replace(".", "").upper() # Determine format from suffix
            if save_format_orig == "TIF": save_format_orig = "TIFF" # Use standard TIFF identifier
            elif save_format_orig == "JPEG": save_format_orig = "JPG"

            print(f"Attempting to save original ({img_to_save_orig.format()}) as {save_format_orig} to {original_save_path}")
            # QImage.save attempts to match format, quality is for lossy formats
            quality = 95 if save_format_orig in ["JPG", "JPEG"] else -1 # Default quality (-1) for lossless

            if not img_to_save_orig.save(original_save_path, format=save_format_orig if save_format_orig else None, quality=quality):
                QMessageBox.warning(self, "Error", f"Failed to save original image to {original_save_path}.")
        else:
             QMessageBox.warning(self, "Error", "Original master image data is missing.")

        # --- Save modified image (Rendered view - likely RGB) ---
        render_scale = 3
        high_res_canvas_width = self.live_view_label.width() * render_scale
        high_res_canvas_height = self.live_view_label.height() * render_scale
        # Use ARGB32 for rendering to support potential transparency (good for PNG)
        # Use RGB888 if saving as JPG/BMP which don't support alpha well.
        save_format_mod = suffix.replace(".", "").upper()
        if save_format_mod in ["JPG", "JPEG", "BMP"]:
            canvas_format = QImage.Format_RGB888
            fill_color = Qt.white # Use white background for opaque formats
        else: # PNG, TIFF
            canvas_format = QImage.Format_ARGB32_Premultiplied # Good for rendering quality with alpha
            fill_color = Qt.transparent # Use transparent background

        high_res_canvas = QImage(
            high_res_canvas_width, high_res_canvas_height, canvas_format
        )
        high_res_canvas.fill(fill_color)

        # Use current self.image for rendering the modified view
        if self.image and not self.image.isNull():
             scaled_image_mod = self.image.scaled(
                 high_res_canvas_width, high_res_canvas_height,
                 Qt.KeepAspectRatio, Qt.SmoothTransformation)

             # Render onto the high-res canvas (render_image_on_canvas draws markers etc.)
             # Pass draw_guides=False to avoid saving them
             self.render_image_on_canvas(
                 high_res_canvas, scaled_image_mod,
                 x_start=0, y_start=0, # Rendering is relative to current self.image extent
                 render_scale=render_scale, draw_guides=False
             )

             # Save the rendered canvas
             quality_mod = 95 if save_format_mod in ["JPG", "JPEG"] else -1
             if not high_res_canvas.save(modified_save_path, format=save_format_mod if save_format_mod else None, quality=quality_mod):
                 QMessageBox.warning(self, "Error", f"Failed to save modified image to {modified_save_path}.")
        else:
             QMessageBox.warning(self, "Error", "No current image to render for saving modified view.")


        # --- Save configuration ---
        config_data = self.get_current_config()
        try:
            with open(config_save_path, "w") as config_file:
                json.dump(config_data, config_file, indent=4)
        except Exception as e:
            QMessageBox.warning(self, "Error", f"Failed to save config file: {e}")
            return False # Indicate save failure


        QMessageBox.information(self, "Saved", f"Files saved successfully:\n- {os.path.basename(original_save_path)}\n- {os.path.basename(modified_save_path)}\n- {os.path.basename(config_save_path)}")

        # Update window title (optional, remove base_name if confusing)
        self.setWindowTitle(f"{self.window_title}::IMAGE SIZE:{self.image.width()}x{self.image.height()}:{base_name_nosuffix}")
        return True # Indicate successful save
            
    def save_image_svg(self):
        """Save the processed image along with markers and labels in SVG format containing EMF data."""
        if not self.image:
            QMessageBox.warning(self, "Warning", "No image to save.")
            return
        
        options = QFileDialog.Options()
        file_path, _ = QFileDialog.getSaveFileName(
            self, "Save Image as SVG for MS Word Image Editing", "", "SVG Files (*.svg)", options=options
        )
    
        if not file_path:
            return
    
        if not file_path.endswith(".svg"):
            file_path += ".svg"
    
        # Create an SVG file with svgwrite
        dwg = svgwrite.Drawing(file_path, profile='tiny', size=(self.image.width(), self.image.height()))
    
        # Convert the QImage to a base64-encoded PNG for embedding
        buffer = QBuffer()
        buffer.open(QBuffer.ReadWrite)
        self.image.save(buffer, "PNG")
        image_data = base64.b64encode(buffer.data()).decode('utf-8')
        buffer.close()
    
        # Embed the image as a base64 data URI
        dwg.add(dwg.image(href=f"data:image/png;base64,{image_data}", insert=(0, 0)))
    
    
        # Add custom markers to the SVG
        for x, y, text, color, font, font_size in getattr(self, "custom_markers", []):
            font_metrics = QFontMetrics(QFont(font, font_size))
            text_width = (font_metrics.horizontalAdvance(text))  # Get text width
            text_height = (font_metrics.height())
    
            dwg.add(
                dwg.text(
                    text,
                    insert=(x-text_width/2, y-text_height/4),
                    fill=color.name(),
                    font_family=font,
                    font_size=f"{font_size}px"
                )
            )
    
        # Add left labels
        for y, text in getattr(self, "left_markers", []):
            font_metrics = QFontMetrics(QFont(self.font_family, self.font_size))
            final_text=f"{text} ⎯ "            
            text_width = int(font_metrics.horizontalAdvance(final_text))  # Get text width
            text_height = font_metrics.height()
    
            dwg.add(
                dwg.text(
                    final_text,
                    insert=(self.left_marker_shift_added-text_width/2, y-text_height/4),
                    fill=self.font_color.name(),
                    font_family=self.font_family,
                    font_size=f"{self.font_size}px",
                    text_anchor="end"  # Aligns text to the right
                )
            )
    
        # Add right labels
        for y, text in getattr(self, "right_markers", []):
            font_metrics = QFontMetrics(QFont(self.font_family, self.font_size))
            final_text=f"{text} ⎯ "            
            text_width = int(font_metrics.horizontalAdvance(final_text))  # Get text width
            text_height = font_metrics.height()

    
            dwg.add(
                dwg.text(
                    f" ⎯ {text}",
                    insert=(self.right_marker_shift_added-text_width/2, y-text_height/4),
                    fill=self.font_color.name(),
                    font_family=self.font_family,
                    font_size=f"{self.font_size}px",
                    text_anchor="start"  # Aligns text to the left
                )
            )
    
        # Add top labels
        for x, text in getattr(self, "top_markers", []):
            font_metrics = QFontMetrics(QFont(self.font_family, self.font_size))
            final_text=f"{text}"
            text_width = int(font_metrics.horizontalAdvance(final_text)) # Get text width
            text_height = font_metrics.height()
    
            dwg.add(
                dwg.text(
                    text,
                    insert=(x-text_width/2, self.top_marker_shift_added-text_height/4),
                    fill=self.font_color.name(),
                    font_family=self.font_family,
                    font_size=f"{self.font_size}px",
                    transform=f"rotate({self.font_rotation}, {x-text_width/2}, {self.top_marker_shift_added+text_height/4})"
                )
            )
    
        # Save the SVG file
        dwg.save()
    
        QMessageBox.information(self, "Success", f"Image saved as SVG at {file_path}.")
    
    
    def copy_to_clipboard(self):
        """Copy the image from live view label to the clipboard."""
        if not self.image:
            print("No image to copy.")
            return
    
        # Define a high-resolution canvas for clipboard copy
        render_scale = 3
        high_res_canvas_width = self.live_view_label.width() * render_scale
        high_res_canvas_height = self.live_view_label.height() * render_scale
        high_res_canvas = QImage(
            high_res_canvas_width, high_res_canvas_height, QImage.Format_RGB888
        )
        high_res_canvas.fill(QColor(255, 255, 255))  # White background
    
        # Define cropping boundaries
        x_start_percent = self.crop_x_start_slider.value() / 100
        y_start_percent = self.crop_y_start_slider.value() / 100
        x_start = int(self.image.width() * x_start_percent)
        y_start = int(self.image.height() * y_start_percent)
    
        # Create a scaled version of the image
        scaled_image = self.image.scaled(
            high_res_canvas_width,
            high_res_canvas_height,
            Qt.KeepAspectRatio,
            Qt.SmoothTransformation,
        )
        self.show_grid_checkbox.setChecked(False)
        self.update_live_view()
        # Render the high-resolution canvas without guides for clipboard
        self.render_image_on_canvas(
            high_res_canvas, scaled_image, x_start, y_start, render_scale, draw_guides=False
        )
        
        # Copy the high-resolution image to the clipboard
        clipboard = QApplication.clipboard()
        clipboard.setImage(high_res_canvas)  # Copy the rendered image
        
    def copy_to_clipboard_SVG(self):
        """Create a temporary SVG file with EMF data, copy it to clipboard."""
        if not self.image:
            QMessageBox.warning(self, "Warning", "No image to save.")
            return
    
        # Get scaling factors to match the live view window
        view_width = self.live_view_label.width()
        view_height = self.live_view_label.height()
        scale_x = self.image.width() / view_width
        scale_y = self.image.height() / view_height
    
        # Create a temporary file to store the SVG
        with NamedTemporaryFile(suffix=".svg", delete=False) as temp_file:
            temp_file_path = temp_file.name
    
        # Create an SVG file with svgwrite
        dwg = svgwrite.Drawing(temp_file_path, profile='tiny', size=(self.image.width(), self.image.height()))
    
        # Convert the QImage to a base64-encoded PNG for embedding
        buffer = QBuffer()
        buffer.open(QBuffer.ReadWrite)
        self.image.save(buffer, "PNG")
        image_data = base64.b64encode(buffer.data()).decode('utf-8')
        buffer.close()
    
        # Embed the image as a base64 data URI
        dwg.add(dwg.image(href=f"data:image/png;base64,{image_data}", insert=(0, 0)))
    
        # Add custom markers to the SVG
        for x, y, text, color, font, font_size in getattr(self, "custom_markers", []):
            dwg.add(
                dwg.text(
                    text,
                    insert=(x * scale_x, y * scale_y),
                    fill=color.name(),
                    font_family=font,
                    font_size=f"{font_size}px"
                )
            )
    
        # Add left labels
        for y, text in getattr(self, "left_markers", []):
            dwg.add(
                dwg.text(
                    text,
                    insert=(self.left_marker_shift_added / scale_x, y),
                    fill=self.font_color.name(),
                    font_family=self.font_family,
                    font_size=f"{self.font_size}px"
                )
            )
    
        # Add right labels
        for y, text in getattr(self, "right_markers", []):
            dwg.add(
                dwg.text(
                    text,
                    insert=((self.image.width() / scale_x + self.right_marker_shift_added / scale_x), y),
                    fill=self.font_color.name(),
                    font_family=self.font_family,
                    font_size=f"{self.font_size}px"
                )
            )
    
        # Add top labels
        for x, text in getattr(self, "top_markers", []):
            dwg.add(
                dwg.text(
                    text,
                    insert=(x, self.top_marker_shift_added / scale_y),
                    fill=self.font_color.name(),
                    font_family=self.font_family,
                    font_size=f"{self.font_size}px",
                    transform=f"rotate({self.font_rotation}, {x}, {self.top_marker_shift_added / scale_y})"
                )
            )
    
        # Save the SVG to the temporary file
        dwg.save()
    
        # Read the SVG content
        with open(temp_file_path, "r", encoding="utf-8") as temp_file:
            svg_content = temp_file.read()
    
        # Copy SVG content to clipboard (macOS-compatible approach)
        clipboard = QApplication.clipboard()
        clipboard.setText(svg_content, mode=clipboard.Clipboard)
    
        QMessageBox.information(self, "Success", "SVG content copied to clipboard.")

        
    def clear_predict_molecular_weight(self):
        self.live_view_label.preview_marker_enabled = False
        self.live_view_label.preview_marker_text = ""
        self.live_view_label.setCursor(Qt.ArrowCursor)
        if hasattr(self, "protein_location"):
            del self.protein_location  # Clear the protein location marker
        self.predict_size=False
        self.bounding_boxes=[]
        self.bounding_box_start = None
        self.live_view_label.bounding_box_start = None
        self.live_view_label.bounding_box_preview = None
        self.quantities=[]
        self.peak_area_list=[]
        self.protein_quantities=[]
        self.standard_protein_values.setText("")
        self.standard_protein_areas=[]
        self.standard_protein_areas_text.setText("")
        self.live_view_label.quad_points=[]
        self.live_view_label.bounding_box_preview = None
        self.live_view_label.rectangle_points = []
        self.latest_calculated_quantities = []
        self.quantities_peak_area_dict={}
        
        self.update_live_view()  # Update the display
        
    def predict_molecular_weight(self):
        """
        Initiates the molecular weight prediction process.
        Determines the available markers, sorts them by position,
        and sets up the mouse click event to get the target protein location.
        """
        self.live_view_label.preview_marker_enabled = False
        self.live_view_label.preview_marker_text = ""
        self.live_view_label.setCursor(Qt.CrossCursor)
        self.run_predict_MW = False # Reset prediction flag

        # Determine which markers to use (left or right)
        markers_raw_tuples = self.left_markers if self.left_markers else self.right_markers
        if not markers_raw_tuples:
            QMessageBox.warning(self, "Error", "No markers available for prediction. Please place markers first.")
            self.live_view_label.setCursor(Qt.ArrowCursor) # Reset cursor
            return

        # --- Crucial Step: Sort markers by Y-position (migration distance) ---
        # This ensures the order reflects the actual separation on the gel/blot
        try:
            # Ensure marker values are numeric for sorting and processing
            numeric_markers = []
            for pos, val_str in markers_raw_tuples:
                try:
                    numeric_markers.append((float(pos), float(val_str)))
                except (ValueError, TypeError):
                    # Skip markers with non-numeric values for prediction
                    print(f"Warning: Skipping non-numeric marker value '{val_str}' at position {pos}.")
                    continue

            if len(numeric_markers) < 2:
                 QMessageBox.warning(self, "Error", "At least two valid numeric markers are needed for prediction.")
                 self.live_view_label.setCursor(Qt.ArrowCursor)
                 return

            sorted_markers = sorted(numeric_markers, key=lambda item: item[0])
            # Separate sorted positions and values
            sorted_marker_positions = np.array([pos for pos, val in sorted_markers])
            sorted_marker_values = np.array([val for pos, val in sorted_markers])

        except Exception as e:
            QMessageBox.critical(self, "Marker Error", f"Error processing marker data: {e}\nPlease ensure markers have valid numeric values.")
            self.live_view_label.setCursor(Qt.ArrowCursor)
            return

        # Ensure there are still at least two markers after potential filtering
        if len(sorted_marker_positions) < 2:
            QMessageBox.warning(self, "Error", "Not enough valid numeric markers remain after filtering.")
            self.live_view_label.setCursor(Qt.ArrowCursor)
            return

        # Prompt user and set up the click handler
        QMessageBox.information(self, "Instruction",
                                "Click on the target protein location in the preview window.\n"
                                "The closest standard set (Gel or WB) will be used for calculation.")
        # Pass the *sorted* full marker data to the click handler
        self.live_view_label.mousePressEvent = lambda event: self.get_protein_location(
            event, sorted_marker_positions, sorted_marker_values
        )

    def get_protein_location(self, event, all_marker_positions, all_marker_values):
        """
        Handles the mouse click event for protein selection.
        Determines the relevant standard set based on click proximity,
        performs regression on that set, predicts MW, and plots the results.
        """
        # --- 1. Get Protein Click Position (Image Coordinates) ---
        pos = event.pos()
        cursor_x, cursor_y = pos.x(), pos.y()

        # Account for zoom and pan
        if self.live_view_label.zoom_level != 1.0:
            cursor_x = (cursor_x - self.live_view_label.pan_offset.x()) / self.live_view_label.zoom_level
            cursor_y = (cursor_y - self.live_view_label.pan_offset.y()) / self.live_view_label.zoom_level

        # Transform cursor position from view coordinates to image coordinates
        displayed_width = self.live_view_label.width()
        displayed_height = self.live_view_label.height()
        image_width = self.image.width() if self.image and self.image.width() > 0 else 1
        image_height = self.image.height() if self.image and self.image.height() > 0 else 1

        scale = min(displayed_width / image_width, displayed_height / image_height)
        x_offset = (displayed_width - image_width * scale) / 2
        y_offset = (displayed_height - image_height * scale) / 2

        protein_y_image = (cursor_y - y_offset) / scale if scale != 0 else 0
        # protein_x_image = (cursor_x - x_offset) / scale if scale != 0 else 0 # Keep X if needed later

        # Store the clicked location in *view* coordinates for drawing the marker later
        self.protein_location = (cursor_x, cursor_y) # Use view coordinates for the marker

        # --- 2. Identify Potential Standard Sets (Partitioning) ---
        transition_index = -1
        # Find the first index where the molecular weight INCREASES after initially decreasing
        # (indicates a likely switch from Gel high->low MW to WB high->low MW)
        initial_decrease = False
        for k in range(1, len(all_marker_values)):
             if all_marker_values[k] < all_marker_values[k-1]:
                 initial_decrease = True # Confirm we've started migrating down
             # Check for increase *after* we've seen a decrease
             if initial_decrease and all_marker_values[k] > all_marker_values[k-1]:
                 transition_index = k
                 break # Found the likely transition

        # --- 3. Select the Active Standard Set based on Click Proximity ---
        active_marker_positions = None
        active_marker_values = None
        set_name = "Full Set" # Default name

        if transition_index != -1:
            # We have two potential sets
            set1_positions = all_marker_positions[:transition_index]
            set1_values = all_marker_values[:transition_index]
            set2_positions = all_marker_positions[transition_index:]
            set2_values = all_marker_values[transition_index:]

            # Check if both sets are valid (at least 2 points)
            valid_set1 = len(set1_positions) >= 2
            valid_set2 = len(set2_positions) >= 2

            if valid_set1 and valid_set2:
                # Calculate the mean Y position for each set
                mean_y_set1 = np.mean(set1_positions)
                mean_y_set2 = np.mean(set2_positions)

                # Assign the click to the set whose mean Y is closer
                if abs(protein_y_image - mean_y_set1) <= abs(protein_y_image - mean_y_set2):
                    active_marker_positions = set1_positions
                    active_marker_values = set1_values
                    set_name = "Set 1 (Gel?)" # Tentative name
                else:
                    active_marker_positions = set2_positions
                    active_marker_values = set2_values
                    set_name = "Set 2 (WB?)" # Tentative name
            elif valid_set1: # Only set 1 is valid
                 active_marker_positions = set1_positions
                 active_marker_values = set1_values
                 set_name = "Set 1 (Gel?)"
            elif valid_set2: # Only set 2 is valid
                 active_marker_positions = set2_positions
                 active_marker_values = set2_values
                 set_name = "Set 2 (WB?)"
            else: # Neither set is valid after splitting
                 QMessageBox.warning(self, "Error", "Could not form valid standard sets after partitioning.")
                 self.live_view_label.setCursor(Qt.ArrowCursor)
                 if hasattr(self, "protein_location"): del self.protein_location
                 return
        else:
            # Only one set detected, use all markers
            if len(all_marker_positions) >= 2:
                active_marker_positions = all_marker_positions
                active_marker_values = all_marker_values
                set_name = "Single Set"
            else: # Should have been caught earlier, but double-check
                 QMessageBox.warning(self, "Error", "Not enough markers in the single set.")
                 self.live_view_label.setCursor(Qt.ArrowCursor)
                 if hasattr(self, "protein_location"): del self.protein_location
                 return

        # --- 4. Perform Regression on the Active Set ---
        # Normalize distances *within the active set*
        min_pos_active = np.min(active_marker_positions)
        max_pos_active = np.max(active_marker_positions)
        if max_pos_active == min_pos_active: # Avoid division by zero if all points are the same
            QMessageBox.warning(self, "Error", "All markers in the selected set have the same position.")
            self.live_view_label.setCursor(Qt.ArrowCursor)
            if hasattr(self, "protein_location"): del self.protein_location
            return

        normalized_distances_active = (active_marker_positions - min_pos_active) / (max_pos_active - min_pos_active)

        # Log transform marker values
        try:
            log_marker_values_active = np.log10(active_marker_values)
        except Exception as e:
             QMessageBox.warning(self, "Error", f"Could not log-transform marker values (are they all positive?): {e}")
             self.live_view_label.setCursor(Qt.ArrowCursor)
             if hasattr(self, "protein_location"): del self.protein_location
             return

        # Perform linear regression (log-linear fit)
        coefficients = np.polyfit(normalized_distances_active, log_marker_values_active, 1)

        # Calculate R-squared for the fit on the active set
        residuals = log_marker_values_active - np.polyval(coefficients, normalized_distances_active)
        ss_res = np.sum(residuals**2)
        ss_tot = np.sum((log_marker_values_active - np.mean(log_marker_values_active))**2)
        r_squared = 1 - (ss_res / ss_tot) if ss_tot > 0 else 1 # Handle case of perfect fit or single point mean

        # --- 5. Predict MW for the Clicked Protein ---
        # Normalize the protein's Y position *using the active set's min/max*
        normalized_protein_position = (protein_y_image - min_pos_active) / (max_pos_active - min_pos_active)

        # Predict using the derived coefficients
        predicted_log10_weight = np.polyval(coefficients, normalized_protein_position)
        predicted_weight = 10 ** predicted_log10_weight

        # --- 6. Update View and Plot ---
        self.update_live_view() # Draw the '*' marker at self.protein_location

        # Plot the results, passing both active and all data for context
        self.plot_molecular_weight_graph(
            # Active set data for fitting and line plot:
            normalized_distances_active,
            active_marker_values,
            10 ** np.polyval(coefficients, normalized_distances_active), # Fitted values for active set
            # Full set data for context (plotting all points):
            all_marker_positions, # Pass original positions
            all_marker_values,
            min_pos_active,       # Pass min/max of *active* set for normalization context
            max_pos_active,
            # Prediction results:
            normalized_protein_position, # Position relative to active set scale
            predicted_weight,
            r_squared,
            set_name # Name of the set used
        )

        # Reset mouse event handler after prediction
        self.live_view_label.mousePressEvent = None
        self.live_view_label.setCursor(Qt.ArrowCursor)
        self.run_predict_MW = True # Indicate prediction was attempted/completed

        
        
    def plot_molecular_weight_graph(
        self,
        # Data for the active set (used for the fit line and highlighted points)
        active_norm_distances,  # Normalized distances for the active set points
        active_marker_values,   # Original MW values of the active set points
        active_fitted_values,   # Fitted MW values corresponding to active_norm_distances

        # Data for *all* markers (for plotting all points for context)
        all_marker_positions,   # *Original* Y positions of all markers
        all_marker_values,      # *Original* MW values of all markers
        active_min_pos,         # Min Y position of the *active* set (for normalizing all points)
        active_max_pos,         # Max Y position of the *active* set (for normalizing all points)

        # Prediction results
        predicted_norm_position,# Predicted protein position normalized relative to active set scale
        predicted_weight,       # Predicted MW value

        # Fit quality and set info
        r_squared,              # R-squared of the fit on the active set
        set_name                # Name of the set used (e.g., "Set 1", "Set 2", "Single Set")
    ):
        """
        Plots the molecular weight prediction graph and displays it in a custom dialog.
        """
        # --- Normalize *all* marker positions using the *active* set's scale ---
        if active_max_pos == active_min_pos: # Avoid division by zero
             print("Warning: Cannot normalize all points - active set min/max are equal.")
             all_norm_distances_for_plot = np.zeros_like(all_marker_positions)
        else:
            all_norm_distances_for_plot = (all_marker_positions - active_min_pos) / (active_max_pos - active_min_pos)

        # --- Create Plot ---
        # Adjust figsize and DPI for a more compact plot suitable for a dialog
        fig, ax = plt.subplots(figsize=(4.5, 3.5)) # Smaller figure size

        # 1. Plot *all* marker points lightly for context
        ax.scatter(all_norm_distances_for_plot, all_marker_values, color="grey", alpha=0.5, label="All Markers", s=25) # Slightly smaller

        # 2. Plot the *active* marker points prominently
        ax.scatter(active_norm_distances, active_marker_values, color="red", label=f"Active Set ({set_name})", s=40, marker='o') # Slightly smaller

        # 3. Plot the fitted line for the *active* set
        sort_indices = np.argsort(active_norm_distances)
        # Move R^2 out of the legend
        ax.plot(active_norm_distances[sort_indices], active_fitted_values[sort_indices],
                 color="blue", label="Fit Line", linewidth=1.5)

        # 4. Plot the predicted protein position
        # Move predicted value out of the legend
        ax.axvline(predicted_norm_position, color="green", linestyle="--",
                    label="Target Protein", linewidth=1.5)

        # --- Configure Plot ---
        ax.set_xlabel(f"Normalized Distance (Relative to {set_name})", fontsize=9)
        ax.set_ylabel("Molecular Weight (units)", fontsize=9)
        ax.set_yscale("log")
        # Smaller legend font size
        ax.legend(fontsize='x-small', loc='best') # Use 'best' location
        ax.set_title(f"Molecular Weight Prediction", fontsize=10) # Simpler title
        ax.grid(True, which='both', linestyle=':', linewidth=0.5)
        ax.tick_params(axis='both', which='major', labelsize=8) # Smaller tick labels

        plt.tight_layout(pad=0.5) # Adjust layout to prevent labels overlapping

        # --- Save Plot to Buffer ---
        pixmap = None
        try:
            buffer = BytesIO()
            # Use a moderate DPI suitable for screen display in a dialog
            plt.savefig(buffer, format="png", bbox_inches="tight", dpi=100)
            buffer.seek(0)
            pixmap = QPixmap()
            pixmap.loadFromData(buffer.read())
            buffer.close()
        except Exception as plot_err:
            print(f"Error generating plot image: {plot_err}")
            QMessageBox.warning(self, "Plot Error", "Could not generate the prediction plot.")
            # pixmap remains None
        finally:
             plt.close(fig) # Ensure figure is always closed using the fig object

        # --- Display Results in a Custom Dialog ---
        if pixmap:
            # Create a custom dialog instead of modifying QMessageBox layout
            dialog = QDialog(self)
            dialog.setWindowTitle("Prediction Result")

            # Layout for the dialog
            layout = QVBoxLayout(dialog)

            # Text label for results
            results_text = (
                f"<b>Prediction using {set_name}:</b><br>"
                f"Predicted MW: <b>{predicted_weight:.2f}</b> units<br>"
                f"Fit R²: {r_squared:.3f}"
            )
            info_label = QLabel(results_text)
            info_label.setTextFormat(Qt.RichText) # Allow basic HTML like <b>
            info_label.setWordWrap(True)
            layout.addWidget(info_label)

            # Label to display the plot
            plot_label = QLabel()
            plot_label.setPixmap(pixmap)
            plot_label.setAlignment(Qt.AlignCenter) # Center the plot
            layout.addWidget(plot_label)

            # OK button
            button_box = QDialogButtonBox(QDialogButtonBox.Ok)
            button_box.accepted.connect(dialog.accept)
            layout.addWidget(button_box)

            dialog.setLayout(layout)
            dialog.exec_() # Show the custom dialog modally
        else:
             # If plot generation failed, show a simpler message box
             QMessageBox.information(self, "Prediction Result (No Plot)",
                f"Used {set_name} for prediction.\n"
                f"Predicted MW: {predicted_weight:.2f} units\n"
                f"Fit R²: {r_squared:.3f}"
             )

  
        
    def reset_image(self):
        # Reset the image to original master (which keeps original format)
        if hasattr(self, 'image_master') and self.image_master and not self.image_master.isNull():
            self.image = self.image_master.copy()
            self.image_before_padding = None # Padding is removed
            self.image_contrasted = self.image.copy() # Reset contrast state
            self.image_before_contrast = self.image.copy()
            self.image_padded = False # Reset padding flag
            print(f"Image reset to master format: {self.image.format()}")
        else:
            # No master image loaded, clear everything
            self.image = None
            self.original_image = None
            self.image_master = None
            self.image_before_padding = None
            self.image_contrasted = None
            self.image_before_contrast = None
            self.image_padded = False
            print("Image state cleared.")

        # (Rest of reset logic for markers, UI elements, etc. remains the same)
        self.warped_image=None
        self.left_markers.clear()
        self.right_markers.clear()
        self.top_markers.clear()
        self.custom_markers.clear()
        self.clear_predict_molecular_weight() # Clears analysis state too

        self.crop_x_start_slider.setValue(0)
        self.crop_x_end_slider.setValue(100)
        self.crop_y_start_slider.setValue(0)
        self.crop_y_end_slider.setValue(100)
        self.orientation_slider.setValue(0)
        self.taper_skew_slider.setValue(0)
        self.high_slider.setValue(100)
        self.low_slider.setValue(100)
        self.gamma_slider.setValue(100)
        self.left_padding_slider.setValue(0)
        self.right_padding_slider.setValue(0)
        self.top_padding_slider.setValue(0)
        self.left_padding_input.setText("0")
        self.right_padding_input.setText("0")
        self.top_padding_input.setText("0")
        self.bottom_padding_input.setText("0")

        self.marker_mode = None
        self.current_left_marker_index = 0
        self.current_right_marker_index = 0
        self.current_top_label_index = 0
        self.left_marker_shift_added = 0
        self.right_marker_shift_added = 0
        self.top_marker_shift_added = 0
        self.live_view_label.mode = None
        self.live_view_label.quad_points = []
        self.live_view_label.bounding_box_preview = None
        self.live_view_label.setCursor(Qt.ArrowCursor)

        try:
            self.combo_box.setCurrentText("Precision Plus All Blue/Unstained")
            self.on_combobox_changed()
        except Exception as e:
            print(f"Error resetting combo box: {e}")

        if self.image and not self.image.isNull():
            try:
                w=self.image.width()
                h=self.image.height()
                ratio=w/h if h > 0 else 1
                self.label_width=int(self.screen_width * 0.28)
                label_height=int(self.label_width/ratio)
                if label_height>self.label_width:
                    label_height=self.label_width
                    self.label_width=int(ratio*label_height)
                self.live_view_label.setFixedSize(int(self.label_width), int(label_height))
                self.left_padding_input.setText(str(int(self.image.width()*0.1)))
                self.right_padding_input.setText(str(int(self.image.width()*0.1)))
                self.top_padding_input.setText(str(int(self.image.height()*0.15)))
            except Exception as e:
                print(f"Error resizing label during reset: {e}")
        else:
            self.live_view_label.clear()
            self.live_view_label.setFixedSize(int(self.screen_width * 0.28), int(self.screen_width * 0.28))

        self.update_live_view()
        print("Image and settings reset complete.")


if __name__ == "__main__":
    try:
        app = QApplication([])
        app.setStyle("Fusion")
        
        app.setStyleSheet("""
        QSlider::handle:horizontal {
            width: 100px;
            height: 100px;
            margin: -5px 0;
            background: #FFFFFF;
            border: 2px solid #555;
            border-radius: 30px;
        }
    """)
        window = CombinedSDSApp()
        window.show()
        app.exec_()
    except Exception as e:
        logging.error("Application crashed", exc_info=True)
